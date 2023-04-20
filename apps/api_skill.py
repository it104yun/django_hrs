import os
import sys
import sysconfig
import datetime
import json
import urllib.parse
import logging
import pymssql
import requests

import pyodbc
from django.apps import apps
from django.conf import settings
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.db.models import F, Q , Sum , Max
from django.db.models.deletion import Collector
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect, FileResponse
from django.shortcuts import render,redirect
from django.utils import timezone
from django.views import View
from django.views.decorators.vary import vary_on_cookie
from django.forms.models import model_to_dict
from sqlalchemy.exc import IntegrityError
import requests as req

from tablib import Dataset
from django.utils import translation


from common.utils import query_serialize

from common.models import UserDefCode,ProcessOptionsTxtDef,RegActRules,FileActionLedger
from system.models import Factory
from apps.kpi.models import EmployeeInfoEasy
from apps.skill_pdca.models import (EmployeeTitle,
                                    JobTitle,
                                    JobSkill,
                                    JobTitleSkill,
                                    MatrixMaster,
                                    MatrixDetail,
                                    PdcaMaster,
                                    PdcaDetail,
                                    PdcaDefinition,
                                    FlowDefinition,
                                    CycleDefinition,
                                    )

from apps.kpi.resources import (EmployeeInfoEasyResource,
                               # EeAttendSummaryResource,
                               excel_set_style
                                )




from django.core.mail import BadHeaderError,send_mail
import ldap

import openpyxl
from openpyxl.styles import Font, PatternFill, colors, Border, Side, Alignment
from googletrans import Translator


logger = logging.getLogger('debug')
tracer = logging.getLogger('trace')


logger = logging.getLogger('debug')
tracer = logging.getLogger('trace')


current_tz = timezone.get_current_timezone()
# now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')
now = timezone.now().astimezone(current_tz)


class ModelHandleView(View):
    @vary_on_cookie
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def parse_url(self, request):
        """
        擷取 model 參數，並找尋相對應的 Model class
        model參數會先依底線切開個單字後，做capitalize再合併。
        找不到對應 model class 回傳 404，反之回傳 Model Class 與 pk(如果有)

        ex.1
        model_parm = 'pds_main'
        model_name = PdsMain

        ex.2
        model_parm = 'PDS_main_testMain'
        model_name = PdsMainTestmain

        """
        model = self.kwargs['model']
        pk = self.kwargs.get('pk', None)
        model_name = ''.join([string.capitalize() for string in model.split('_')])
        model_list = apps.get_models()
        try:
            model_class = [m for m in model_list if m.__name__ == model_name][0]
        except IndexError as e:
             logger.debug(e)
             return JsonResponse({"success": False})
        else:
            return model_class, pk

    def get(self, request, *args, **kwargs):
        """
        GET 做兩件事: 讀取 / 刪除

        - 讀取
            <str:model>?sort=xx&order=OO&limit=100&factory=1000&other_parameters
            以 other_parameters 篩選資料之後再處理sorting, ordering, limit 和 factory

        - 刪除
            <str:model>/<str:pk>
            取得 pk 刪除指定的資料。
        """
        model, pk = self.parse_url(request)
        get_param = request.GET.dict()
        tracer.info('get %s' % model._meta.model_name)
        identity_str = request.META['QUERY_STRING']
        if not pk:
            try:
                sort = get_param.pop('sort', None)
            except:
                pass
            try:
                order = get_param.pop('order', None)
            except:
                pass
            try:
                limit = get_param.pop('limit', None)
            except:
                pass
            try:
                latest = get_param.pop('latest', None)
            except:
                pass

            factory = get_param.pop('factory', None)
            cache_str = '%s_%s' % (model._meta.model_name, identity_str)
            query = Q()
            datas = cache.get(cache_str)

            if datas:
                return JsonResponse(datas, safe=False)
            for key, value in get_param.items():
                # _id 結尾通常是 foreignkey, 不支援 __contains
                # key = key + '__contains' if not '_id' in key else key
                if ( key == 'filterRules'):   # 不執行 , 因應前端js有'enableFilter'  2021/11/30 增加
                    continue
                if ( value.find(',')==-1 ):     #queryParams單一個,單選
                    query &= Q(**{key: value})
                else:                          #多選
                    value_tuple = tuple(value.split(','))
                    query &= Q(**{'%s%s' % (key, '__in'): value_tuple})
            if factory:
                query &= Q(factory__id=factory)
            if latest:
                three_months = timezone.now() - datetime.timedelta(days=30)
                query &= Q(create_time__gt=three_months)
            data = model.objects.filter(query)
            # TODO: 優化

            if sort:
                order_syntext = '-' if order == 'desc' else ''
                data = data.order_by(order_syntext + sort)
            if limit:
                data = data[:int(limit)]

            dict_data = self._serialize(data, True)
            cache.add(cache_str, dict_data, 60)
            return JsonResponse(dict_data, safe=False)
        else:
            try:
                model.objects.filter(pk=pk).delete()
                cache.clear()
                return JsonResponse({"success": True})
            except:
                return JsonResponse({"success": False,"message":"***有『關聯』資料，不允刪除***"})


    def post(self, request, *args, **kwargs):
        # return JsonResponse({"success": True})
        """
        POST 做兩件事: 新增 / 修改

        - 新增
            <str:model>
            新增資料。

        - 修改
            <str:model>/<str:pk>
            取得 pk 更新指定的資料。
        """
        model, pk = self.parse_url(request)
        post_param = request.POST.dict()

        # 拿掉 csrfmiddlewaretoken
        if post_param.get('csrfmiddlewaretoken', None):
            post_param.pop('csrfmiddlewaretoken')

        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')
        if pk:
            instance = model.objects.filter(pk=pk)
            if instance:
                post_param.update({
                    'change_time': now,
                    'changer': request.user.username,
                })

                ''' Add by liyun at 2020/9/30
                ＊＊＊前端傳回的資料，若有未輸入會傳回〝空字串〞。＊＊＊
                空字串('') : (1)-在日期，數值，邏輯...等資料在存檔時，會出現錯誤，無法存檔。
                                使得每個欄位都〝必要輸入〞資料，這並不符合實際需要。
                            (2)-所以只要是〝空字串〞就指定為None。
                這樣的做法，是將Field是否需要輸入的控制權，交給forms.py( 假設models.py的field 的屬性 blank=True, null=True )。
                '''
                for key, value in post_param.items():
                    if value == '':
                        post_param.update({
                            key: None,
                        })
                    elif value == 'false':                             # add at 2021/09/24
                        post_param.update({
                            key: False,
                        })
                    elif value == 'true':                               # add at 2021/09/24
                        post_param.update({
                            key: True,
                        })
                try:
                    if ( str(model._meta) == 'skill_pdca.jobtitle'  ):
                        post_param.pop('job_parent_parent_id')
                        post_param.pop('job_parent_id')
                    instance.update(**post_param)
                except ValidationError as e:
                    logger.debug(e)

                    if e.code == 'invalid_time':
                        return JsonResponse({"success": False, "message": "\"%s\" 格式正確，卻非有效時間。" % e.params['value']})
                    else:
                        print(post_param)
                        return JsonResponse({"success": False, "message": "請檢查欄位資料是否正確。"})
                except ValueError as e:
                    logger.debug(e)
                    return JsonResponse({"success": False, "message": "欄位不能為空，且只能是數字。"})
                else:
                    cache.clear()
                    return JsonResponse({"success": True})
        else:
            '''
            前端傳回的資料，若有未輸入會傳回〝空字串〞，
            空字串('') : 在日期，數值，邏輯...等資料在存檔時，會出現錯誤，無法存檔。
            '''
            for key, value in post_param.items():
                if value == '':
                    post_param.update({
                        key: None,
                    })
                if value == 'false':  # add at 2021/09/24
                    post_param.update({
                        key: False,
                    })
                elif value == 'true':  # add at 2021/09/24
                    post_param.update({
                        key: True,
                    })

            post_param.update({
                'change_time': now,
                'changer': request.user.username,
                'create_time': now,
                'creator': request.user.username,
            })

            try:
                model.objects.create(**post_param)
            except ValidationError as e:
                logger.debug(e)
                if e.code == 'invalid_time':
                    return JsonResponse({"success": False, "message": "\"%s\" 格式正確，卻非有效時間。" % e.params['value']})
                else:
                    return JsonResponse({"success": False, "message": "請檢查欄位數值是否正確。"})
            except ValueError as e:
                logger.debug(e)
                return JsonResponse({"success": False, "message": "欄位不能為空，且只能是數字。"})
            except pymssql.DatabaseError as e:
                logger.debug(e)
                return JsonResponse({"success": False, "message": "已存在資料庫中"})
            except Exception as e:
                logger.debug(e)
                return JsonResponse({"success": False, "message": "已存在資料庫中"})
            else:
                cache.clear()
                return JsonResponse({"success": True})

    def _serialize(self, obj, related=True):
        return query_serialize(obj, related)



class ModelHandleDistinctView(View):

    @vary_on_cookie
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def parse_url(self, request):
        """
        擷取 model 參數，並找尋相對應的 Model class
        model參數會先依底線切開個單字後，做capitalize再合併。
        找不到對應 model class 回傳 404，反之回傳 Model Class 與 pk(如果有)

        ex.1
        model_parm = 'pds_main'
        model_name = PdsMain

        ex.2
        model_parm = 'PDS_main_testMain'
        model_name = PdsMainTestmain

        """
        model = self.kwargs['model']
        pk = self.kwargs.get('pk', None)
        model_name = ''.join([string.capitalize() for string in model.split('_')])
        model_list = apps.get_models()
        try:
            model_class = [m for m in model_list if m.__name__ == model_name][0]
        except IndexError as e:
             logger.debug(e)
             return JsonResponse({"success": False})
        else:
            return model_class, pk

    def get(self, request, *args, **kwargs):
        """
        GET 做兩件事: 讀取 / 刪除

        - 讀取
            <str:model>?sort=xx&order=OO&limit=100&factory=1000&other_parameters
            以 other_parameters 篩選資料之後再處理sorting, ordering, limit 和 factory

        - 刪除
            <str:model>/<str:pk>
            取得 pk 刪除指定的資料。
        """
        model, pk = self.parse_url(request)
        get_param = request.GET.dict()
        tracer.info('get %s' % model._meta.model_name)
        identity_str = request.META['QUERY_STRING']
        if not pk:
            try:
                sort = get_param.pop('sort', None)
            except:
                pass
            try:
                order = get_param.pop('order', None)
            except:
                pass
            try:
                limit = get_param.pop('limit', None)
            except:
                pass
            try:
                latest = get_param.pop('latest', None)
            except:
                pass

            factory = get_param.pop('factory', None)
            cache_str = '%s_%s' % (model._meta.model_name, identity_str)

            query = Q()
            datas = cache.get(cache_str)

            if datas:
                return JsonResponse(datas, safe=False)
            for key, value in get_param.items():
                # _id 結尾通常是 foreignkey, 不支援 __contains
                # key = key + '__contains' if not '_id' in key else key
                query &= Q(**{key: value})
            if factory:
                query &= Q(factory__id=factory)
            if latest:
                three_months = timezone.now() - datetime.timedelta(days=30)
                query &= Q(create_time__gt=three_months)

            data = model.objects.filter(query).distinct()
            # TODO: 優化

            if sort:
                order_syntext = '-' if order == 'desc' else ''
                data = data.order_by(order_syntext + sort)
            if limit:
                data = data[:int(limit)]

            dict_data = self._serialize(data, True)
            cache.add(cache_str, dict_data, 60)
            return JsonResponse(dict_data, safe=False)
        else:
            try:
                model.objects.filter(pk=pk).delete()
                cache.clear()
                return JsonResponse({"success": True})
            except:
                return JsonResponse({"success": False, "message": "***有『關聯』資料，不允刪除***"})

    def _serialize(self, obj, related=True):
        return query_serialize(obj, related)



def cmpx_search_handle(request, model):
    """
    多條件查詢
    0313
    """
    format = request.GET.get('format', None)
    factory = request.GET.get('factory', None)
    def relation(operator):
        mapping = {
            '>': '__gt',
            '<': '__lt',
            '=': '',
            'include': '__contains',
            '!=': '',
            'exclude': '__contains'
        }
        return mapping[operator]

    model_name = ''.join([string.capitalize() for string in model.split('_')])
    model_list = apps.get_models()
    try:
        # tracer.info('cpx, %s' % model_name)
        model_class = [m for m in model_list if m.__name__ == model_name][0]
    except IndexError as e:
        # logger.debug('Failed on cpx: %s' % e)
        return JsonResponse({"status": False})
    else:
        post_param = request.POST
        data = json.loads(post_param['data'])
        logic = data.get('logic', 'all')
        filters = data.get('filter')

        query = Q()
        if logic == 'all':
            for key, item in filters.items():
                if item['relation'] in ['exclude', '!=']:
                    query &= ~Q(**{'%s%s' % (item['field'], relation(item['relation'])): item['value']})
                else:
                    query &= Q(**{'%s%s' % (item['field'], relation(item['relation'])): item['value']})
        elif logic == 'any':
            for key, item in filters.items():
                if item['relation'] in ['exclude', '!=']:
                    query |= ~Q(**{'%s%s' % (item['field'], relation(item['relation'])): item['value']})
                else:
                    query |= Q(**{'%s%s' % (item['field'], relation(item['relation'])): item['value']})
        if format:
            query &= Q(**{'_format':format})
        if factory:
            query &= Q(**{'factory_id':factory})
        try:
            result = model_class.objects.filter(query)
        except Exception as e:
            # logger.debug('Failed on cpx: %s' % e)
            return JsonResponse({"status": False})
        return JsonResponse(query_serialize(result, True), safe=False)



#員工基本資料的匯出,2020/10/20目前未使用
def employee_export(request,pk=None):
    employee_resource = EmployeeInfoEasyResource()
    dataset = employee_resource.export()
    response = HttpResponse( dataset.xls , content_type = 'application/vnd.ms-excel' )
    response['Content-Disposition'] = 'attachment; filename="persons.csv"'
    return response


#員工基本資料的匯入
def employee_import(request):
    if request.method == 'POST':
        # EmployeeInfoEasy.objects
        employee_resource = EmployeeInfoEasyResource()
        new_employees = request.FILES['myfile']
        prepared_data = Dataset().load(new_employees.read(), format='xlsx' ,headers= True )    #headers=True...可使用string做為欄名
        err_message = "匯入資料錯誤"
        try:
            result = employee_resource.import_data( prepared_data , dry_run=True, raise_errors = True) #dry_run=True　沒有正式匯入，測試資料匯入是否有error
        except:
            # 資料有誤，顯示錯誤訊息
            cache.clear()
            return render(request,'kpi/import_error.html',context={"success": False, "message":err_message})
        else:
            if result.has_errors():
                # 匯入有誤，顯示錯誤訊息
                cache.clear()      #清除cache, 好讓〝程式〞重新取得最新匯入的資料
                return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
            else:
                result = employee_resource.import_data( prepared_data, dry_run=False)  # 實際將資料匯入
                if result.has_errors():
                    cache.clear()
                    return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
                else:
                    # 匯入成功,返回上一頁
                    cache.clear()
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER','/'),{"success": True})

def get_employee_data_factory(request,pk=None):
    results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(factory_id=pk)         # 排除'-000'共同指標, '-100'出勤指標

    dataList = []
    for T in results:
        dataList.append({
            'value': T[0],
            'text': T[0]+" "+T[1],
        })
    return JsonResponse(dataList,safe=False)



def get_director_data_factory(request,pk=None):
    commQ = Q(work_code__in = EmployeeInfoEasy.objects.values_list('director_id',flat=True).filter(factory_id=pk).order_by('director_id').distinct('director_id'))
    results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(commQ)  # 排除'-000'共同指標, '-100'出勤指標
    dataList = []
    for T in results:
        dataList.append({
            'value': T[0],
            'text': T[0]+" "+T[1],
        })
    return JsonResponse(dataList,safe=False)


def get_common_udc(request,pk=None):
    fieldList = [ 'id',
                  'desc1'
                 ]
    results = UserDefCode.objects.filter(topic_code_id=pk).values_list(*fieldList).order_by('topic_code_id','udc')
    dataList = []
    for T in results:
        dataList.append({
            'value': T[0],
            'text': str(T[0])+" "+T[1],
        })
    return JsonResponse(dataList,safe=False)


def gen_matrix_master(request,pk=None):
    err_message = ""
    if pk!=None:
        #過濾條件
        #  1.沒有離職日期
        #  2. [ today()-到職日 ] >= 3個月
        #  3. [ today()-異動日 ] >= 3個月
        commQ1 = ~Q(work_code__endswith='-000')          #排除共同指標
        commQ2 = Q(factory_id=pk)
        commQ3 = Q(resign_date__isnull=True)
        three_month = ( timezone.now() - datetime.timedelta(days=90) ).date()
        commQ4 = Q(arrival_date__lte=three_month)        # 非空,到職日在三個月以前(小於等於)
        commQ5 = Q(trans_date__isnull=True)              # 非空,異動日在三個月以前(小於等於)
        commQ6 = Q(trans_date__lte=three_month)          # 非空,異動日在三個月以前(小於等於)
        if pk=='ZZ01':          #總公司( 所有人, 不過濾公司)
            commQ9 = Q(work_code__in = EmployeeInfoEasy.objects.values_list('work_code').filter(commQ1 & commQ3 & commQ4 & (commQ5|commQ6) ) )
        else:
            commQ9 = Q(work_code__in = EmployeeInfoEasy.objects.values_list('work_code').filter(commQ1 & commQ2 & commQ3 & commQ4 & (commQ5|commQ6) ))
        commQx = Q(enable=True)       # 啟用
        results = EmployeeTitle.objects.filter(commQ9,commQx)
    today = datetime.date.today()
    year = today.year
    month = today.month

    masterList =[]
    for instance in results:
        masterList.append({
            'work_code_title': instance,
            'year' : year,
            'month' : month,
        })

    bulk_data = [MatrixMaster(
        change_time=now,
        changer=request.user.username,
        create_time=now,
        creator=request.user.username,
        work_code_title=master.get('work_code_title'),
        year=master.get('year'),
        month=master.get('month'),
    ) for master in masterList]

    try:
        MatrixMaster.objects.bulk_create(bulk_data)
        return JsonResponse({"success": True})
    except ValidationError as e:
        return JsonResponse({"success": False})


def matrix_score(request,pk1=None,pk2=None):     #pk1 : job_title  pk2 : userID
    select_dept = request.session.get('select_dept', None)
    view_name = (request.path).split('/')[2]         #['', 'skill_pdca', 'tt407_2pdf']

    fields = [
        'id','year','month','bpm',

        # 向上連結 : EmployeeTitle wrok_code_title__work_code   欄位名+外連接檔案的欄位名
        'work_code_title__work_code','work_code_title__job_title','work_code_title__enable',

        #向上連結 : MatrixStatus
        'bpm__bpm_number','bpm__bpm_status_desc1',

        # 向下連結 : matrix_detail__cr001  連接至本檔案的"releated_name"+連接至本檔案的欄位名
        'matrix_detail__cr001','matrix_detail__cr002','matrix_detail__cr003','matrix_detail__cr004','matrix_detail__cr005','matrix_detail__cr006','matrix_detail__cr007','matrix_detail__cr008',
        'matrix_detail__cr009','matrix_detail__cr010','matrix_detail__cr011','matrix_detail__cr012','matrix_detail__cr013','matrix_detail__cr014','matrix_detail__cr015','matrix_detail__cr016',
        'matrix_detail__cr017','matrix_detail__cr018','matrix_detail__cr019','matrix_detail__cr020',
        'matrix_detail__ge001','matrix_detail__ge002','matrix_detail__ge003','matrix_detail__ge004','matrix_detail__ge005','matrix_detail__ge006','matrix_detail__ge007','matrix_detail__ge008',
        'matrix_detail__ge009','matrix_detail__ge010','matrix_detail__ge011','matrix_detail__ge012','matrix_detail__ge013','matrix_detail__ge014','matrix_detail__ge015','matrix_detail__ge016',
        'matrix_detail__ge017','matrix_detail__ge018','matrix_detail__ge019','matrix_detail__ge020',
        'matrix_detail__ma001','matrix_detail__ma002','matrix_detail__ma003','matrix_detail__ma004','matrix_detail__ma005','matrix_detail__ma006','matrix_detail__ma007','matrix_detail__ma008',
        'matrix_detail__ma009','matrix_detail__ma010','matrix_detail__ma011','matrix_detail__ma012','matrix_detail__ma013','matrix_detail__ma014','matrix_detail__ma015','matrix_detail__ma016',
        'matrix_detail__ma017','matrix_detail__ma018','matrix_detail__ma019','matrix_detail__ma020','matrix_detail__ma021','matrix_detail__ma022','matrix_detail__ma023','matrix_detail__ma024',
        'matrix_detail__ma025','matrix_detail__ma026','matrix_detail__ma027','matrix_detail__ma028','matrix_detail__ma029','matrix_detail__ma030','matrix_detail__ma031','matrix_detail__ma032',
        'matrix_detail__ma033','matrix_detail__ma034','matrix_detail__ma035','matrix_detail__ma036','matrix_detail__ma037','matrix_detail__ma038','matrix_detail__ma039','matrix_detail__ma040',
        'matrix_detail__ma041','matrix_detail__ma042','matrix_detail__ma043','matrix_detail__ma044','matrix_detail__ma045','matrix_detail__ma046','matrix_detail__ma047','matrix_detail__ma048',
        'matrix_detail__ma049','matrix_detail__ma050','matrix_detail__ma051','matrix_detail__ma052','matrix_detail__ma053','matrix_detail__ma054','matrix_detail__ma055','matrix_detail__ma056',
        'matrix_detail__ma057','matrix_detail__ma058','matrix_detail__ma059','matrix_detail__ma060',
        'matrix_detail__pr001','matrix_detail__pr002','matrix_detail__pr003','matrix_detail__pr004','matrix_detail__pr005','matrix_detail__pr006','matrix_detail__pr007','matrix_detail__pr008',
        'matrix_detail__pr009','matrix_detail__pr010','matrix_detail__pr011','matrix_detail__pr012','matrix_detail__pr013','matrix_detail__pr014','matrix_detail__pr015','matrix_detail__pr016',
        'matrix_detail__pr017','matrix_detail__pr018','matrix_detail__pr019','matrix_detail__pr020','matrix_detail__pr021','matrix_detail__pr022','matrix_detail__pr023','matrix_detail__pr024',
        'matrix_detail__pr025','matrix_detail__pr026','matrix_detail__pr027','matrix_detail__pr028','matrix_detail__pr029','matrix_detail__pr030','matrix_detail__pr031','matrix_detail__pr032',
        'matrix_detail__pr033','matrix_detail__pr034','matrix_detail__pr035','matrix_detail__pr036','matrix_detail__pr037','matrix_detail__pr038','matrix_detail__pr039','matrix_detail__pr040',
        'matrix_detail__pr041','matrix_detail__pr042','matrix_detail__pr043','matrix_detail__pr044','matrix_detail__pr045','matrix_detail__pr046','matrix_detail__pr047','matrix_detail__pr048',
        'matrix_detail__pr049','matrix_detail__pr050','matrix_detail__pr051','matrix_detail__pr052','matrix_detail__pr053','matrix_detail__pr054','matrix_detail__pr055','matrix_detail__pr056',
        'matrix_detail__pr057','matrix_detail__pr058','matrix_detail__pr059','matrix_detail__pr060','matrix_detail__pr061','matrix_detail__pr062','matrix_detail__pr063','matrix_detail__pr064',
        'matrix_detail__pr065','matrix_detail__pr066','matrix_detail__pr067','matrix_detail__pr068','matrix_detail__pr069','matrix_detail__pr070','matrix_detail__pr071','matrix_detail__pr072',
        'matrix_detail__pr073','matrix_detail__pr074','matrix_detail__pr075','matrix_detail__pr076','matrix_detail__pr077','matrix_detail__pr078','matrix_detail__pr079','matrix_detail__pr080',
        'matrix_detail__pr081','matrix_detail__pr082','matrix_detail__pr083','matrix_detail__pr084','matrix_detail__pr085','matrix_detail__pr086','matrix_detail__pr087','matrix_detail__pr088',
        'matrix_detail__pr089','matrix_detail__pr090','matrix_detail__pr091','matrix_detail__pr092','matrix_detail__pr093','matrix_detail__pr094','matrix_detail__pr095','matrix_detail__pr096',
        'matrix_detail__pr097','matrix_detail__pr098','matrix_detail__pr099','matrix_detail__pr100','matrix_detail__pr101','matrix_detail__pr102','matrix_detail__pr103','matrix_detail__pr104',
        'matrix_detail__pr105','matrix_detail__pr106','matrix_detail__pr107','matrix_detail__pr108','matrix_detail__pr109','matrix_detail__pr110','matrix_detail__pr111','matrix_detail__pr112',
        'matrix_detail__pr113','matrix_detail__pr114','matrix_detail__pr115','matrix_detail__pr116','matrix_detail__pr117','matrix_detail__pr118','matrix_detail__pr119','matrix_detail__pr120',
        'matrix_detail__xa001','matrix_detail__xa002','matrix_detail__xb001','matrix_detail__xb002','matrix_detail__xb003',
    ]

    qr0 = Q(bpm__isnull=True)
    qr1 = Q(bpm__bpm_status_desc1__exact='')  # 未送BPM   留待js控制按鈕
    qr2 = Q(bpm__bpm_status_desc1='reject')  # BPM退回   留待js控制按鈕
    qr3 = Q(work_code_title__job_title=pk1)
    qr4 = Q(work_code_title__work_code__in=EmployeeInfoEasy.objects.filter(direct_supv=pk2,dept_flevel_id=select_dept)) #找到『直接主管』是userID的工號<--這個在view裏抓
    results = None
    # if (view_name=='tt407_2pdf'):
    #     # session_key = "query"+pk1+"_"+str(select_dept)
    #     session_key = pk2+str(select_dept)+pk1
    #     query_list = request.session.get(session_key)
    #     query_x = Q()
    #     if query_list:
    #         for ql in query_list:
    #             query_x = query_x | Q ( year = ql.get('year'), month = ql.get('month') , work_code_title__job_title= ql.get('job_title'),work_code_title__work_code=ql.get('work_code') )
    #         results = MatrixMaster.objects.filter( query_x ).values(*fields)
    # else:
    #     results = MatrixMaster.objects.filter( (qr0 | qr1 | qr2), qr3 , qr4).values(*fields)
    results = MatrixMaster.objects.filter( (qr0 | qr1 | qr2), qr3 , qr4).values(*fields)
    # results = MatrixMaster.objects.filter( (qr1 | qr2), qr3 , qr4).values(*fields)

    dataList = []
    if results:
        for data in results:
            fieldsList = [
                'chi_name',
                'dept',
                'arrival_date',
                'resign_date',
                'trans_date',
                'trans_type',
                'rank',
                'pos',
                'factory',
            ]


            ee = model_to_dict( EmployeeInfoEasy.objects.get(work_code=data.get('work_code_title__work_code')) , fields=fieldsList )

            # js easyui datagrid 配合column同名, 才能assign給grid------------------------------------------------------------begin
            dataList.append({
                'id' : data.get('id',''),
                'work_code': data.get('work_code_title__work_code',''),
                'chi_name': ee.get('chi_name',''),
                'year': data.get('year',''),
                'month': data.get('month',''),
                'job_title': data.get('work_code_title__job_title',''),
                'job_title_desc': JobTitle.objects.get(job_code=data.get('work_code_title__job_title')).job_name,
                # 'bpm_status': data.get('work_code_title__bpm_status'),
                'dept_desc': "" if ee.get('dept') == None else UserDefCode.objects.get(id=ee.get('dept')).desc1,
                'arrival_date': ee.get('arrival_date'),
                'trans_date': ee.get('trans_date'),
                'trans_type': ee.get('trans_type'),
                'rank_desc': "" if ee.get('rank') == None else UserDefCode.objects.get(id=ee.get('rank')).desc1,
                'pos_desc': "" if ee.get('pos') == None else UserDefCode.objects.get(id=ee.get('pos')).desc1,
                'pos_shc1': "" if ee.get('pos') == None else UserDefCode.objects.get(id=ee.get('pos')).shc1,    #是否為『主管職』,管理職能盤點的依據
                'pos_shc2': "" if ee.get('pos') == None else UserDefCode.objects.get(id=ee.get('pos')).shc2,    #主管職者, 營業報告書是否要盤點
                'factory_desc': "" if ee.get('factory') == None else Factory.objects.get(id=ee.get('factory')).name,
                'bpm_number':data.get('bpm__bpm_number',''),
                # 'bpm_status':data.get('bpm__bpm_status',''),
                'bpm_status_desc1':data.get('bpm__bpm_status_desc1',''),
                # 'bpm_status_desc2':data.get('bpm__bpm_status_desc2',''),

                'cr001': data.get('matrix_detail__cr001',''),
                'cr002': data.get('matrix_detail__cr002',''),
                'cr003': data.get('matrix_detail__cr003',''),
                'cr004': data.get('matrix_detail__cr004',''),
                'cr005': data.get('matrix_detail__cr005',''),
                'cr006': data.get('matrix_detail__cr006',''),
                'cr007': data.get('matrix_detail__cr007',''),
                'cr008': data.get('matrix_detail__cr008',''),
                'cr009': data.get('matrix_detail__cr009',''),
                'cr010': data.get('matrix_detail__cr010',''),
                'cr011': data.get('matrix_detail__cr011',''),
                'cr012': data.get('matrix_detail__cr012',''),
                'cr013': data.get('matrix_detail__cr013',''),
                'cr014': data.get('matrix_detail__cr014',''),
                'cr015': data.get('matrix_detail__cr015',''),
                'cr016': data.get('matrix_detail__cr016',''),
                'cr017': data.get('matrix_detail__cr017',''),
                'cr018': data.get('matrix_detail__cr018',''),
                'cr019': data.get('matrix_detail__cr019',''),
                'cr020': data.get('matrix_detail__cr020',''),

                'ma001': data.get('matrix_detail__ma001', ''),
                'ma002': data.get('matrix_detail__ma002', ''),
                'ma003': data.get('matrix_detail__ma003', ''),
                'ma004': data.get('matrix_detail__ma004', ''),
                'ma005': data.get('matrix_detail__ma005', ''),
                'ma006': data.get('matrix_detail__ma006', ''),
                'ma007': data.get('matrix_detail__ma007', ''),
                'ma008': data.get('matrix_detail__ma008', ''),
                'ma009': data.get('matrix_detail__ma009', ''),
                'ma010': data.get('matrix_detail__ma010', ''),
                'ma011': data.get('matrix_detail__ma011', ''),
                'ma012': data.get('matrix_detail__ma012', ''),
                'ma013': data.get('matrix_detail__ma013', ''),
                'ma014': data.get('matrix_detail__ma014', ''),
                'ma015': data.get('matrix_detail__ma015', ''),
                'ma016': data.get('matrix_detail__ma016', ''),
                'ma017': data.get('matrix_detail__ma017', ''),
                'ma018': data.get('matrix_detail__ma018', ''),
                'ma019': data.get('matrix_detail__ma019', ''),
                'ma020': data.get('matrix_detail__ma020', ''),
                'ma021': data.get('matrix_detail__ma021', ''),
                'ma022': data.get('matrix_detail__ma022', ''),
                'ma023': data.get('matrix_detail__ma023', ''),
                'ma024': data.get('matrix_detail__ma024', ''),
                'ma025': data.get('matrix_detail__ma025', ''),
                'ma026': data.get('matrix_detail__ma026', ''),
                'ma027': data.get('matrix_detail__ma027', ''),
                'ma028': data.get('matrix_detail__ma028', ''),
                'ma029': data.get('matrix_detail__ma029', ''),
                'ma030': data.get('matrix_detail__ma030', ''),
                'ma031': data.get('matrix_detail__ma031', ''),
                'ma032': data.get('matrix_detail__ma032', ''),
                'ma033': data.get('matrix_detail__ma033', ''),
                'ma034': data.get('matrix_detail__ma034', ''),
                'ma035': data.get('matrix_detail__ma035', ''),
                'ma036': data.get('matrix_detail__ma036', ''),
                'ma037': data.get('matrix_detail__ma037', ''),
                'ma038': data.get('matrix_detail__ma038', ''),
                'ma039': data.get('matrix_detail__ma039', ''),
                'ma040': data.get('matrix_detail__ma040', ''),
                'ma041': data.get('matrix_detail__ma041', ''),
                'ma042': data.get('matrix_detail__ma042', ''),
                'ma043': data.get('matrix_detail__ma043', ''),
                'ma044': data.get('matrix_detail__ma044', ''),
                'ma045': data.get('matrix_detail__ma045', ''),
                'ma046': data.get('matrix_detail__ma046', ''),
                'ma047': data.get('matrix_detail__ma047', ''),
                'ma048': data.get('matrix_detail__ma048', ''),
                'ma049': data.get('matrix_detail__ma049', ''),
                'ma050': data.get('matrix_detail__ma050', ''),
                'ma051': data.get('matrix_detail__ma051', ''),
                'ma052': data.get('matrix_detail__ma052', ''),
                'ma053': data.get('matrix_detail__ma053', ''),
                'ma054': data.get('matrix_detail__ma054', ''),
                'ma055': data.get('matrix_detail__ma055', ''),
                'ma056': data.get('matrix_detail__ma056', ''),
                'ma057': data.get('matrix_detail__ma057', ''),
                'ma058': data.get('matrix_detail__ma058', ''),
                'ma059': data.get('matrix_detail__ma059', ''),
                'ma060': data.get('matrix_detail__ma060', ''),

                'ge001': data.get('matrix_detail__ge001', ''),
                'ge002': data.get('matrix_detail__ge002', ''),
                'ge003': data.get('matrix_detail__ge003', ''),
                'ge004': data.get('matrix_detail__ge004', ''),
                'ge005': data.get('matrix_detail__ge005', ''),
                'ge006': data.get('matrix_detail__ge006', ''),
                'ge007': data.get('matrix_detail__ge007', ''),
                'ge008': data.get('matrix_detail__ge008', ''),
                'ge009': data.get('matrix_detail__ge009', ''),
                'ge010': data.get('matrix_detail__ge010', ''),
                'ge011': data.get('matrix_detail__ge011', ''),
                'ge012': data.get('matrix_detail__ge012', ''),
                'ge013': data.get('matrix_detail__ge013', ''),
                'ge014': data.get('matrix_detail__ge014', ''),
                'ge015': data.get('matrix_detail__ge015', ''),
                'ge016': data.get('matrix_detail__ge016', ''),
                'ge017': data.get('matrix_detail__ge017', ''),
                'ge018': data.get('matrix_detail__ge018', ''),
                'ge019': data.get('matrix_detail__ge019', ''),
                'ge020': data.get('matrix_detail__ge020', ''),

                'pr001': data.get('matrix_detail__pr001', ''),
                'pr002': data.get('matrix_detail__pr002', ''),
                'pr003': data.get('matrix_detail__pr003', ''),
                'pr004': data.get('matrix_detail__pr004', ''),
                'pr005': data.get('matrix_detail__pr005', ''),
                'pr006': data.get('matrix_detail__pr006', ''),
                'pr007': data.get('matrix_detail__pr007', ''),
                'pr008': data.get('matrix_detail__pr008', ''),
                'pr009': data.get('matrix_detail__pr009', ''),
                'pr010': data.get('matrix_detail__pr010', ''),
                'pr011': data.get('matrix_detail__pr011', ''),
                'pr012': data.get('matrix_detail__pr012', ''),
                'pr013': data.get('matrix_detail__pr013', ''),
                'pr014': data.get('matrix_detail__pr014', ''),
                'pr015': data.get('matrix_detail__pr015', ''),
                'pr016': data.get('matrix_detail__pr016', ''),
                'pr017': data.get('matrix_detail__pr017', ''),
                'pr018': data.get('matrix_detail__pr018', ''),
                'pr019': data.get('matrix_detail__pr019', ''),
                'pr020': data.get('matrix_detail__pr020', ''),
                'pr021': data.get('matrix_detail__pr021', ''),
                'pr022': data.get('matrix_detail__pr022', ''),
                'pr023': data.get('matrix_detail__pr023', ''),
                'pr024': data.get('matrix_detail__pr024', ''),
                'pr025': data.get('matrix_detail__pr025', ''),
                'pr026': data.get('matrix_detail__pr026', ''),
                'pr027': data.get('matrix_detail__pr027', ''),
                'pr028': data.get('matrix_detail__pr028', ''),
                'pr029': data.get('matrix_detail__pr029', ''),
                'pr030': data.get('matrix_detail__pr030', ''),
                'pr031': data.get('matrix_detail__pr031', ''),
                'pr032': data.get('matrix_detail__pr032', ''),
                'pr033': data.get('matrix_detail__pr033', ''),
                'pr034': data.get('matrix_detail__pr034', ''),
                'pr035': data.get('matrix_detail__pr035', ''),
                'pr036': data.get('matrix_detail__pr036', ''),
                'pr037': data.get('matrix_detail__pr037', ''),
                'pr038': data.get('matrix_detail__pr038', ''),
                'pr039': data.get('matrix_detail__pr039', ''),
                'pr040': data.get('matrix_detail__pr040', ''),
                'pr041': data.get('matrix_detail__pr041', ''),
                'pr042': data.get('matrix_detail__pr042', ''),
                'pr043': data.get('matrix_detail__pr043', ''),
                'pr044': data.get('matrix_detail__pr044', ''),
                'pr045': data.get('matrix_detail__pr045', ''),
                'pr046': data.get('matrix_detail__pr046', ''),
                'pr047': data.get('matrix_detail__pr047', ''),
                'pr048': data.get('matrix_detail__pr048', ''),
                'pr049': data.get('matrix_detail__pr049', ''),
                'pr050': data.get('matrix_detail__pr050', ''),
                'pr051': data.get('matrix_detail__pr051', ''),
                'pr052': data.get('matrix_detail__pr052', ''),
                'pr053': data.get('matrix_detail__pr053', ''),
                'pr054': data.get('matrix_detail__pr054', ''),
                'pr055': data.get('matrix_detail__pr055', ''),
                'pr056': data.get('matrix_detail__pr056', ''),
                'pr057': data.get('matrix_detail__pr057', ''),
                'pr058': data.get('matrix_detail__pr058', ''),
                'pr059': data.get('matrix_detail__pr059', ''),
                'pr060': data.get('matrix_detail__pr060', ''),
                'pr061': data.get('matrix_detail__pr061', ''),
                'pr062': data.get('matrix_detail__pr062', ''),
                'pr063': data.get('matrix_detail__pr063', ''),
                'pr064': data.get('matrix_detail__pr064', ''),
                'pr065': data.get('matrix_detail__pr065', ''),
                'pr066': data.get('matrix_detail__pr066', ''),
                'pr067': data.get('matrix_detail__pr067', ''),
                'pr068': data.get('matrix_detail__pr068', ''),
                'pr069': data.get('matrix_detail__pr069', ''),
                'pr070': data.get('matrix_detail__pr070', ''),
                'pr071': data.get('matrix_detail__pr071', ''),
                'pr072': data.get('matrix_detail__pr072', ''),
                'pr073': data.get('matrix_detail__pr073', ''),
                'pr074': data.get('matrix_detail__pr074', ''),
                'pr075': data.get('matrix_detail__pr075', ''),
                'pr076': data.get('matrix_detail__pr076', ''),
                'pr077': data.get('matrix_detail__pr077', ''),
                'pr078': data.get('matrix_detail__pr078', ''),
                'pr079': data.get('matrix_detail__pr079', ''),
                'pr080': data.get('matrix_detail__pr080', ''),
                'pr081': data.get('matrix_detail__pr081', ''),
                'pr082': data.get('matrix_detail__pr082', ''),
                'pr083': data.get('matrix_detail__pr083', ''),
                'pr084': data.get('matrix_detail__pr084', ''),
                'pr085': data.get('matrix_detail__pr085', ''),
                'pr086': data.get('matrix_detail__pr086', ''),
                'pr087': data.get('matrix_detail__pr087', ''),
                'pr088': data.get('matrix_detail__pr088', ''),
                'pr089': data.get('matrix_detail__pr089', ''),
                'pr090': data.get('matrix_detail__pr090', ''),
                'pr091': data.get('matrix_detail__pr091', ''),
                'pr092': data.get('matrix_detail__pr092', ''),
                'pr093': data.get('matrix_detail__pr093', ''),
                'pr094': data.get('matrix_detail__pr094', ''),
                'pr095': data.get('matrix_detail__pr095', ''),
                'pr096': data.get('matrix_detail__pr096', ''),
                'pr097': data.get('matrix_detail__pr097', ''),
                'pr098': data.get('matrix_detail__pr098', ''),
                'pr099': data.get('matrix_detail__pr099', ''),
                'pr100': data.get('matrix_detail__pr100', ''),
                'pr101': data.get('matrix_detail__pr101', ''),
                'pr102': data.get('matrix_detail__pr102', ''),
                'pr103': data.get('matrix_detail__pr103', ''),
                'pr104': data.get('matrix_detail__pr104', ''),
                'pr105': data.get('matrix_detail__pr105', ''),
                'pr106': data.get('matrix_detail__pr106', ''),
                'pr107': data.get('matrix_detail__pr107', ''),
                'pr108': data.get('matrix_detail__pr108', ''),
                'pr109': data.get('matrix_detail__pr109', ''),
                'pr110': data.get('matrix_detail__pr110', ''),
                'pr111': data.get('matrix_detail__pr111', ''),
                'pr112': data.get('matrix_detail__pr112', ''),
                'pr113': data.get('matrix_detail__pr113', ''),
                'pr114': data.get('matrix_detail__pr114', ''),
                'pr115': data.get('matrix_detail__pr115', ''),
                'pr116': data.get('matrix_detail__pr116', ''),
                'pr117': data.get('matrix_detail__pr117', ''),
                'pr118': data.get('matrix_detail__pr118', ''),
                'pr119': data.get('matrix_detail__pr119', ''),
                'pr120': data.get('matrix_detail__pr120', ''),

                'xa001': data.get('matrix_detail__xa001', ''),
                'xa002': data.get('matrix_detail__xa002', ''),

                'xb001': data.get('matrix_detail__xb001', ''),
                'xb002': data.get('matrix_detail__xb002', ''),
                'xb003': data.get('matrix_detail__xb003', ''),
            })


    if (view_name=='tt407_2pdf'):
        # #取出session後要立即清除, 否則下次列印, 又會再跑出來
        # if session_key in request.session:
        #     request.session.pop(session_key)
        # print()
        # print("~"*200)
        # print(" "*16+"技能成熟度:")
        # for ll in dataList:
        #     print(" "*20,ll)
        # print("-"*200)
        # print()
        return dataList
    else:
        # query_list = []
        # for data in results:
        #     # 存起來，給下一個也呼叫此app的view=tt407_2pdf 使用
        #     query_list.append({
        #         'year': data.get('year'),
        #         'month': data.get('month'),
        #         'job_title': data.get('work_code_title__job_title'),
        #         'work_code': data.get('work_code_title__work_code'),
        #     })
        # 存起來，給 view=tt407_2pdf query 使用
        # 取出session後要立即清除, 否則下次列印, 又會再跑出來
        # session_key = pk2+str(select_dept)+pk1   #pk1 : job_title
        # request.session[session_key] = query_list
        return JsonResponse(dataList,safe=False)


class update_tabs_datagrid_rows(View):
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        post_param = list( request.POST.dict() )    #整個所有物件內容, 是一個key
        data = json.loads(post_param[0])            #取得dict第一個key值

        update_results = []
        for key,val in data.items():
            dict = {key: value for i, (key, value) in enumerate(val.items()) if key[0:2] in ('id','cr','ma','ge','pr','xa','xb')}
            dict['master_id'] = dict.pop("id")
            dict.update({
                'create_time': now,
                'change_time': now,
                'changer': request.user.username,
                'creator': request.user.username,
            })
            try:
                MatrixDetail.objects.create(**dict)
            except:
                try:
                    instance = MatrixDetail.objects.filter(master_id=dict.get('master_id'))
                except:
                    print("資料錯誤, 無法建立, 也無法更新")
                    update_results.append({
                        "job_title": val.get('job_title'),
                        "work_code":val.get('work_code'),
                        "chi_name":val.get('chi_name'),
                        "year":val.get('year'),
                        "month":val.get('month'),
                        "success":False,
                    })
                else:
                    instance.update(**dict)
                    update_results.append({
                        "job_title": val.get('job_title'),
                        "work_code": val.get('work_code'),
                        "chi_name": val.get('chi_name'),
                        "year": val.get('year'),
                        "month": val.get('month'),
                        "success":True,
                    })
            else:
                update_results.append({
                    "job_title": val.get('job_title'),
                    "work_code": val.get('work_code'),
                    "chi_name": val.get('chi_name'),
                    "year": val.get('year'),
                    "month": val.get('month'),
                    "success": True,
                })
        return JsonResponse({"results": update_results})


class export_matrix_detail(View):

    def get_common_skills(self):
        dataList1 = []
        dataList2 = []
        cr = JobSkill.objects.values_list('skill_name', flat=True).filter(
            skill_class=UserDefCode.objects.get(udc='S001'))  # 核心職能
        for idx, itm in enumerate(cr):
            field_name = 'cr' + ('00' if idx < 9 else '0' if idx>8 and idx < 99 else '') + str(idx + 1)
            dataList1.append(field_name)
            dataList2.append(itm)

        ma = JobSkill.objects.values_list('skill_name','chk_yn').filter(
            skill_class=UserDefCode.objects.get(udc='S002'))  # 管理職能
        for idx,itm in enumerate(ma):
            field_name = 'ma' + ('00' if idx < 9 else '0' if idx>8 and idx < 99 else '') + str(idx + 1)
            dataList1.append(field_name)
            dataList2.append(itm[0])


        ge = JobSkill.objects.values_list('skill_name', flat=True).filter(
            skill_class=UserDefCode.objects.get(udc='S003'))  # 一般職能
        for idx, itm in enumerate(ge):
            field_name = 'ge' + ('00' if idx < 9 else '0' if idx>8 and idx < 99 else '') + str(idx + 1)
            dataList1.append(field_name)
            dataList2.append(itm)

        return dataList1,dataList2

    def get_pro_skills(self,job_title):
        dataList1 = []
        dataList2 = []
        pr = JobSkill.objects.values_list('skill_name', flat=True).filter(
            skill_code__in=JobTitleSkill.objects.values_list('job_skill').filter(job_title=job_title))  #專業職能，只有專業職能，會依不同職務而不同
        for idx, itm in enumerate(pr):  # 結果:[{}]<---List's element have namy dict-->to js-->Array have many element
            field_name = 'pr' + ('00' if idx < 9 else '0' if idx>8 and idx < 99 else '') + str(idx + 1)
            dataList1.append(field_name)
            dataList2.append(itm)

        return dataList1,dataList2



    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        post_param = list( request.POST.dict() )    #整個所有物件內容, 是一個key
        sele = json.loads(post_param[0])            #取得dict第一個key值
        factory = sele.get('factory_select')
        date_s = sele.get('month_start').split('-')
        date_e = sele.get('month_end').split('-')
        jobtitle_list = sele.get('jobtitle_select')
        jobtitle_list.sort()

        fields = ['id','year','month','dept_desc1','work_code','chi_name','arrival_date','rank_desc1','pos_desc1','job_name','job_title',

                  'cr001','cr002','cr003','cr004','cr005','cr006','cr007','cr008','cr009','cr010','cr011','cr012','cr013','cr014','cr015','cr016','cr017','cr018','cr019','cr020',

                  'ma001','ma002','ma003','ma004','ma005','ma006','ma007','ma008','ma009','ma010','ma011','ma012','ma013','ma014','ma015','ma016','ma017','ma018','ma019','ma020',
                  'ma021','ma022','ma023','ma024','ma025','ma026','ma027','ma028','ma029','ma030','ma031','ma032','ma033','ma034','ma035','ma036','ma037','ma038','ma039','ma040',
                  'ma041','ma042','ma043','ma044','ma045','ma046','ma047','ma048','ma049','ma050','ma051','ma052','ma053','ma054','ma055','ma056','ma057','ma058','ma059','ma060',

                  'ge001', 'ge002', 'ge003', 'ge004', 'ge005', 'ge006', 'ge007', 'ge008', 'ge009', 'ge010', 'ge011',
                  'ge012', 'ge013', 'ge014', 'ge015', 'ge016', 'ge017', 'ge018', 'ge019', 'ge020',

                  'pr001','pr002','pr003','pr004','pr005','pr006','pr007','pr008','pr009','pr010','pr011','pr012','pr013','pr014','pr015','pr016','pr017','pr018','pr019','pr020',
                  'pr021','pr022','pr023','pr024','pr025','pr026','pr027','pr028','pr029','pr030','pr031','pr032','pr033','pr034','pr035','pr036','pr037','pr038','pr039','pr040',
                  'pr041','pr042','pr043','pr044','pr045','pr046','pr047','pr048','pr049','pr050','pr051','pr052','pr053','pr054','pr055','pr056','pr057','pr058','pr059','pr060',
                  'pr061','pr062','pr063','pr064','pr065','pr066','pr067','pr068','pr069','pr070','pr071','pr072','pr073','pr074','pr075','pr076','pr077','pr078','pr079','pr080',
                  'pr081','pr082','pr083','pr084','pr085','pr086','pr087','pr088','pr089','pr090','pr091','pr092','pr093','pr094','pr095','pr096','pr097','pr098','pr099','pr100',
                  'pr101','pr102','pr103','pr104','pr105','pr106','pr107','pr108','pr109','pr110','pr111','pr112','pr113','pr114','pr115','pr116','pr117','pr118','pr119','pr120',

                  'xa001','xa002','xb001','xb002','xb003'
        ]



        Color = ['000000', 'ffffff', 'ff0000', '0000ff', '00ff00', 'ffff00', '000055']  # 黑,白,紅,藍,綠,黃
        fille = PatternFill('solid', fgColor=Color[6])
        font = Font(u'微軟正黑體', size=12, bold=True, italic=False, strike=False, color=Color[5])  # 设置字体样式
        align = Alignment(horizontal="center", vertical="center",wrap_text=True)
        # 寫Excel檔
        wb = openpyxl.Workbook()
        for job in jobtitle_list:  #依不同職務明稱寫入sheet
            query = ""
            query = Q(work_code_title__work_code__factory=factory)
            if (date_s[0] == date_e[0]):  # 年度相等
                query &= Q(year=date_s[0], month__gte=date_s[1], month__lte=date_e[1])
            else:  # 年度不同
                query = Q(year=date_s[0], month__gte=date_s[1], month__lte=12)
                query |= Q(year=date_e[0], month__gte=1, month__lte=date_e[1])
            query &= Q(work_code_title__job_title=job)
            results = list( MatrixMaster.objects.filter(query).exclude(bpm__isnull=True).
                    annotate( #field alias
                        dept_desc1 = F('work_code_title__work_code__dept__desc1'),
                        work_code = F('work_code_title__work_code'),
                        chi_name = F('work_code_title__work_code__chi_name'),
                        arrival_date = F('work_code_title__work_code__arrival_date'),
                        rank_desc1 = F('work_code_title__work_code__rank__desc1'),
                        pos_desc1 = F('work_code_title__work_code__pos__desc1'),
                        job_title = F('work_code_title__job_title'),
                        job_name = F('work_code_title__job_title__job_name'),
                        cr001 = F('matrix_detail__cr001'),
                        cr002 = F('matrix_detail__cr002'),
                        cr003 = F('matrix_detail__cr003'),
                        cr004 = F('matrix_detail__cr004'),
                        cr005 = F('matrix_detail__cr005'),
                        cr006 = F('matrix_detail__cr006'),
                        cr007 = F('matrix_detail__cr007'),
                        cr008 = F('matrix_detail__cr008'),
                        cr009 = F('matrix_detail__cr009'),
                        cr010 = F('matrix_detail__cr010'),
                        cr011 = F('matrix_detail__cr011'),
                        cr012 = F('matrix_detail__cr012'),
                        cr013 = F('matrix_detail__cr013'),
                        cr014 = F('matrix_detail__cr014'),
                        cr015 = F('matrix_detail__cr015'),
                        cr016 = F('matrix_detail__cr016'),
                        cr017 = F('matrix_detail__cr017'),
                        cr018 = F('matrix_detail__cr018'),
                        cr019 = F('matrix_detail__cr019'),
                        cr020 = F('matrix_detail__cr020'),

                        ma001 = F('matrix_detail__ma001'),
                        ma002 = F('matrix_detail__ma002'),
                        ma003 = F('matrix_detail__ma003'),
                        ma004 = F('matrix_detail__ma004'),
                        ma005 = F('matrix_detail__ma005'),
                        ma006 = F('matrix_detail__ma006'),
                        ma007 = F('matrix_detail__ma007'),
                        ma008 = F('matrix_detail__ma008'),
                        ma009 = F('matrix_detail__ma009'),
                        ma010 = F('matrix_detail__ma010'),
                        ma011 = F('matrix_detail__ma011'),
                        ma012 = F('matrix_detail__ma012'),
                        ma013 = F('matrix_detail__ma013'),
                        ma014 = F('matrix_detail__ma014'),
                        ma015 = F('matrix_detail__ma015'),
                        ma016 = F('matrix_detail__ma016'),
                        ma017 = F('matrix_detail__ma017'),
                        ma018 = F('matrix_detail__ma018'),
                        ma019 = F('matrix_detail__ma019'),
                        ma020 = F('matrix_detail__ma020'),
                        ma021 = F('matrix_detail__ma021'),
                        ma022 = F('matrix_detail__ma022'),
                        ma023 = F('matrix_detail__ma023'),
                        ma024 = F('matrix_detail__ma024'),
                        ma025 = F('matrix_detail__ma025'),
                        ma026 = F('matrix_detail__ma026'),
                        ma027 = F('matrix_detail__ma027'),
                        ma028 = F('matrix_detail__ma028'),
                        ma029 = F('matrix_detail__ma029'),
                        ma030 = F('matrix_detail__ma030'),
                        ma031 = F('matrix_detail__ma031'),
                        ma032 = F('matrix_detail__ma032'),
                        ma033 = F('matrix_detail__ma033'),
                        ma034 = F('matrix_detail__ma034'),
                        ma035 = F('matrix_detail__ma035'),
                        ma036 = F('matrix_detail__ma036'),
                        ma037 = F('matrix_detail__ma037'),
                        ma038 = F('matrix_detail__ma038'),
                        ma039 = F('matrix_detail__ma039'),
                        ma040 = F('matrix_detail__ma040'),
                        ma041 = F('matrix_detail__ma041'),
                        ma042 = F('matrix_detail__ma042'),
                        ma043 = F('matrix_detail__ma043'),
                        ma044 = F('matrix_detail__ma044'),
                        ma045 = F('matrix_detail__ma045'),
                        ma046 = F('matrix_detail__ma046'),
                        ma047 = F('matrix_detail__ma047'),
                        ma048 = F('matrix_detail__ma048'),
                        ma049 = F('matrix_detail__ma049'),
                        ma050 = F('matrix_detail__ma050'),
                        ma051 = F('matrix_detail__ma051'),
                        ma052 = F('matrix_detail__ma052'),
                        ma053 = F('matrix_detail__ma053'),
                        ma054 = F('matrix_detail__ma054'),
                        ma055 = F('matrix_detail__ma055'),
                        ma056 = F('matrix_detail__ma056'),
                        ma057 = F('matrix_detail__ma057'),
                        ma058 = F('matrix_detail__ma058'),
                        ma059 = F('matrix_detail__ma059'),
                        ma060 = F('matrix_detail__ma060'),

                        ge001 = F('matrix_detail__ge001'),
                        ge002 = F('matrix_detail__ge002'),
                        ge003 = F('matrix_detail__ge003'),
                        ge004 = F('matrix_detail__ge004'),
                        ge005 = F('matrix_detail__ge005'),
                        ge006 = F('matrix_detail__ge006'),
                        ge007 = F('matrix_detail__ge007'),
                        ge008 = F('matrix_detail__ge008'),
                        ge009 = F('matrix_detail__ge009'),
                        ge010 = F('matrix_detail__ge010'),
                        ge011 = F('matrix_detail__ge011'),
                        ge012 = F('matrix_detail__ge012'),
                        ge013 = F('matrix_detail__ge013'),
                        ge014 = F('matrix_detail__ge014'),
                        ge015 = F('matrix_detail__ge015'),
                        ge016 = F('matrix_detail__ge016'),
                        ge017 = F('matrix_detail__ge017'),
                        ge018 = F('matrix_detail__ge018'),
                        ge019 = F('matrix_detail__ge019'),
                        ge020 = F('matrix_detail__ge020'),

                        pr001 = F('matrix_detail__pr001'),
                        pr002 = F('matrix_detail__pr002'),
                        pr003 = F('matrix_detail__pr003'),
                        pr004 = F('matrix_detail__pr004'),
                        pr005 = F('matrix_detail__pr005'),
                        pr006 = F('matrix_detail__pr006'),
                        pr007 = F('matrix_detail__pr007'),
                        pr008 = F('matrix_detail__pr008'),
                        pr009 = F('matrix_detail__pr009'),
                        pr010 = F('matrix_detail__pr010'),
                        pr011 = F('matrix_detail__pr011'),
                        pr012 = F('matrix_detail__pr012'),
                        pr013 = F('matrix_detail__pr013'),
                        pr014 = F('matrix_detail__pr014'),
                        pr015 = F('matrix_detail__pr015'),
                        pr016 = F('matrix_detail__pr016'),
                        pr017 = F('matrix_detail__pr017'),
                        pr018 = F('matrix_detail__pr018'),
                        pr019 = F('matrix_detail__pr019'),
                        pr020 = F('matrix_detail__pr020'),
                        pr021 = F('matrix_detail__pr021'),
                        pr022 = F('matrix_detail__pr022'),
                        pr023 = F('matrix_detail__pr023'),
                        pr024 = F('matrix_detail__pr024'),
                        pr025 = F('matrix_detail__pr025'),
                        pr026 = F('matrix_detail__pr026'),
                        pr027 = F('matrix_detail__pr027'),
                        pr028 = F('matrix_detail__pr028'),
                        pr029 = F('matrix_detail__pr029'),
                        pr030 = F('matrix_detail__pr030'),
                        pr031 = F('matrix_detail__pr031'),
                        pr032 = F('matrix_detail__pr032'),
                        pr033 = F('matrix_detail__pr033'),
                        pr034 = F('matrix_detail__pr034'),
                        pr035 = F('matrix_detail__pr035'),
                        pr036 = F('matrix_detail__pr036'),
                        pr037 = F('matrix_detail__pr037'),
                        pr038 = F('matrix_detail__pr038'),
                        pr039 = F('matrix_detail__pr039'),
                        pr040 = F('matrix_detail__pr040'),
                        pr041 = F('matrix_detail__pr041'),
                        pr042 = F('matrix_detail__pr042'),
                        pr043 = F('matrix_detail__pr043'),
                        pr044 = F('matrix_detail__pr044'),
                        pr045 = F('matrix_detail__pr045'),
                        pr046 = F('matrix_detail__pr046'),
                        pr047 = F('matrix_detail__pr047'),
                        pr048 = F('matrix_detail__pr048'),
                        pr049 = F('matrix_detail__pr049'),
                        pr050 = F('matrix_detail__pr050'),
                        pr051 = F('matrix_detail__pr051'),
                        pr052 = F('matrix_detail__pr052'),
                        pr053 = F('matrix_detail__pr053'),
                        pr054 = F('matrix_detail__pr054'),
                        pr055 = F('matrix_detail__pr055'),
                        pr056 = F('matrix_detail__pr056'),
                        pr057 = F('matrix_detail__pr057'),
                        pr058 = F('matrix_detail__pr058'),
                        pr059 = F('matrix_detail__pr059'),
                        pr060 = F('matrix_detail__pr060'),
                        pr061 = F('matrix_detail__pr061'),
                        pr062 = F('matrix_detail__pr062'),
                        pr063 = F('matrix_detail__pr063'),
                        pr064 = F('matrix_detail__pr064'),
                        pr065 = F('matrix_detail__pr065'),
                        pr066 = F('matrix_detail__pr066'),
                        pr067 = F('matrix_detail__pr067'),
                        pr068 = F('matrix_detail__pr068'),
                        pr069 = F('matrix_detail__pr069'),
                        pr070 = F('matrix_detail__pr070'),
                        pr071 = F('matrix_detail__pr071'),
                        pr072 = F('matrix_detail__pr072'),
                        pr073 = F('matrix_detail__pr073'),
                        pr074 = F('matrix_detail__pr074'),
                        pr075 = F('matrix_detail__pr075'),
                        pr076 = F('matrix_detail__pr076'),
                        pr077 = F('matrix_detail__pr077'),
                        pr078 = F('matrix_detail__pr078'),
                        pr079 = F('matrix_detail__pr079'),
                        pr080 = F('matrix_detail__pr080'),
                        pr081 = F('matrix_detail__pr081'),
                        pr082 = F('matrix_detail__pr082'),
                        pr083 = F('matrix_detail__pr083'),
                        pr084 = F('matrix_detail__pr084'),
                        pr085 = F('matrix_detail__pr085'),
                        pr086 = F('matrix_detail__pr086'),
                        pr087 = F('matrix_detail__pr087'),
                        pr088 = F('matrix_detail__pr088'),
                        pr089 = F('matrix_detail__pr089'),
                        pr090 = F('matrix_detail__pr090'),
                        pr091 = F('matrix_detail__pr091'),
                        pr092 = F('matrix_detail__pr092'),
                        pr093 = F('matrix_detail__pr093'),
                        pr094 = F('matrix_detail__pr094'),
                        pr095 = F('matrix_detail__pr095'),
                        pr096 = F('matrix_detail__pr096'),
                        pr097 = F('matrix_detail__pr097'),
                        pr098 = F('matrix_detail__pr098'),
                        pr099 = F('matrix_detail__pr099'),
                        pr100 = F('matrix_detail__pr100'),
                        pr101 = F('matrix_detail__pr101'),
                        pr102 = F('matrix_detail__pr102'),
                        pr103 = F('matrix_detail__pr103'),
                        pr104 = F('matrix_detail__pr104'),
                        pr105 = F('matrix_detail__pr105'),
                        pr106 = F('matrix_detail__pr106'),
                        pr107 = F('matrix_detail__pr107'),
                        pr108 = F('matrix_detail__pr108'),
                        pr109 = F('matrix_detail__pr109'),
                        pr110 = F('matrix_detail__pr110'),
                        pr111 = F('matrix_detail__pr111'),
                        pr112 = F('matrix_detail__pr112'),
                        pr113 = F('matrix_detail__pr113'),
                        pr114 = F('matrix_detail__pr114'),
                        pr115 = F('matrix_detail__pr115'),
                        pr116 = F('matrix_detail__pr116'),
                        pr117 = F('matrix_detail__pr117'),
                        pr118 = F('matrix_detail__pr118'),
                        pr119 = F('matrix_detail__pr119'),
                        pr120 = F('matrix_detail__pr120'),

                        xa001 = F('matrix_detail__xa001'),
                        xa002 = F('matrix_detail__xa002'),

                        xb001 = F('matrix_detail__xb001'),
                        xb002 = F('matrix_detail__xb002'),
                        xb003 = F('matrix_detail__xb003'),
                    ).values( *fields ).order_by('year','month','job_title','work_code'))

            efields = []
            etitles = []
            if results:
                rrr = JobTitle.objects.filter(job_code=job).values_list('job_name')
                for r in rrr:
                    job_name = r[0]
                # 建立sheet
                efields.extend(['no', 'dept_desc', 'work_code', 'chi_name', 'arrival_date', 'rank_desc', 'pos_desc',
                           'job_name'])
                etitles.extend(['項次', '單位', '卡號', '姓名', '到職日期', '職等', '職稱', '職務名稱'])
                field_com, title_com = self.get_common_skills()
                field_pro, title_pro = self.get_pro_skills(job)
                efields.extend(field_com)
                efields.extend(field_pro)
                efields.extend(['xa001', 'xa002', 'xb001', 'xb002', 'xb003'])
                etitles.extend(title_com)
                etitles.extend(title_pro)
                etitles.extend(['佔總分的比率 ％', '成熟度', '教育訓練', '課程名稱', '預定輪調的職務'])
                sheetX = wb.create_sheet(job_name, 0)
                # sheetX.append(efields)
                sheetX.append(etitles)
                for i in range(0, len(etitles)):
                    sheetX.cell(row=1, column=i + 1).fill = fille
                    sheetX.cell(row=1, column=i + 1).font = font
                    sheetX.cell(row=1, column=i + 1).alignment = align

                rowCount = 0
                for dt in results:
                    rowCount = rowCount + 1
                    rowX = []
                    rowX.extend([
                        rowCount,
                        dt.get('dept_desc1'),
                        dt.get('work_code'),
                        dt.get('chi_name'),
                        dt.get('arrival_date'),
                        dt.get('rank_desc1'),
                        dt.get('pos_desc1'),
                        dt.get('job_name'),
                    ])
                    for idx, (key, val) in enumerate(dt.items()):
                        if idx>10 and val!=None:
                            rowX.append(val)
                    sheetX.append(rowX)
                    fontX = Font(u'微軟正黑體', size=24, italic=False, strike=False)  # 设置字体样式
                    fontY = Font(color=Color[3],bold=True)  # 设置字体样式
                    alignX = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    special_width = 15
                    # for i in range(66, 72):
                    #     sheetX.column_dimensions[chr(i)].width = special_width
                    sheetX.column_dimensions['C'].width = special_width
                    sheetX.column_dimensions['E'].width = special_width

                    for i in range(0, len(etitles)):
                        if sheetX.cell(row=rowCount+1, column=i + 1).value in ["◔","◑","◕"]:
                            sheetX.cell(row=rowCount+1, column=i + 1).font = fontX

                        if sheetX.cell(row=rowCount+1, column=i + 1).value not in ["○","◔","◑","◕","●","X"]:
                            sheetX.cell(row=rowCount+1, column=i + 1).font = fontY
                        sheetX.cell(row=rowCount+1, column=i + 1).alignment = alignX

        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y%m%d_%H%M%S')
        filename = 'tt422_技能盤點資料匯出_' + now + '.xlsx'
        output_file = '%s/reports/xlsx/%s' % (settings.MEDIA_ROOT, filename)
        wb.save(output_file)
        openFile = '/media-files/reports/xlsx/'+filename
        return JsonResponse({"success": True,"filename":filename,"openFile":openFile})


def get_last_job_code(request,job_code_l2=None):
    last_job_code = JobTitle.objects.filter(job_code__startswith=job_code_l2).values_list('job_code',).order_by('job_code').last()
    return JsonResponse({"last_job_code": str(int(last_job_code[0])+1).rjust(6,'0') })


def get_last_skill_code(request,skill_class=None):
    last_skill_code = JobSkill.objects.filter(skill_class=skill_class).values_list('skill_code',).order_by('skill_code').last()[0]
    if last_skill_code[0:2] == 'PR':
        number_str = "0123456789"

        find_result =  number_str.find(last_skill_code[2:3])

        if number_str.find(last_skill_code[2:3])==1:
            last_skill_code = last_skill_code[0:2] + str(int(last_skill_code[2:6]) + 1).rjust(4, '0')
        else:
            last_skill_code = last_skill_code[0:3] + str(int(last_skill_code[3:6]) + 1).rjust(3, '0')
    else:
        last_skill_code = last_skill_code[0:2]+str(int(last_skill_code[2:6]) + 1).rjust(4, '0')
    return JsonResponse({"last_skill_code": last_skill_code})


def get_job_title_l3(request):
    results = JobTitle.objects.filter(level_number=3).values_list('job_code','job_name')
    dataList = []
    if results!=None:
        for T in results:
            dataList.append({
                'job_code': T[0],
                'job_name': T[1],
        })
    return JsonResponse(dataList,safe=False)


def get_job_tiitle_skill_order_number(request,job_title=None):     #pk1 : job_title  pk2 : userID
    try:
        last_number = JobTitleSkill.objects.filter(job_title=job_title).values_list('order_number').order_by('job_title','order_number').last()[0]
    except:
        last_number = 0
    next_number = last_number + 1
    return JsonResponse({"next_number": next_number})



def valid_employee_title_skill(request,job_title=None):
    # job_title_l3 = JobTitleSkill.objects.filter(job_title_id=job_title)
    skill_count = JobTitleSkill.objects.filter(job_title_id=job_title).count()
    if skill_count > 0:
        return JsonResponse({"success": True})
    else:
        return JsonResponse({"success": False,"message":"***職務未綁定職能，不允新增***"})


def employee_title(request):
    fields = [
        'work_code',
        'work_code__chi_name',
        'job_title',
        'job_title__job_name',
        'id',
    ]
    results = EmployeeTitle.objects.all().values_list(*fields)
    dataList = []
    if results != None:
        for T in results:
            dataList.append({
                'job_code': T[2],
                'job_name': T[3],
                'work_code': T[0],
                'chi_name': T[1],
                'id': T[4],
            })
    return JsonResponse(dataList, safe=False)


def matrix_master_director(request):
    fields = [
        'direct_supv',
        'supv_name',
        'dept',
        'year',
        'month',
        'bpm',
        'bpm_desc1',
        'report_url',
        'dept_udc',
        # 'work_code_title',
    ]

    # results = MetricsSetup.objects.filter(Q(work_code=pk1), Q(date_yyyy=pk2), Q(order_number__lt=900)).values_list('work_code', 'date_yyyy', 'date_mm').order_by('work_code', 'date_yyyy', 'date_mm').distinct()
    results = MatrixMaster.objects.all().annotate(
        direct_supv = F('work_code_title__work_code__direct_supv'),
        supv_name = F('work_code_title__work_code__direct_supv__chi_name'),
        dept_udc=F('work_code_title__work_code__dept_flevel'),
        dept = F('work_code_title__work_code__dept_flevel__desc1'),
        bpm_desc1 = F('bpm__bpm_status_desc1'),
        report_url = F('bpm__report_url'),
    ).values_list(*fields).order_by('direct_supv','dept','year','month','bpm').distinct()
    dataList = []
    if results != None:
        for T in results:
            dataList.append({
                'direct_supv':T[0],
                'supv_name':T[1],
                'dept':T[2],
                # 'work_code':T[3],
                # 'chi_name':T[4],
                # 'job_code':T[5],
                # 'job_name':T[6],
                'year':T[3],
                'month':T[4],
                'bpm':T[5],
                'bpm_desc1': T[6],
                'bpm_desc2': UserDefCode.objects.get(topic_code_id='bpm_status',desc1=T[6]).desc2 if T[6] else '',
                'report_url':T[7].replace("media","media-files") if T[7] else '',
                'dept_udc':T[8]
                # 'work_code_title':T[8],
            })
    return JsonResponse(dataList, safe=False)



def supv_work_code_title(request,direct_supv=None,dept_udc=None):
    fields = [ 'id',
               'work_code',
               'work_code__chi_name',
               'job_title',
               'job_title__job_name'
               ]

    qry1 = Q(enable=True)  # 啟用
    if direct_supv=='undefined':   #取全部
        dataList = [ { 'value':"", 'text':'---------' } ]
        r = EmployeeTitle.objects.filter(qry1).values_list(*fields)
    else:
        dataList = []
        qry2 = Q(work_code__in=EmployeeInfoEasy.objects.values_list('work_code').filter(direct_supv=direct_supv,
                                                                                        dept_flevel=dept_udc))
        r = EmployeeTitle.objects.filter(qry1, qry2).values_list(*fields)


    for T in r:
        dataList.append({
            'value': T[0],
            'text': T[1]+" "+T[2]+" "+T[3]+" "+T[4],
        })
    return JsonResponse(dataList, safe=False)


def matrix_master_employee(request,direct_supv=None,dept=None,year=None,month=None,bpm=None):
    fields = [
        'work_code_title__work_code__direct_supv',
        'work_code_title__work_code__direct_supv__chi_name',
        'work_code_title__work_code__dept_flevel__desc1',
        'work_code_title__work_code',
        'work_code_title__work_code__chi_name',
        'work_code_title__job_title',
        'work_code_title__job_title__job_name',
        'year',
        'month',
        'bpm',
        'bpm__bpm_status_desc1',
        'bpm__report_url',
        'work_code_title',
        'id',
    ]


    if bpm=='null':
        results = MatrixMaster.objects.filter(
            work_code_title__work_code__direct_supv = direct_supv,
            work_code_title__work_code__dept_flevel = dept,
            year = year , month = month ,  bpm__isnull=True ).values_list(*fields)
    else:
        results = MatrixMaster.objects.filter(
            work_code_title__work_code__direct_supv = direct_supv,
            work_code_title__work_code__dept_flevel = dept,
            year = year , month = month ,bpm=bpm).values_list(*fields)

    dataList = []
    if results != None:
        for T in results:
            detail_count = MatrixDetail.objects.filter(master_id=T[13]).count()
            dataList.append({
                'direct_supv':T[0],
                'supv_name':T[1],
                'dept':T[2],
                'work_code':T[3],
                'chi_name':T[4],
                'job_code':T[5],
                'job_name':T[6],
                'year':T[7],
                'month':T[8],
                'bpm':T[9],
                'bpm_desc1': T[10],
                'bpm_desc2': UserDefCode.objects.get(topic_code_id='bpm_status',desc1=T[10]).desc2 if T[10] else '',
                'report_url':T[11].replace("media","media-files") if T[11] else '',
                'work_code_title': T[12],
                'detail_yn': 'Y' if detail_count>=1 else 'N',
                'id':T[13],
            })
    return JsonResponse(dataList, safe=False)


def get_work_code_direct_supv(request,work_code=None):
    ee =  model_to_dict( EmployeeInfoEasy.objects.get(work_code=work_code) )
    direct_supv = ee.get('direct_supv', None)
    dept_desc = UserDefCode.objects.get( id = ee.get('dept_flevel',None) ).desc1
    supv_name = EmployeeInfoEasy.objects.get(work_code=direct_supv).chi_name
    return JsonResponse({'success':True,'direct_supv':direct_supv,'supv_name':supv_name,'dept':dept_desc})


def pdca_export_data_count(request,master_id=None):
    now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')
    instance = PdcaMaster.objects.get(id=master_id)
    dt = model_to_dict(instance)
    exp_count = dt.get('exp_count', None)
    data={
        'exp_count': exp_count + 1,
        'exp_time': now,
    }
    instance.__dict__.update(**data)
    instance.save()
    return JsonResponse({'success':True})



def pdca_detail(request,job_title=None,userId=None,add_version=None):
    choice_language = request.session.get('choice_language')
    view_name = (request.path).split('/')[2]  # ['', 'skill_pdca', 'tt602_2pdf']
    def_line = PdcaDetail.objects.get(id=0).order_number      #在 order_number 定義總行數

    qr1 = Q(work_code=userId)
    qr2 = Q(job_title=job_title)
    last_the_time = PdcaMaster.objects.filter(qr1, qr2).values_list('work_code','job_title','the_time','id').order_by('the_time').last()
    # print("\n"*5)
    # print("="*220)
    # print(" last_the_time=",last_the_time)
    if ( add_version == "Y" ):
        qrx = Q(the_time=last_the_time[2]+1)
        last_the_time = PdcaMaster.objects.filter(qr1, qr2,qrx).values_list('work_code', 'job_title', 'the_time','id').order_by('the_time')

    # order_number > 9990 是合計
    qr3 = Q(master_id=last_the_time[3]) if last_the_time else None          #未送BPM   留待js控制按鈕
    qr4 = Q(master__bpm_status_desc1__exact='')  #未送BPM   留待js控制按鈕
    qr5 = Q(master__bpm_status_desc1='reject')   #BPM退回   留待js控制按鈕
    qr6 = Q(order_number__lt=9990)               #BPM退回   留待js控制按鈕

    view_name_pdf = ['tt602_2pdf','tt603_tab2pdf']
    # if (view_name == 'tt602_2pdf' or view_name == 'tt603_tab2pdf'):
    if (view_name in view_name_pdf):
        # 合計, 要在報表顯示
        r = PdcaDetail.objects.filter(qr3).order_by('master','order_number')
    else:
        # 合計, 由當下使用者修改, 重新計算
        if last_the_time:
            r = PdcaDetail.objects.filter( qr3,qr6 ).order_by('master','order_number')
        else:
            r = None


    json_lists = []
    if r:
        for idx,itm in enumerate(r):
            # print(idx,itm)
            json_dict = model_to_dict(itm)
            # if (view_name == 'tt602_2pdf' or  view_name == 'tt603_tab2pdf'):
            if (view_name in view_name_pdf):
                json_dict.update({"job_title":job_title})
            json_lists.append(json_dict)
        # if (view_name != 'tt602_2pdf' ) :
        if ( not view_name in view_name_pdf ) :
            if (idx+1) < def_line:       #行數不足,補足行數
                r = PdcaDetail.objects.filter(id=0)  # id=0 預設值, 為取得table schema(field)使用
                blank_line = def_line - idx
                for j in range( 1,  blank_line ):
                    blank_order_number = j + idx + 1
                    for i in r:
                        json_dict = model_to_dict(i)
                        json_dict.update({'order_number' : blank_order_number} )
                        json_lists.append(json_dict)

    else:
        r = PdcaDetail.objects.filter(id=0)    # id=0 預設值, 為取得table schema(field)使用
        for j in range(1,def_line+1):
            for i in r:
                json_dict = model_to_dict(i)
                json_dict.update({'order_number':j})
                json_lists.append(json_dict)

    # print("\n"*5)
    # print("*"*210)
    # print("="*210)
    # for idx,itm in enumerate(json_lists):
    #     print("idx=",idx)
    #     print("-"*50)
    #     print(itm)
    #     print("-" * 50)
    # print("="*210)
    # print("*"*210)
    # print("\n"*5)

    translator = Translator()
    for idx,itm in enumerate(json_lists):
        if itm.get('order_number') < 9000:
            source_text = itm.get("work_item")
            if source_text :
                text_lang = translator.detect(source_text).lang
                if (text_lang in ['zh-TW', 'zh-CN']) and (choice_language in ["zh-hant", "zh-hans"]):
                    # 中文, 不用翻譯
                    # print("work_item = ", source_text, "     The translator has nothing to do.")
                    break
                else:
                    # 有找到翻譯的語言
                    if choice_language==itm.get('lang_code_out'):
                        json_lists[idx].update({
                            'work_item':itm.get('fwork_item'),
                            'cycl01':itm.get('fcycl01')
                        })
                        for i in range(0,13):
                            field_name = 'flow' + ('0' if i < 9 else '') + str(i + 1)
                            ffield_name = 'f'+field_name
                            json_lists[idx].update({
                                field_name:itm.get(ffield_name)
                            })

    for idx, itm in enumerate(json_lists):
        # 刪除不需列印的資料--------------------------------------------------------begin
        json_lists[idx].pop('lang_code_in')
        json_lists[idx].pop('lang_code_out')
        json_lists[idx].pop('fwork_item')
        json_lists[idx].pop('fcycl01')
        for i in range(0, 12):
            ffield_name = 'fflow' + ('0' if i < 9 else '') + str(i + 1)
            json_lists[idx].pop(ffield_name)
        # 刪除不需列印的資料--------------------------------------------------------end
        # for idx,itm in enumerate(json_lists):
        #     print("idx=",idx)
        #     print("-"*50)
        #     print(itm)
        #     print("-" * 50)
    # print("="*210)
    # print("*"*210)
    # print("\n"*5)

    if ( view_name in view_name_pdf ):
        return json_lists
    else:
        return JsonResponse(json_lists, safe=False)


# def google_translator(curr_lang,dtext):
#     translator = Translator()
#     text_lang = translator.detect(dtext)
#     print("text_lang detect=",text_lang)
#     print("curr_lang=",curr_lang)


class update_tabs_datagrid_rows_pdca(View):
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        post_param = list( request.POST.dict() )    #整個所有物件內容, 是一個key
        data = json.loads(post_param[0])            #取得dict第一個key值
        # del data['current_lang']
        choice_language = request.session.get('choice_language')
        translator = Translator()
        update_results = []
        for key,val in data.items():
            dt_detail = { key: ( float(value) if key[0:4] in ('pdca','calc','ptot','dtot','ctot','atot') else value ) for i, (key, value) in enumerate(val.items()) }
            if dt_detail.get('work_item',None) ==  None : continue
            work_code = dt_detail.pop('work_code'),
            the_time = dt_detail.pop('the_time'),
            job_title = dt_detail.pop('job_title'),
            job_title_desc = dt_detail.pop('job_title_desc'),
            dt_master = ({
                "work_code_id": work_code[0],
                "the_time": the_time[0],
                "job_title_id": job_title[0],
                "create_time": now,
                "change_time": now,
                "changer": request.user.username,
                "creator": request.user.username,
            })
            try:
                instance_master = PdcaMaster.objects.create(**dt_master)
                instance_master = PdcaMaster.objects.filter(work_code=work_code[0], the_time=the_time[0],
                                                         job_title=job_title[0]).values_list('id')
            except:
                 try:
                     instance_master = PdcaMaster.objects.filter(work_code=work_code[0], the_time=the_time[0],
                                                                 job_title=job_title[0]).values_list('id')
                 except:
                     print("資料錯誤, 無法建立, 也無法更新")

            if instance_master:
                for T in instance_master:
                    master_id = T[0]
                    source_text = dt_detail.get("work_item")
                    text_lang = translator.detect(source_text).lang
                    print("="*200)
                    print("text_lang:",text_lang,"    choice_language=",choice_language)
                    print("*"*200)
                    print(source_text)
                    if (text_lang in ['zh-TW', 'zh-CN','en']) and (choice_language in ["zh-hant","zh-hans"]):
                        # 中文, 不用翻譯
                        print("work_item = ",source_text,"The translator has nothing to do.")
                    else:
                        # --------------翻譯開始-------------------------------------------------------------------------------------翻譯開始
                        if dt_detail.get("order_number") < 9000:
                            translator = Translator()
                            source_text = dt_detail.get("work_item")
                            dt_detail['lang_code_in'] = text_lang
                            dt_detail['lang_code_out'] = choice_language
                            if text_lang in ['zh-TW', 'zh-CN']:  # 資料, 是"繁體中文"
                                source_text = dt_detail.get("work_item")
                                if choice_language not in ["zh-hant","zh-hans"]:  # 網頁選擇的不是"繁體中文/簡體中文"
                                    if source_text:
                                        trans_foreign = translator.translate(source_text, dest=choice_language).text
                                        dt_detail['fwork_item'] = trans_foreign
                                    for i in range(0, 12):
                                        field_name = 'flow' + ('0' if i < 9 else '') + str(i + 1)
                                        source_text = dt_detail.get(field_name)
                                        if source_text:
                                            trans_foreign = translator.translate(source_text, dest=choice_language).text
                                            # dt_detail['fflow0' + str(i)] = trans_foreign
                                            dt_detail['f'+field_name ] = trans_foreign
                                    source_text = dt_detail.get("cycl01")
                                    if source_text:
                                        trans_foreign = translator.translate(source_text, dest=choice_language).text
                                        dt_detail['fcycl01'] = trans_foreign
                            else:  # 資料, 不是"繁體中文", 就要翻譯一份繁體中
                                #先儲存外國語言
                                dt_detail['fwork_item'] = dt_detail['work_item']
                                for i in range(0, 12):
                                    field_name = 'flow' + ('0' if i < 9 else '') + str(i + 1)
                                    source_text = dt_detail.get(field_name)
                                    dt_detail['f'+field_name] = source_text
                                source_text = dt_detail.get("cycl01")
                                dt_detail['fcycl01'] = source_text

                                #將外國言言翻譯成中文後, 儲存在原來的位置
                                if source_text:
                                    trans_TW = translator.translate(source_text, dest='zh-TW').text
                                    dt_detail['work_item'] = trans_TW
                                for i in range(0, 12):
                                    flow_name = 'flow0' + str(i)
                                    field_name = 'flow' + ('0' if i < 9 else '') + str(i + 1)
                                    source_text = dt_detail.get(field_name)
                                    if source_text:
                                        trans_TW = translator.translate(source_text, dest='zh-TW').text
                                        dt_detail[flow_name] = trans_TW
                                source_text = dt_detail.get("cycl01")
                                if source_text:
                                    trans_TW = translator.translate(source_text, dest=choice_language).text
                                    dt_detail['cycl01'] = trans_TW
                        # --------------翻譯結束--------------------------------------------------------------------------------------翻譯結束

                    dt_detail.update({
                        'master_id':master_id,
                        "create_time": now,
                        "change_time": now,
                        "changer": request.user.username,
                        "creator": request.user.username,
                    })
                    try:
                        PdcaDetail.objects.create(**dt_detail)
                        update_results.append({
                            "job_title_desc": job_title_desc,
                            "order_number": dt_detail.get('order_number'),
                            "success": True,
                        })
                    except:
                        try:
                            instance_detail = PdcaDetail.objects.filter(master_id=master_id,order_number=dt_detail.get('order_number') )
                            if instance_detail:
                                dt_detail.update({
                                    "create_time": now,
                                    "change_time": now,
                                    "changer": request.user.username,
                                    "creator": request.user.username,
                                })
                                # try:
                                # for x in dt_detail:
                                #     print(x)
                                instance_detail.update(**dt_detail)
                                #     update_results.append({
                                #         "job_title_desc": job_title_desc,
                                #         "order_number": dt_detail.get('order_number'),
                                #         "success": True,
                                #     })
                                # except:
                                #     print("資料錯誤, 無法建立, 也無法更新")
                                #     update_results.append({
                                #         "job_title_desc": job_title_desc,
                                #         "order_number": dt_detail.get('order_number'),
                                #         "success": False,
                                #     })
                        except:
                            pass
            else:
                update_results.append({
                    "job_title_desc": job_title_desc,
                    "order_number": dt_detail.get('order_number'),
                    "success": False,
                })
                return JsonResponse({"results": update_results})
        # 計算非"summary"的行號有幾行
        row_length = 0
        new_key = ""
        old_key = ""
        for key, val in data.items():
            work_code = val.get('work_code')
            the_time = val.get('the_time')
            job_title = val.get('job_title')
            new_key = work_code + "-" + str(the_time) + "-" + job_title
            old_key = new_key if (key == 0) else old_key
            if ( val.get('order_number') < 9000 ):
                row_length = (row_length + 1) if ( new_key == old_key ) else 1
            if ( val.get('order_number') == 9993 ):
                qry1 = Q( master__work_code=work_code,master__the_time=the_time,master__job_title=job_title )
                qry2 = Q(order_number__lt=9000)
                detail_dele = PdcaDetail.objects.filter(qry1,qry2).order_by('order_number')
                if ( len(detail_dele) > 0 and row_length < len(detail_dele) ):
                    # print("***** row_length<len(detail_dele)")
                    data_length = 0
                    for x_instance in detail_dele:
                        data_length += 1
                        print("data_length=", data_length)
                        if data_length > row_length:
                            print("***此筆要刪除", x_instance)
                            print("-" * 50)
                            # 刪除，只更新前幾筆，最後幾筆，不會更新，還留在ＴＡＢＬＥ裏，留在ＴＡＢＬＥ在Datagrid是空白，必需刪除
                            x_instance.delete()

            old_key = new_key
        return JsonResponse({"results": update_results})


def pdca_master(request):
    fields = [
        'work_code',
        'work_code__chi_name',
        'the_time',
        'job_title__job_code',
        'job_title__job_name',
        'bpm_number',
        'bpm_status_desc1',
        'report_name',
        'report_url',
    ]

    results = PdcaMaster.objects.all().values_list(*fields)
    dataList = []
    if results != None:
        for T in results:
            dataList.append({
                'work_code': T[0],
                'chi_name': T[1],
                'the_time': T[2],
                'job_code': T[3],
                'job_name': T[4],
                'bpm_number': T[5],
                'bpm_desc1': T[6],
                'bpm_desc2': UserDefCode.objects.get(topic_code_id='bpm_status',desc1=T[6]).desc2 if T[6] else '',
                'report_name': T[7],
                'report_url': T[8].replace("media","media-files") if T[8] else '',
            })
    return JsonResponse(dataList, safe=False)



#員工基本資料的匯入
def pdca_detail_import(request):
    if request.method == 'POST':
        fn = request.FILES['myfile']
        workbook = openpyxl.load_workbook(fn)

        # 獲得所有sheet的名稱
        sheets = workbook.get_sheet_names()
        # 獲得當前正在顯示的sheet
        sheet = workbook.active
        alist = []

        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')


        # import_list = []
        # for i in range(2, sheet.max_row):
        #     alist = list(sheet.rows)[i]
        #     instance=None
        #     import_tuple = ()
        #     update_tuple = ()
        #     update_list = []
        #     update_param = {}
        #     import_param = {}
        #     for cell in alist:
        #         if alist.index(cell)==0:  #第一格是工號
        #             instance = EmployeeInfoEasy.objects.filter(work_code=cell.value)
        #         if instance:
        #             update_tuple += tuple([cell.value])   #將字串轉成單個list,再轉成單個tuple,再做tuple累加
        #         else:
        #             import_tuple += tuple([cell.value])   #將字串轉成單個list,再轉成單個tuple,再做tuple累加



# 取得下屬的工作明細
def get_pdca_subs_data(request,pk1=None):
    dataList = []
    if pk1:
        fieldList = ['work_code',
                     'chi_name',
                     'dept',
                     'director',
                     'corp',
                     'factory',
                     'arrival_date',
                     'resign_date',
                     'pos_id',
                     'rank',
                     'bonus_factor',
                     'eval_class',
                     'nat',
                     'factory_area',
                     ]
        results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(direct_supv_id=pk1,pdca_agent=True)

        # print("\n"*2)
        # print("-"*200)
        dataList = []
        for dataTuple in results:
            dept_name = "" if dataTuple[2] == None else UserDefCode.objects.get(id=dataTuple[2]).desc1
            director_name = "" if dataTuple[3] == None else EmployeeInfoEasy.objects.get(work_code=dataTuple[3]).chi_name
            corp_name = "" if dataTuple[4] == None else UserDefCode.objects.get(id=dataTuple[4]).desc1
            factory_name = "" if dataTuple[5] == None else Factory.objects.get(id=dataTuple[5]).name
            pos_name = "" if dataTuple[8] == None else UserDefCode.objects.get(id=dataTuple[8]).desc1
            rank_name = "" if dataTuple[9] == None else UserDefCode.objects.get(id=dataTuple[9]).desc1
            bonus_factor_name = "" if dataTuple[10] == None else UserDefCode.objects.get(id=dataTuple[10]).desc1      # 獎金點數
            eval_class_name = "" if dataTuple[11] == None else UserDefCode.objects.get(id=dataTuple[11]).desc1        # kpi/bsc
            nat_name = "" if dataTuple[12] == None else UserDefCode.objects.get(id=dataTuple[12]).desc1               # 國籍
            factory_area_name = "" if dataTuple[13] == None else UserDefCode.objects.get(id=dataTuple[13]).desc1      # 廠區    2021/08/04 add
            results_a = PdcaMaster.objects.filter(work_code=dataTuple[0]).aggregate(Max('the_time'))
            results_b = PdcaMaster.objects.filter(work_code=dataTuple[0]).aggregate(Max('the_time'))

            # pm406 : 主管審核, 增加『狀態』
            dataList.append({
                'work_code': dataTuple[0],
                'chi_name': dataTuple[1],
                'dept_name': dept_name,
                'director_id': dataTuple[3],
                'director_name': director_name,
                'corp_name': corp_name,
                'factory_name': factory_name,
                'arrival_date': dataTuple[6],
                'resign_date': dataTuple[7],
                'pos_name': pos_name,
                'rank': rank_name,
                'bonus_factor': bonus_factor_name,
                'eval_class': eval_class_name,
                'nat': nat_name,
                'factory_area':factory_area_name,
                'the_time':results_a['the_time__max'] if results_a else "",                 #dcaMaster抓取，抓不到，就是空白
                # 'change_time':results_a['change_time__max'] if results_a else "",              #PdcaMaster抓取，抓不到，就是空白
            })
        # print("-"*200)
        # print("\n"*2)

    return JsonResponse(dataList, safe=False)


#tt406 技能盤點選取＂部門＂使用
def update_session(request,select_dept=None):
    request.session['select_dept'] = select_dept
    return JsonResponse({'success':True})


def update_session_pdca(request,select_work_code=None,select_chi_name=None):
    request.session['select_work_code'] = select_work_code
    request.session['select_chi_name'] = select_chi_name
    return JsonResponse({'success':True})


def download_pdf(request,url):
    file = open(url, 'rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="1234567890_abcdefg.pdf"'
    return response


def set_language_code(request,choice_language=None):
    settings.LANGUAGE_CODE = choice_language
    request.session['choice_language'] = choice_language
    settings.LANGUAGE_CODE = choice_language
    translation.activate(choice_language)
    return JsonResponse({'success':True})



