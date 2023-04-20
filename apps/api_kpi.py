import os
import os
import sys
import sysconfig
import datetime
import time
import json
import urllib.parse
import logging
import pymssql

# import psycopg2
import pyodbc
from django.apps import apps
from django.conf import settings
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.db.models import F, Q , Sum , Max
from django.db import connections,close_old_connections
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect,HttpResponseServerError
from django.shortcuts import render,redirect
from django.utils import timezone
from django.views import View
from django.views.decorators.vary import vary_on_cookie
from django.forms.models import model_to_dict
from sqlalchemy.exc import IntegrityError

from tablib import Dataset

from common.utils import query_serialize
from common.models import UserDefCode,ProcessOptionsTxtDef,RegActRules,FileActionLedger
from system.models import Factory
from apps.kpi.models import (EmployeeInfoEasy,
                             MetricsSetup,
                             MetricsCalc,
                             ScoreSheet,
                             WorkingYM,
                             WorkcodeMapping,
                             DeptSupervisor
                             )


from apps.kpi.resources import (EmployeeInfoEasyResource,
                               # EeAttendSummaryResource,
                               excel_set_style
                                )

from django.core.mail import BadHeaderError,send_mail
import ldap

import openpyxl
from openpyxl.styles import Font, PatternFill, colors, Border, Side, Alignment


logger = logging.getLogger('debug')
tracer = logging.getLogger('trace')

current_tz = timezone.get_current_timezone()
# now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')
now = timezone.now().astimezone(current_tz)

class ModelHandleView(View):
    @vary_on_cookie
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def parse_url(self, request):
        """
        擷取 model 參數，並找尋相對應的 Model class
        model參數會先依底線切開個單字後，做capitalize再合併。
        找不到對應 model class 回傳 404，反之回傳 Model Class 與 pk(如果有)

        ex.1
        model_parm = 'pds_main'
        model_name = PdsMain

        ex.2
        model_parm = 'PDS_main_testMain'
        model_name = PdsMainTestmain

        """
        model = self.kwargs['model']
        pk = self.kwargs.get('pk', None)
        model_name = ''.join([string.capitalize() for string in model.split('_')])
        model_list = apps.get_models()
        try:
            model_class = [m for m in model_list if m.__name__ == model_name][0]
        except IndexError as e:
             logger.debug(e)
             return JsonResponse({"success": False})
        else:
            return model_class, pk

    def get(self, request, *args, **kwargs):
        """
        GET 做兩件事: 讀取 / 刪除

        - 讀取
            <str:model>?sort=xx&order=OO&limit=100&factory=1000&other_parameters
            以 other_parameters 篩選資料之後再處理sorting, ordering, limit 和 factory

        - 刪除
            <str:model>/<str:pk>
            取得 pk 刪除指定的資料。
        """
        model, pk = self.parse_url(request)
        get_param = request.GET.dict()
        tracer.info('get %s' % model._meta.model_name)
        identity_str = request.META['QUERY_STRING']
        if not pk:
            try:
                sort = get_param.pop('sort', None)
            except:
                pass
            try:
                order = get_param.pop('order', None)
            except:
                pass
            try:
                limit = get_param.pop('limit', None)
            except:
                pass
            try:
                latest = get_param.pop('latest', None)
            except:
                pass

            factory = get_param.pop('factory', None)
            cache_str = '%s_%s' % (model._meta.model_name, identity_str)
            query = Q()
            datas = cache.get(cache_str)

            if datas:
                return JsonResponse(datas, safe=False)
            for key, value in get_param.items():
                # _id 結尾通常是 foreignkey, 不支援 __contains
                # key = key + '__contains' if not '_id' in key else key
                if ( value.find(',')==-1 ):     #queryParams單一個,單選
                    query &= Q(**{key: value})
                else:                          #多選
                    value_tuple = tuple(value.split(','))
                    query &= Q(**{'%s%s' % (key, '__in'): value_tuple})
            if factory:
                query &= Q(factory__id=factory)
            if latest:
                three_months = timezone.now() - datetime.timedelta(days=30)
                query &= Q(create_time__gt=three_months)
            data = model.objects.filter(query)
            # TODO: 優化

            if sort:
                order_syntext = '-' if order == 'desc' else ''
                data = data.order_by(order_syntext + sort)
            if limit:
                data = data[:int(limit)]

            dict_data = self._serialize(data, True)
            cache.add(cache_str, dict_data, 60)
            return JsonResponse(dict_data, safe=False)
        else:
            try:
                model.objects.filter(pk=pk).delete()
                cache.clear()
                return JsonResponse({"success": True})
            except:
                return JsonResponse({"success": False, "message": "***有『關聯』資料，不允刪除***"})

    def post(self, request, *args, **kwargs):
        # return JsonResponse({"success": True})
        """
        POST 做兩件事: 新增 / 修改

        - 新增
            <str:model>
            新增資料。

        - 修改
            <str:model>/<str:pk>
            取得 pk 更新指定的資料。
        """
        model, pk = self.parse_url(request)
        post_param = request.POST.dict()
        if post_param.get('csrfmiddlewaretoken', None):
            post_param.pop('csrfmiddlewaretoken')
        # 拿掉 csrfmiddlewaretoken
        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')
        if pk:
            instance = model.objects.filter(pk=pk)
            if instance:
                post_param.update({
                    'change_time': now,
                    # 'changer': request.user.name,
                    'changer': request.user.username,
                })

                ''' Add by liyun at 2020/9/30
                ＊＊＊前端傳回的資料，若有未輸入會傳回〝空字串〞。＊＊＊
                空字串('') : (1)-在日期，數值，邏輯...等資料在存檔時，會出現錯誤，無法存檔。
                                使得每個欄位都〝必要輸入〞資料，這並不符合實際需要。
                            (2)-所以只要是〝空字串〞就指定為None。
                這樣的做法，是將Field是否需要輸入的控制權，交給forms.py( 假設models.py的field 的屬性 blank=True, null=True )。
                '''
                for key, value in post_param.items():
                    if value == '':
                        post_param.update({
                            key: None,
                        })

                try:
                    instance.update(**post_param)
                except ValidationError as e:
                    logger.debug(e)

                    if e.code == 'invalid_time':
                        return JsonResponse({"success": False, "message": "\"%s\" 格式正確，卻非有效時間。" % e.params['value']})
                    else:
                        return JsonResponse({"success": False, "message": "請檢查欄位數值是否接正確。"})
                except ValueError as e:
                    logger.debug(e)
                    return JsonResponse({"success": False, "message": "欄位不能為空，且只能是數字。"})
                else:
                    cache.clear()
                    return JsonResponse({"success": True})
        else:
            post_param.update({
                'change_time': now,
                # 'changer': request.user.name,
                'changer': request.user.username,
                'create_time': now,
                # 'creator': request.user.name,
                'creator': request.user.username,
            })

            '''
            前端傳回的資料，若有未輸入會傳回〝空字串〞，
            空字串('') : 在日期，數值，邏輯...等資料在存檔時，會出現錯誤，無法存檔。
            '''
            for key, value in post_param.items():
                if value == '':
                    post_param.update({
                        key: None,
                    })
            try:
                model.objects.create(**post_param)
            except ValidationError as e:
                logger.debug(e)
                if e.code == 'invalid_time':
                    return JsonResponse({"success": False, "message": "\"%s\" 格式正確，卻非有效時間。" % e.params['value']})
                else:
                    return JsonResponse({"success": False, "message": "請檢查欄位數值是否正確。"})
            except ValueError as e:
                logger.debug(e)
                return JsonResponse({"success": False, "message": "欄位不能為空，且只能是數字。"})
            except pymssql.DatabaseError as e:
                logger.debug(e)
                return JsonResponse({"success": False, "message": "已存在資料庫中"})
            except Exception as e:
                logger.debug(e)
                return JsonResponse({"success": False, "message": "已存在資料庫中"})
            else:
                cache.clear()
                return JsonResponse({"success": True})

    def _serialize(self, obj, related=True):
        return query_serialize(obj, related)



class ModelHandleDistinctView(View):

    @vary_on_cookie
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def parse_url(self, request):
        """
        擷取 model 參數，並找尋相對應的 Model class
        model參數會先依底線切開個單字後，做capitalize再合併。
        找不到對應 model class 回傳 404，反之回傳 Model Class 與 pk(如果有)

        ex.1
        model_parm = 'pds_main'
        model_name = PdsMain

        ex.2
        model_parm = 'PDS_main_testMain'
        model_name = PdsMainTestmain

        """
        model = self.kwargs['model']
        pk = self.kwargs.get('pk', None)
        model_name = ''.join([string.capitalize() for string in model.split('_')])
        model_list = apps.get_models()
        try:
            model_class = [m for m in model_list if m.__name__ == model_name][0]
        except IndexError as e:
             logger.debug(e)
             return JsonResponse({"success": False})
        else:
            return model_class, pk

    def get(self, request, *args, **kwargs):
        """
        GET 做兩件事: 讀取 / 刪除

        - 讀取
            <str:model>?sort=xx&order=OO&limit=100&factory=1000&other_parameters
            以 other_parameters 篩選資料之後再處理sorting, ordering, limit 和 factory

        - 刪除
            <str:model>/<str:pk>
            取得 pk 刪除指定的資料。
        """
        model, pk = self.parse_url(request)
        get_param = request.GET.dict()
        tracer.info('get %s' % model._meta.model_name)
        identity_str = request.META['QUERY_STRING']
        if not pk:
            try:
                sort = get_param.pop('sort', None)
            except:
                pass
            try:
                order = get_param.pop('order', None)
            except:
                pass
            try:
                limit = get_param.pop('limit', None)
            except:
                pass
            try:
                latest = get_param.pop('latest', None)
            except:
                pass

            factory = get_param.pop('factory', None)
            cache_str = '%s_%s' % (model._meta.model_name, identity_str)
            query = Q()
            datas = cache.get(cache_str)

            if datas:
                return JsonResponse(datas, safe=False)
            for key, value in get_param.items():
                # _id 結尾通常是 foreignkey, 不支援 __contains
                # key = key + '__contains' if not '_id' in key else key
                query &= Q(**{key: value})
            if factory:
                query &= Q(factory__id=factory)
            if latest:
                three_months = timezone.now() - datetime.timedelta(days=30)
                query &= Q(create_time__gt=three_months)

            data = model.objects.filter(query).distinct()
            # TODO: 優化

            if sort:
                order_syntext = '-' if order == 'desc' else ''
                data = data.order_by(order_syntext + sort)
            if limit:
                data = data[:int(limit)]

            dict_data = self._serialize(data, True)
            cache.add(cache_str, dict_data, 60)
            return JsonResponse(dict_data, safe=False)
        else:
            try:
                model.objects.filter(pk=pk).delete()
                cache.clear()
                return JsonResponse({"success": True})
            except:
                return JsonResponse({"success": False, "message": "***有『關聯』資料，不允刪除***"})

    def _serialize(self, obj, related=True):
        return query_serialize(obj, related)



def cmpx_search_handle(request, model):
    """
    多條件查詢
    0313
    """
    format = request.GET.get('format', None)
    factory = request.GET.get('factory', None)
    def relation(operator):
        mapping = {
            '>': '__gt',
            '<': '__lt',
            '=': '',
            'include': '__contains',
            '!=': '',
            'exclude': '__contains'
        }
        return mapping[operator]

    model_name = ''.join([string.capitalize() for string in model.split('_')])
    model_list = apps.get_models()
    try:
        # tracer.info('cpx, %s' % model_name)
        model_class = [m for m in model_list if m.__name__ == model_name][0]
    except IndexError as e:
        # logger.debug('Failed on cpx: %s' % e)
        return JsonResponse({"status": False})
    else:
        post_param = request.POST
        data = json.loads(post_param['data'])
        logic = data.get('logic', 'all')
        filters = data.get('filter')

        query = Q()
        if logic == 'all':
            for key, item in filters.items():
                if item['relation'] in ['exclude', '!=']:
                    query &= ~Q(**{'%s%s' % (item['field'], relation(item['relation'])): item['value']})
                else:
                    query &= Q(**{'%s%s' % (item['field'], relation(item['relation'])): item['value']})
        elif logic == 'any':
            for key, item in filters.items():
                if item['relation'] in ['exclude', '!=']:
                    query |= ~Q(**{'%s%s' % (item['field'], relation(item['relation'])): item['value']})
                else:
                    query |= Q(**{'%s%s' % (item['field'], relation(item['relation'])): item['value']})
        if format:
            query &= Q(**{'_format':format})
        if factory:
            query &= Q(**{'factory_id':factory})
        try:
            result = model_class.objects.filter(query)
        except Exception as e:
            # logger.debug('Failed on cpx: %s' % e)
            return JsonResponse({"status": False})
        return JsonResponse(query_serialize(result, True), safe=False)


def employee_info_easy_handle(View):
    pass


def score_sheet_handle(request, pk=None):
    if request.method == 'POST':
        pass
    else:
        if pk:
            pass
        else:
            pass


#讀取剛剛修改或新增的employee data
def get_add_update_employee_data(request,pk=None):

    if pk==None:
        return JsonResponse({"data":"","success":True})
    else:
        results = EmployeeInfoEasy.objects.filter(work_code=pk)
        for itm in results:
            r = model_to_dict(itm)
        # dataDict={
        #     'work_code': r.get('work_code'),
        #     'chi_name': r.get('chi_name'),
        #     'direct_supv': r.get('direct_supv')+" "+EmployeeInfoEasy.objects.get(work_code=r.get('direct_supv')).chi_name,
        #     'director': r.get('director')+" "+EmployeeInfoEasy.objects.get(work_code=r.get('director')).chi_name,
        #     'factory': Factory.objects.get(id=r.get('factory')).name,
        #     'dept_flevel': UserDefCode.objects.get(id=r.get('dept_flevel')).desc1,
        #     'dept': UserDefCode.objects.get(id=r.get('dept')).desc1,
        #     'pos': UserDefCode.objects.get(id=r.get('pos')).desc1,
        #     'nat': UserDefCode.objects.get(id=r.get('nat')).desc1,
        #     'rank': UserDefCode.objects.get(id=r.get('rank')).desc1,
        #     'eval_class': UserDefCode.objects.get(id=r.get('eval_class')).desc1,
        #     'kpi_diy': r.get('kpi_diy'),
        #     'bonus_type': UserDefCode.objects.get(id=r.get('bonus_type')).desc1,
        #     'bonus_factor': UserDefCode.objects.get(id=r.get('bonus_factor')).desc1,
        #     'factory_area': UserDefCode.objects.get(id=r.get('factory_area')).desc1,
        #     'email': r.get('email'),
        #     'arrival_date': r.get('arrival_date'),
        #     'resign_date': r.get('resign_date'),
        #     'trans_date': r.get('trans_date'),
        #     'trans_type': r.get('trans_type'),
        #     'labor_type': UserDefCode.objects.get(id=r.get('labor_type')).desc1,
        #     'service_status': UserDefCode.objects.get(id=r.get('service_status')).desc1,
        #     'dept_description': r.get('dept_description'),
        # }
        # return JsonResponse({"data":dataDict,"success":True})
        return JsonResponse({"data":r,"success":True})


def ee_attend_summary_handle(request, pk=None):
    if request.method == 'POST':
        pass
    else:
        if pk:
            pass
        else:
            pass


def metrics_setup_handle(request, pk=None):
    if request.method == 'POST':
        pass
    else:
        if pk:
            pass
        else:
            pass



def metrics_calc_handle(request, pk=None):
    if request.method == 'POST':
        pass
    else:
        if pk:
            pass
        else:
            pass


def get_metrics_calc(request,pk=None):
    fields = ['order_number',
              'calc_content',
              'lower_limit',
              'upper_limit',
              'score']
    if pk:
        results = MetricsCalc.objects.values_list(*fields).filter(metrics_id=pk)
        all_list = []
        if results:
            for T in results:
                all_list.append({
                    'order_number': T[0],
                    'calc_content': T[1],
                    'lower_limit':T[2],
                    'upper_limit':T[3],
                    'score' : T[4],
                })
    return JsonResponse(all_list,safe=False)


def get_calc_score(request,pk=None):
    '''
    #方法一 : 這種方法的好處是方便控制最終返回字典值的格式
    obj = MetricsCalc.objects.get(id=pk)
    myDict = {
        'id':obj.id,
        'calc_content':obj.calc_content,
        'score':obj.score,
        'efDate':obj.ef_date.strftime('%Y-%m-%d'),
        'expDate': obj.exp_date.strftime('%Y-%m-%d')
    }

    #方法二 : 直接轉dict
    obj = MetricsCalc.objects.get(id=pk).__dict__
    for k,v in obj.items():
        print(k," --> ",v)
     '''
    #方法三(此方法, 無法轉-->editable=False，有auto_now_add=True,auto_now=True屬性的field)
    myCalc = model_to_dict( MetricsCalc.objects.get(id=pk) , fields=['calc_content','score'] )
    return JsonResponse(myCalc)



#員工基本資料的匯出,2020/10/20目前未使用
def employee_export(request,pk=None):
    employee_resource = EmployeeInfoEasyResource()
    dataset = employee_resource.export()
    response = HttpResponse( dataset.xls , content_type = 'application/vnd.ms-excel' )
    response['Content-Disposition'] = 'attachment; filename="persons.csv"'
    return response


#員工基本資料的匯入
def employee_import(request):
    if request.method == 'POST':
        # EmployeeInfoEasy.objects
        employee_resource = EmployeeInfoEasyResource()
        new_employees = request.FILES['myfile']
        prepared_data = Dataset().load(new_employees.read(), format='xlsx' ,headers= True )    #headers=True...可使用string做為欄名
        err_message = "匯入資料錯誤"
        # try:
        result = employee_resource.import_data( prepared_data , dry_run=True, raise_errors = True) #dry_run=True　沒有正式匯入，測試資料匯入是否有error
        # except:
        #     # 資料有誤，顯示錯誤訊息
        #     cache.clear()
        #     return render(request,'kpi/import_error.html',context={"success": False, "message":err_message})
        # else:
        #     if result.has_errors():
        #         # 匯入有誤，顯示錯誤訊息
        #         cache.clear()      #清除cache, 好讓〝程式〞重新取得最新匯入的資料
        #         return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
        #     else:
        #         result = employee_resource.import_data( prepared_data, dry_run=False)  # 實際將資料匯入
        #         if result.has_errors():
        #             cache.clear()
        #             return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
        #         else:
        #             # 匯入成功,返回上一頁
        #             cache.clear()
        #             return HttpResponseRedirect(request.META.get('HTTP_REFERER','/'),{"success": True})


#將文字資料,轉成對應的id
def employee_import_data_transfer(excelTuple):
    # 將中文或CODE轉換成對應的id--------------------------------------------------------------------------------------------------------Begin
    # 2 factory.name
    alist = [excelTuple[0],excelTuple[1]]
    try:
        alist.append( Factory.objects.get(name=excelTuple[2]).id )
    except:
        print(" 公司:",excelTuple[2],end=" ; ")
        alist.append("ZZZZZZ")

    # 3 nat_id (用國籍編號)
    try:
        alist.append( UserDefCode.objects.get(topic_code_id="nat_id",udc=excelTuple[3]).id )
    except:
        print(" 國家:", excelTuple[3], end=" ; ")
        alist.append(999999999)

    # 4 service_status_id
    try:
        alist.append( UserDefCode.objects.get(topic_code_id="service_status_id",desc1=excelTuple[4]).id )
    except:
        print(" 服務狀態:", excelTuple[4], end=" ; ")
        alist.append(999999999)

    # 5 factory_area_id
    try:
        alist.append( UserDefCode.objects.get(topic_code_id="factory_area_id",desc1=excelTuple[5]).id )
    except:
        print(" 廠區:", excelTuple[5], end=" ; ")
        alist.append(999999999)

    # 6 emial : 不變(依使用者輸入)
    alist.append(excelTuple[6])

    # 7 kpi_diy
    alist.append( True if excelTuple[14]=="間接" else False )

    # 8 dept_flevel_id (用一級部門編號)
    try:
        alist.append( UserDefCode.objects.get(topic_code_id="dept_flevel_id",udc=excelTuple[8].strip()).id )
    except:
        print("一級部門:",excelTuple[8], end=" ; ")
        alist.append(999999999)

    # 9 dept_id (用部門編號)
    try:
        alist.append( UserDefCode.objects.get(topic_code_id="dept_id",udc=excelTuple[9].strip()).id )
    except:
        print("部門別:", excelTuple[9], end=" ; ")
        alist.append( 999999999 )

    #10,11,12 :  dept_description, director_id,direct_supv_id : 不變(依使用者輸入)
    alist.extend([excelTuple[10],excelTuple[11],excelTuple[12]])
    #13 pos_id
    try:
        alist.append( UserDefCode.objects.get(topic_code_id="pos_id",desc1=excelTuple[13]).id )
    except:
        print("職位:", excelTuple[13], end=" ; ")
        alist.append( 999999999 )

    #14 labor_type_id
    try:
        alist.append( UserDefCode.objects.get(topic_code_id="labor_type_id",desc1=excelTuple[14]).id )
    except:
        print("直/間接:", excelTuple[14], end=" ; ")
        alist.append( 999999999 )

    #15,16,17,18 :  arrival_date,resign_date,trans_date,trans_type 不變(依使用者輸入)
    alist.extend( [excelTuple[15], excelTuple[16], excelTuple[17], excelTuple[18]] )

    #19 bonus_type_id
    try:
        alist.append( UserDefCode.objects.get(topic_code_id="bonus_type_id",desc1=excelTuple[19].strip()).id )
    except:
        print("獎金類別:",excelTuple[19], end=" ; ")
        alist.append( 999999999 )

    #20 bonus_factor_id
    try:
        alist.append( UserDefCode.objects.get(topic_code_id="bonus_factor_id",desc1=excelTuple[20]).id )
    except:
        print("獎金點數:", excelTuple[20], end=" ; ")
        alist.append( 999999999 )

    #21 rank_id
    try:
        alist.append( UserDefCode.objects.get(topic_code_id="rank_id",desc1=excelTuple[21]).id )
    except:
        print("職等:", excelTuple[21], end=" ; ")
        alist.append( 999999999 )

    #22 eval_class_id : BSC/KPI
    if excelTuple[22] in ["KPI","BSC"]:
        alist.append(UserDefCode.objects.get(topic_code_id="eval_class_id", desc1=excelTuple[22]).id)
    else:
        # print("KPI/BSC:", excelTuple[22], end=" ; ")
        bsc_kpi = "BSC" if int(excelTuple[21]) < 8 else "KPI"
        alist.append( UserDefCode.objects.get(topic_code_id="eval_class_id",desc1=bsc_kpi).id )
    print()
    # 將中文或CODE轉換成對應的id--------------------------------------------------------------------------------------------------------Ending
    return tuple(alist)


#員工基本資料的匯入
def employee_import_update(request):
    if request.method == 'POST':
        try:
            fn = request.FILES['myfile']
        except:
            return render(request, 'kpi/import_error.html', context={"success": False, "message": "沒有選擇檔案"})
        workbook = openpyxl.load_workbook(fn)

        # 獲得所有sheet的名稱
        sheets = workbook.get_sheet_names()
        # 獲得當前正在顯示的sheet
        sheet = workbook.active

        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')
        now_str = timezone.now().astimezone(current_tz).strftime('%Y%m%d_%H%M%S')
        filename = 'HR2PMS_UPDATE_ADD_' + now_str + "_" + request.user.username + '.xlsx'

        import_list = []
        for i in range(2, sheet.max_row):
            aTuple = list(sheet.rows)[i]
            excelTuple = ()
            for cell in aTuple:
                excelTuple += tuple([cell.value])
            aTuple = employee_import_data_transfer(excelTuple)         #將文字資料,轉成對應的id
            instance=None
            import_tuple = ()
            update_tuple = ()
            update_list = []
            update_param = {}
            import_param = {}
            for idx,value in enumerate(aTuple):
                # print(idx,":",value,end=" , ")
                if idx==0:  #第一格是工號
                    instance = EmployeeInfoEasy.objects.filter(work_code=value)
                if instance:
                    update_tuple += tuple([value])   #將字串轉成單個list,再轉成單個tuple,再做tuple累加
                else:
                    import_tuple += tuple([value])   #將字串轉成單個list,再轉成單個tuple,再做tuple累加
            if instance:
                update_param.update({
                    'change_time':now,
                    'changer': request.user.username,
                    'creator': request.user.username,
                    'corp_id': None,
                    'work_code': update_tuple[0],
                    'chi_name' : update_tuple[1],
                    'factory_id': update_tuple[2],
                    'nat_id': update_tuple[3],
                    'service_status_id': update_tuple[4],
                    'factory_area_id': update_tuple[5],
                    'email': update_tuple[6],
                    'kpi_diy': True if update_tuple[7] == 'Y' else False,
                    'dept_flevel_id': update_tuple[8],
                    'dept_id': update_tuple[9],
                    'dept_description': update_tuple[10],
                    'director_id': update_tuple[11],
                    'direct_supv_id': update_tuple[12],
                    'pos_id': update_tuple[13],
                    'labor_type_id': update_tuple[14],
                    'arrival_date' : update_tuple[15],
                    'resign_date' : update_tuple[16],
                    'trans_date' : update_tuple[17],
                    'trans_type' : update_tuple[18],
                    'bonus_type_id': update_tuple[19],
                    'bonus_factor_id': update_tuple[20],
                    'rank_id': update_tuple[21],
                    'eval_class_id':update_tuple[22],
                    'urrf':request.user.username+" "+str(request.user),
                    'urrf1':now_str+"比對HR，匯入，修改",
                    'urrf2':str(fn),
                })
                try:
                    instance.update(**update_param)
                except ValidationError as e:
                    logger.debug(e)
                    if e.code == 'invalid_time':
                        err_message = "\"%s\" 格式正確，卻非有效時間。   工號:%s  姓名:%s" % e.params['value'] % update_tuple[0] % update_tuple[1]
                        return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
                    else:
                        err_message = "請檢查欄位數值是否接正確。 工號:%s  姓名:%s "  % update_tuple[0] % update_tuple[1]
                        return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
                except ValueError as e:
                    logger.debug(e)
                    err_message ="欄位不能為空，且只能是數字。 工號:%s  姓名:%s "  % update_tuple[0] % update_tuple[1]
                    return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
                except Exception as e:
                    logger.debug(e)
                    error_value = str(e.args)
                    return render(request, 'kpi/import_data_error.html',
                                  context={"success": False, "message1": "工號:"+update_tuple[0], "message2": "姓名:"+update_tuple[1],
                                           "message3": error_value})
                else:
                    cache.clear()
                    # err_message = "匯入成功"
                    # return render(request, 'kpi/import_success.html', context={"success": False, "message": err_message})
            else:
                import_param.update({
                    'create_time': now,
                    'change_time': now,
                    'changer': request.user.username,
                    'creator': request.user.username,
                    'corp_id': None,
                    'work_code': import_tuple[0],
                    'chi_name': import_tuple[1],
                    'factory_id': import_tuple[2],
                    'nat_id': import_tuple[3],
                    'service_status_id': import_tuple[4],
                    'factory_area_id': import_tuple[5],
                    'email': import_tuple[6],
                    'kpi_diy': True if import_tuple[7] == 'Y' else False,
                    'dept_flevel_id': import_tuple[8],
                    'dept_id': import_tuple[9],
                    'dept_description': import_tuple[10],
                    'director_id': import_tuple[11],
                    'direct_supv_id': import_tuple[12],
                    'pos_id': import_tuple[13],
                    'labor_type_id': import_tuple[14],
                    'arrival_date': import_tuple[15],
                    'resign_date': import_tuple[16],
                    'trans_date': import_tuple[17],
                    'trans_type': import_tuple[18],
                    'bonus_type_id': import_tuple[19],
                    'bonus_factor_id': import_tuple[20],
                    'rank_id': import_tuple[21],
                    'eval_class_id': import_tuple[22],
                    'urrf': request.user.username+" "+str(request.user),
                    'urrf1': now_str+"比對HR，匯入，修改",
                    'urrf2': str(fn),
                })
                try:
                    EmployeeInfoEasy.objects.create(**import_param)
                except ValidationError as e:
                    logger.debug(e)
                    if e.code == 'invalid_time':
                        err_message = "\"%s\" 格式正確，卻非有效時間。 工號:%s  姓名:%s" % e.params['value']  % import_tuple[0] % import_tuple[1]
                        return render(request, 'kpi/import_error.html',
                                      context={"success": False, "message": err_message})
                    else:
                        err_message = "請檢查欄位數值是否接正確。 工號:%s  姓名:%s "  % import_tuple[0] % import_tuple[1]
                        return render(request, 'kpi/import_error.html',
                                      context={"success": False, "message": err_message})
                except ValueError as e:
                    logger.debug(e)
                    err_message = "欄位不能為空，且只能是數字。 工號:%s  姓名:%s "  % import_tuple[0] % import_tuple[1]
                    return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
                except Exception as e:
                    logger.debug(e)
                    error_value = str(e.args)
                    return render(request, 'kpi/import_data_error.html',
                                  context={"success": False, "message1": "工號:"+import_tuple[0], "message2": "姓名:"+import_tuple[1],
                                           "message3": error_value})
                else:
                    cache.clear()
            # 將定義好的的sheet, 寫入檔案
            filename = 'HR異動資料匯入_' + now_str + "_" + request.user.username+request.user.name + '.xlsx'
            output_file = '%s/reports/xlsx/tt002/import/%s' % (settings.MEDIA_ROOT, filename)
            workbook.save(output_file)
    err_message = "匯入成功"
    return render(request, 'kpi/import_success.html', context={"success": True, "message": err_message})


''''
#出勤彙總資料的匯入
def ee_attend_summary_import(request):
    if request.method == 'POST':
        ee_attend_resource = EeAttendSummaryResource()
        new_ee_attend = request.FILES['myfile']
        prepared_data = Dataset().load(new_ee_attend.read(), format='xlsx',headers=True)  # headers=True...可使用string做為欄名
        err_message = "匯入資料錯誤"
        try:
            result = ee_attend_resource.import_data(prepared_data, dry_run=True,raise_errors=True)  # dry_run=True　沒有正式匯入，測試資料匯入是否有error
        except:
            # 資料有誤，顯示錯誤訊息
            cache.clear()
            return render(request,'kpi/import_error.html',context={"success": False, "message":err_message})
        else:
            if result.has_errors():
                # 匯入有誤，顯示錯誤訊息
                cache.clear()      #清除cache, 好讓〝程式〞重新取得最新匯入的資料
                return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
            else:
                result = ee_attend_resource.import_data(prepared_data, dry_run=False)  # 實際將資料匯入
                if result.has_errors():
                    cache.clear()
                    return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
                else:
                    # 匯入成功,返回上一頁
                    cache.clear()
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER','/'),{"success": True})
'''


def get_factory_employee_data(request,pk1=None,pk2=None):
    factory = pk1.split('-')
    commQ = Q(eval_class__in=UserDefCode.objects.filter(Q(topic_code_id='eval_class_id'),~Q(desc1='BSC')))  # 2021/07/14 ADD 拿掉BSC的人
    dept_id = UserDefCode.objects.get(topic_code_id='dept_id',udc=pk2).id
    results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(Q(factory_id=factory[0]),Q(dept_id=dept_id),commQ)         # 排除'-000'共同指標, '-100'出勤指標

    dataList = []
    for dataTuple in results:
        dataList.append({
            'value': dataTuple[0],
            'text': dataTuple[0]+" "+dataTuple[1],
        })
    return JsonResponse(dataList,safe=False)




def get_employee_data(request,pk=None):
    commQ = Q(eval_class__in=UserDefCode.objects.filter(Q(topic_code_id='eval_class_id'), ~Q(desc1='BSC')))   #2021/07/14 ADD 拿掉BSC的人
    commQ1 = ~Q(work_code__endswith='-000')   # '-000'共同指標, '-100'出勤指標
    commQ2 = ~Q(work_code__endswith='-100')
    commQ3 = Q(director=pk)
    # results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(~Q(work_code__endswith='-000'),~Q(work_code__endswith='-100'))  # '-000'共同指標, '-100'出勤指標
    if (pk==None):
        # results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(commQ1,commQ2,commQ)         # 排除'-000'共同指標, '-100'出勤指標
        results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(commQ1,commQ2)         # 排除'-000'共同指標, '-100'出勤指標
    else:
        if pk.find('_expand')==-1:      #沒找到
            if pk.find('_copy') == -1:  #沒找到
                    results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(commQ1,commQ2,commQ3,commQ)  # 只取得下屬的資料
            else:              #有找到copy
                pk_list = pk.split('_')
                # pk_list[0]工號  pk_list[1]=copy   pk_list[2]年度   pk_list[3]月份
                return JsonResponse(employee_have_MetricsCalc(pk_list[0], pk_list[1], pk_list[2],pk_list[3]), safe=False)
        else:            #找到expand
            pk_list = pk.split('_')
            # pk_list[0]工號  pk_list[1]=expand   pk_list[2]年度   pk_list[3]月份=0
            return JsonResponse( employee_have_MetricsCalc(pk_list[0],pk_list[1],pk_list[2]), safe=False)


    dataList = []
    for dataTuple in results:
        dataList.append({
            'value': dataTuple[0],
            'text': dataTuple[0]+" "+dataTuple[1],
        })
    return JsonResponse(dataList,safe=False)


#有配分,有衡量指標,才出現在expand的dialog的工號下拉選單
def employee_have_MetricsCalc(pk1=None,pk2=None,pk3=None,pk4=None):
    # pk1:work_code  pk2:copy/expand/valid   pk3:date_yyyy   pk4:date_mm
    commQ = Q(eval_class__in=UserDefCode.objects.filter(Q(topic_code_id='eval_class_id'),~Q(desc1='BSC')))  # 2021/07/14 ADD 拿掉BSC的人
    results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(Q(director=pk1),commQ)
    dataList = []
    have_calc = False
    if (results):   #這個主管有下屬
        for T in results:
            if pk2=='copy':
                commQ = (Q(work_code=T[0]) & Q(date_yyyy=pk3) & Q(date_mm=pk4))
            elif pk2=='expand':
                commQ = (Q(work_code=T[0]) & Q(date_yyyy=pk3) & Q(date_mm=0))
            resultsX = MetricsSetup.objects.order_by('work_code', 'date_yyyy', 'date_mm'). \
                filter(commQ,order_number__lt=900).values_list('work_code', 'date_yyyy', 'date_mm'). \
                annotate(Sum('allocation'))      #跟據 work_code,date_yyyy,date_mm累加衡量指標的配分
            if (resultsX): #這個主管的1個下屬有建衡量指標
                have_calc = False               #是否有得分0,得分最高配分
                score_less_allocation = False  #得分大於最高配分
                # have_calc_seq = False         #依照遞增.遞減做計算方式
                for TX in resultsX:
                    if (TX[3]==100):      #配分等於100分,再去檢查有沒有計算方式
                        # resultsY = MetricsSetup.objects.filter(commQ, order_number__lt=900).values_list('metrics_id',flat=True)    #整個月的metrics_id
                        resultsY = MetricsSetup.objects.filter(commQ, order_number__lt=900).values_list('metrics_id','allocation','score_type')    #整個月的metrics_id
                        for TZ in resultsY:      #這個月的每一個metrics_id
                            # 有沒有0配分? 有沒有最高配分?
                            score0X_count=MetricsCalc.objects.filter(metrics_id=TZ[0],score__in=(0,TZ[1])).count()
                            scoreMax = MetricsCalc.objects.values_list('score', flat=True).filter(metrics_id=TZ[0]).annotate(Max('score'))
                            allocation = MetricsSetup.objects.get(metrics_id=TZ[0]).allocation #取出最高配分
                            if (score0X_count==2):               #這個月的每個metrics_id, 有沒有計算方式
                                # 有0配分 也有最高配分
                                # 檢查計算方式,是否有依順序高底設置(score_type)
                                have_calc = True
                            else:
                                have_calc = False                   #只要有一個衡量指標沒有計算方式就離開
                                break

                            if (scoreMax[0]<=allocation):           #評算方式的得分的最高分, 小於或等於...衡量指標旳配分
                                score_less_allocation = True
                            else:
                                score_less_allocation = False
                                break

                        if have_calc and score_less_allocation:     #計算方式得分有(0,最高配分), 得分未高於最高配分...可列為工號的選擇
                            have_calc = False
                            dataList.append({
                                'value': T[0],
                                'text': T[0] + " " + T[1],
                            })
    return dataList



def valid_my_metrics(request,pk1,pk2,pk3):
    # pk1=work_code   pk2=date_yyyy       pk3=date_mm
    # use by pm402.js line 530 檢查是否
    # ＊＊＊檢查衡量指標，未符合下列，就不可評分＊＊＊
    # 　   衡量指標該月"必定"滿100分
    #     每一項至少有二筆"計算方式"(得分0 / 最高配分)
    #     每一項計算方式"的得分不可高於最高配分
    pass





# def get_employee_data_factory(request,pk=None):
def get_employee_data_factory(request):
    # if (pk==None):
    results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(~Q(work_code__contains='-'))         # 排除'-000'共同指標, '-100'出勤指標
    # else:
    #     results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(factory_id=pk)         # 依公司來取得資料

    dataList = []
    for dataTuple in results:
        value_str = dataTuple[0] if dataTuple[0] else ''
        text_str = dataTuple[1] if dataTuple[1] else ''
        dataList.append({
            'value': value_str,
            'text': value_str+" "+text_str,
        })
    return JsonResponse(dataList,safe=False)



def get_dept_data_by_factory(request,pk1=None,pk2=None):   #pk1 : 使用udc的欄位, pk2:facotry
    commQ = Q(id__in=EmployeeInfoEasy.objects.values_list(pk1, flat=True).filter(~Q(work_code__contains='-'),factory_id=pk2).order_by(pk1).distinct(pk1))
    if ( pk2==None ):
        commQ = Q(id__in=EmployeeInfoEasy.objects.values_list(pk1, flat=True).filter(~Q(work_code__contains='-')).order_by(pk1).distinct(pk1))

    fieldList = [ 'id',
                  'udc',
                  'desc1',
                  'desc2',
                 ]
    print("\n"*2)
    print("="*200)
    print(pk1,pk2)
    print(commQ)
    print("="*200)
    print("\n"*2)
    results = UserDefCode.objects.filter(commQ).values_list(*fieldList).order_by('udc')
    dataList = []
    for dataTuple in results:
        dataList.append({
            'value': dataTuple[0],
            'text': str(dataTuple[1])+" "+dataTuple[2],
        })
    return JsonResponse(dataList,safe=False)



def get_manager_data(request,pk=None):
    commQ = Q(work_code__in=EmployeeInfoEasy.objects.values_list(pk, flat=True).filter(~Q(work_code__contains='-')).order_by(pk).distinct(pk) )
    results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(commQ)  # 排除'-000'共同指標, '-100'出勤指標

    dataList = []
    for dataTuple in results:
        dataList.append({
            'value': dataTuple[0],
            'text': dataTuple[0]+" "+dataTuple[1],
        })
    return JsonResponse(dataList,safe=False)


def get_director_data_factory(request):
    # if (pk==None):
    commQ = Q(work_code__in=EmployeeInfoEasy.objects.values_list('director_id', flat=True).filter(~Q(work_code__contains='-')).order_by('director_id').distinct('director_id'))
    results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(commQ)  # 排除'-000'共同指標, '-100'出勤指標


    dataList = []
    for dataTuple in results:
        dataList.append({
            'value': dataTuple[0],
            'text': dataTuple[0]+" "+dataTuple[1],
        })
    return JsonResponse(dataList,safe=False)



def get_employee_common_data(request,pk1=None,pk2=None):
    results = None
    if (pk1 == None and pk2 == None):
        results = EmployeeInfoEasy.objects.values_list('work_code', 'chi_name').filter(work_code__endswith='-000')         # 排除'-000'共同指標, '-100'出勤指標
    else:
        commQ = Q(work_code__in=MetricsSetup.objects.filter( Q(date_yyyy=pk1) & Q(date_mm=pk2)).values_list('work_code', flat=True).order_by('work_code','date_yyyy','date_mm'))
        results = EmployeeInfoEasy.objects.values_list('work_code','chi_name').filter(commQ,Q(work_code__endswith='-000'))
    dataList = []
    for dataTuple in results:
        dataList.append({
            'value': dataTuple[0],
            'text': dataTuple[0]+" "+dataTuple[1],
        })
    return JsonResponse(dataList,safe=False)



def get_dept_data(request,pk1=None):
    if pk1==None:
        results = UserDefCode.objects.values_list('udc', 'desc1').filter(topic_code_id='dept_id')
    else:
        factory = pk1.split('-')
        resultAA=EmployeeInfoEasy.objects.values_list('dept_id').filter(factory_id=factory[0]).order_by('factory_id','dept_id').distinct('factory_id', 'dept_id')
        commQ = Q(id__in=EmployeeInfoEasy.objects.values_list('dept_id').filter(factory_id=factory[0]).order_by('factory_id','dept_id').distinct('factory_id','dept_id'))
        results = UserDefCode.objects.values_list('udc', 'desc1').filter(commQ)
    dataList = []
    for dataTuple in results:
        dataList.append({
            'value': dataTuple[0],
            'text': dataTuple[0]+" "+dataTuple[1],
        })
    return JsonResponse(dataList,safe=False)


#查詢衡量指標
def get_metrics_setup_data(request,pk1=None,pk2=None,pk3=None,pk4=None):
    fieldList = ['work_code',
                 'chi_name',
                 'dept',
                 'director',
                 'corp',
                 'factory',
                 'arrival_date',
                 'resign_date',
                 'pos_id',
                 'rank',
                 'bonus_factor',
                 'eval_class',
                 'nat']
    # flat=True，返回的結果為單個值，不是Tuple ; distinct前，給予要取的值做order_by
    commQ = Q(work_code__in=MetricsSetup.objects.values_list('work_code', flat=True).order_by('work_code').distinct())
    commQ_pk1 = Q(work_code__contains=pk1)
    commQ_pk2 = Q(chi_name__contains=pk2)
    commQ_director = Q(director=pk4)    #2021/02/24增加,過濾這個工號(work_code_key)是否在主管欄位的工號

    if pk3!='x' and pk3!=None:
        # pk3 : dept(部門)，ForeignKey
        # _fk = UserDefCode.objects.filter(desc1__contains=pk3)
        # commQ_pk3_fk = _fk
        commQ_pk3_fk = Q(dept__in=UserDefCode.objects.filter(desc1__contains=pk3))

    # Case 0
    if (pk1==None and pk2==None and pk3==None) or (pk1=='x' and pk2=='x' and pk3=='x'):
        # view呼叫:什麼都沒輸入 pk(1,2,3)==None   搜尋呼叫:什麼都沒輸入 pk(1,2,3)=='x'
        # Use a list inside values_list : add '*'( * : call with arguments unpacked from a list )
        if pk4=="all":
            commQ2 = ~Q(work_code__endswith='-000')
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ,commQ2)
        else:
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_director,commQ)
        # 模糊/交叉...搜尋-------------------------------------------------------------------------------------------------------------------------------------------Begin
        # Case 1
    elif pk1!='x' and pk2=='x' and pk3=='x':
        # pk1 : wrok_code(只輸入工號)
        if pk4 == "all":
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_pk1,commQ)
        else:
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_director,commQ_pk1,commQ)
        # Case 2
    elif pk1=='x' and pk2!='x' and pk3=='x':
        # pk2 : chi_name(只輸入姓名)
        if pk4 == "all":
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_pk2,commQ)
        else:
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_director,commQ_pk2,commQ)
        # Case 3
    elif pk1=='x' and pk2=='x' and pk3!='x':
        # pk3 : dept(只輸入部門)，ForeignKey
        if pk4 == "all":
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_pk3_fk,commQ)
        else:
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_director,commQ_pk3_fk,commQ)
        # Case 4
    elif pk1=='x' and pk2!='x' and pk3!='x':
        # pk1 : work_code(只有工號 "沒" 輸入)   pk3 : dept(部門)，ForeignKey
        if pk4 == "all":
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_pk2,commQ_pk3_fk,commQ)
        else:
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_director,commQ_pk2,commQ_pk3_fk,commQ)
        # Case 5
    elif pk1!='x'and pk2=='x' and pk3!='x':
        # pk2 : chi_name(只有姓名 "沒" 輸入)    pk3 : dept(部門)，ForeignKey
        if pk4 == "all":
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_pk1,commQ_pk3_fk,commQ)
        else:
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_director,commQ_pk1,commQ_pk3_fk,commQ)
        # Case 6
    elif pk1!='x' and pk2!='x' and pk3=='x':
        # pk3 : dept(只有部門 "沒" 輸入)
        if pk4 == "all":
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_pk1,commQ_pk2,commQ)
        else:
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_director,commQ_pk1,commQ_pk2,commQ)
        # Case 7
    elif pk1!='x' and pk2!='x' and pk3!='x':
        # pk(1,2,3) : work_code,chi_name,dept(全都輸入了)   pk3 : dept(部門)，ForeignKey
        if pk4 == "all":
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_pk1,commQ_pk2,commQ_pk3_fk,commQ)
        else:
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_director,commQ_pk1,commQ_pk2,commQ_pk3_fk,commQ)

    # 模糊/交叉...搜尋------------------------------------------------------------------------------------------------------------------------------------------------Ending


    dataList = []
    for dataTuple in results:
        if dataTuple[2]==None:
            # 避免空值，無法往回找到對應的id
            dept_name=""
        else:
            dept_name=UserDefCode.objects.get(id=dataTuple[2]).desc1

        if dataTuple[3]==None:
            # 避免空值，無法往回找到對應的id
            director_name=""
        else:
            director_name=EmployeeInfoEasy.objects.get(work_code=dataTuple[3]).chi_name

        if dataTuple[4]==None:
            # 避免空值，無法往回找到對應的id
            corp_name=""
        else:
            corp_name=UserDefCode.objects.get(id=dataTuple[4]).desc1

        if dataTuple[5]==None:
            # 避免空值，無法往回找到對應的id
            factory_name=""
        else:
            factory_name=Factory.objects.get(id=dataTuple[5]).name

        if dataTuple[8]==None:
            # 避免空值，無法往回找到對應的id
            pos_name=""
        else:
            pos_name=UserDefCode.objects.get(id=dataTuple[8]).desc1

        if dataTuple[9]==None:
            # 避免空值，無法往回找到對應的id
            rank_name=""
        else:
            rank_name=UserDefCode.objects.get(id=dataTuple[9]).desc1

        if dataTuple[10]==None:
            # 避免空值，無法往回找到對應的id
            bonus_factor_name=""
        else:
            bonus_factor_name=UserDefCode.objects.get(id=dataTuple[10]).desc1

        if dataTuple[11]==None:
            # 避免空值，無法往回找到對應的id
            eval_class_name=""
        else:
            eval_class_name=UserDefCode.objects.get(id=dataTuple[11]).desc1

        if dataTuple[12]==None:
            # 避免空值，無法往回找到對應的id
            nat_name=""
        else:
            nat_name=UserDefCode.objects.get(id=dataTuple[12]).desc1

        dataList.append({
            'work_code': dataTuple[0],
            'chi_name': dataTuple[1],
            'dept_name': dept_name,
            'director_id': dataTuple[3],
            'director_name': director_name,
            'corp_name': corp_name,
            'factory_name': factory_name,
            'arrival_date': dataTuple[6],
            'resign_date': dataTuple[7],
            'pos_name': pos_name,
            'rank': rank_name,
            'bonus_factor': bonus_factor_name,
            'eval_class': eval_class_name,
            'nat': nat_name,
        })
    return JsonResponse(dataList,safe=False)


# 取得下屬的衡量指標
def get_metrics_setup_subs_data(request,pk1=None,pk2=None):
# def get_metrics_setup_subs_data(request,pk1=None):
    dataList = []
    if pk1==None:
        return JsonResponse(dataList, safe=False)
    else:
        fieldList = ['work_code',
                     'chi_name',
                     'dept',
                     'director',
                     'corp',
                     'factory',
                     'arrival_date',
                     'resign_date',
                     'pos_id',
                     'rank',
                     'bonus_factor',
                     'eval_class',
                     'nat',
                     'factory_area',]
        # flat=True，返回的結果為單個值，不是Tuple ; distinct前，給予要取的值做order_by
        YM = WorkingYM.objects.values_list('date_yyyy', 'date_mm').get(id=1)
        # commQ = Q(work_code__in=MetricsSetup.objects.values_list('work_code',flat=True).order_by('work_code').distinct())
        # 2021/3/11 增加 : 只取得工作年有指標的下屬
        # commQ = Q(work_code__in=MetricsSetup.objects.values_list('work_code',flat=True).filter(date_yyyy=YM[0],order_number__lt=900).order_by('work_code').distinct())
        commX = Q(eval_class__in=UserDefCode.objects.filter(Q(topic_code_id='eval_class_id'),
                                                            ~Q(desc1='BSC')))  # 2021/07/14 ADD 拿掉BSC的人
        if pk1=='all':    #for pm802 : 衡量指標查詢(人資)
            commQ = Q(work_code__in=MetricsSetup.objects.values_list('work_code',flat=True).order_by('work_code').distinct())
            commQ2 = ~Q(work_code__endswith='-000')
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ,commQ2,commX)
        else:
            commQ = Q(work_code__in=MetricsSetup.objects.values_list('work_code',flat=True).filter(order_number__lt=900).order_by('work_code').distinct())
            commQ_pk1 = Q(director=pk1)
            results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ_pk1, commQ, commX)

        dataList = []
        for dataTuple in results:
            dept_name = "" if dataTuple[2] == None else UserDefCode.objects.get(id=dataTuple[2]).desc1
            director_name = "" if dataTuple[3] == None else EmployeeInfoEasy.objects.get(work_code=dataTuple[3]).chi_name
            corp_name = "" if dataTuple[4] == None else UserDefCode.objects.get(id=dataTuple[4]).desc1
            factory_name = "" if dataTuple[5] == None else Factory.objects.get(id=dataTuple[5]).name
            pos_name = "" if dataTuple[8] == None else UserDefCode.objects.get(id=dataTuple[8]).desc1
            rank_name = "" if dataTuple[9] == None else UserDefCode.objects.get(id=dataTuple[9]).desc1
            bonus_factor_name = "" if dataTuple[10] == None else UserDefCode.objects.get(id=dataTuple[10]).desc1      # 獎金點數
            eval_class_name = "" if dataTuple[11] == None else UserDefCode.objects.get(id=dataTuple[11]).desc1        # kpi/bsc
            nat_name = "" if dataTuple[12] == None else UserDefCode.objects.get(id=dataTuple[12]).desc1               # 國籍
            factory_area_name = "" if dataTuple[13] == None else UserDefCode.objects.get(id=dataTuple[13]).desc1      # 廠區    2021/08/04 add


            # pm406 : 主管審核, 增加『狀態』
            if (pk2=='Y'):
                YM = WorkingYM.objects.values_list('date_yyyy', 'date_mm').get(id=1)
                commQx = Q(metrics_id__in=MetricsSetup.objects.filter(work_code=dataTuple[0],date_yyyy=YM[0],date_mm=YM[1]).order_by('work_code','date_yyyy','date_mm','order_number','order_item'))
                # fieldList = ['metrics__work_code', 'metrics__date_yyyy', 'metrics__date_mm','metrics__order_number', 'metrics__order_item','metrics__metrics_content',
                #              'metrics_id','actual_score','calc_content','metrics_calc', 'last_status_id']
                # score_result = ScoreSheet.objects.filter(commQx)       # instance < work_code chi_name ........>
                score_sheet_last_status = ScoreSheet.objects.filter(commQx).values_list('last_status_id').distinct()   #看整筆整月的狀態
                curr_status = ''
                for T in score_sheet_last_status:
                    curr_status = T[0]

                dataList.append({
                    'work_code': dataTuple[0],
                    'chi_name': dataTuple[1],
                    'dept_name': dept_name,
                    'director_id': dataTuple[3],
                    'director_name': director_name,
                    'corp_name': corp_name,
                    'factory_name': factory_name,
                    'arrival_date': dataTuple[6],
                    'resign_date': dataTuple[7],
                    'pos_name': pos_name,
                    'rank': rank_name,
                    'bonus_factor': bonus_factor_name,
                    'eval_class': eval_class_name,
                    'nat': nat_name,
                    'factory_area':factory_area_name,
                    'last_status': RegActRules.objects.get(id=curr_status).last_status if curr_status else '',
                    'status_desc': RegActRules.objects.get(id=curr_status).status_desc if curr_status else '自評未開始',
                })

            else:
                dataList.append({
                    'work_code': dataTuple[0],
                    'chi_name': dataTuple[1],
                    'dept_name': dept_name,
                    'director_id': dataTuple[3],
                    'director_name': director_name,
                    'corp_name': corp_name,
                    'factory_name': factory_name,
                    'arrival_date': dataTuple[6],
                    'resign_date': dataTuple[7],
                    'pos_name': pos_name,
                    'rank': rank_name,
                    'bonus_factor': bonus_factor_name,
                    'eval_class': eval_class_name,
                    'nat': nat_name,
                    'factory_area':factory_area_name,
                })


        return JsonResponse(dataList, safe=False)


# 取得期同衡量指標
def get_metrics_setup_common(request):
    dataList = []
    fieldList = ['work_code',
                 'chi_name',
                 'dept',
                 'factory']
    # flat=True，返回的結果為單個值，不是Tuple ; distinct前，給予要取的值做order_by
    YM = WorkingYM.objects.values_list('date_yyyy', 'date_mm').get(id=1)
    # commQ1 = Q(work_code__in=MetricsSetup.objects.values_list('work_code', flat=True).filter(date_yyyy=YM[0]).order_by('work_code').distinct())
    commQ1 = Q(work_code__in=MetricsSetup.objects.values_list('work_code', flat=True).order_by('work_code').distinct())
    # commQ = Q(work_code__in=MetricsSetup.objects.values_list('work_code',flat=True).order_by('work_code').distinct())
    # 2021/3/11 增加 : 只取得工作年有指標的下屬
    commQ2 = Q(work_code__endswith='-000')  # '-000'共同指標, '-100'出勤指標
    results = EmployeeInfoEasy.objects.values_list(*fieldList).filter(commQ1,commQ2)

    dataList = []
    for dataTuple in results:
        if dataTuple[2] == None:
            # 避免空值，無法往回找到對應的id
            dept_name = ""
        else:
            dept_name = UserDefCode.objects.get(id=dataTuple[2]).desc1

        if dataTuple[3] == None:
            # 避免空值，無法往回找到對應的id
            factory_name = ""
        else:
            factory_name = Factory.objects.get(id=dataTuple[3]).name


        dataList.append({
            'work_code': dataTuple[0],
            'chi_name': dataTuple[1],
            'dept_name': dept_name,
            'factory_name': factory_name,
        })
    return JsonResponse(dataList, safe=False)


def get_metrics_setupDate_data(request,pk1=None,pk2=None,pk3=None):
    # distinct前，先做order_by，再給予要取的值(取空值) : 做一個空的grid
    results = None
    if pk3:
        results = MetricsSetup.objects.filter(Q(work_code=pk1), Q(date_yyyy=pk2), Q(date_mm=pk3), Q(order_number__lt=900)).values_list('work_code', 'date_yyyy','date_mm').order_by('work_code', 'date_yyyy', 'date_mm').distinct()
    else:
        if pk2:
            results = MetricsSetup.objects.filter(Q(work_code=pk1),Q(date_yyyy=pk2), Q(order_number__lt=900)).values_list('work_code', 'date_yyyy','date_mm').order_by('work_code', 'date_yyyy', 'date_mm').distinct()
        else:
            if pk1:
                # results = MetricsSetup.objects.filter(work_code=pk1).values_list('work_code', 'date_yyyy','date_mm').order_by('work_code', 'date_yyyy', 'date_mm').distinct()
                #20210222 4:58修改
                results = MetricsSetup.objects.filter(Q(work_code=pk1), Q(date_yyyy=pk2), Q(order_number__lt=900)).values_list('work_code','date_yyyy','date_mm').order_by('work_code', 'date_yyyy', 'date_mm').distinct()
            else:
                results=None
    dataList = []
    if results!=None:
        for dataTuple in results:
            dataList.append({
                'work_code': dataTuple[0],
                'date_yyyy': dataTuple[1],
                'date_mm': dataTuple[2],
        })
    return JsonResponse(dataList,safe=False)



def get_metrics_setupDate_data_search(request,pk1=None,pk2=None,pk3=None):
    # distinct前，先做order_by，再給予要取的值(取空值) : 做一個空的grid
    results = None
    if pk3:
        results = MetricsSetup.objects.filter(Q(work_code=pk1), Q(date_yyyy=pk2), Q(date_mm=pk3)).values_list('work_code', 'date_yyyy','date_mm').order_by('work_code', 'date_yyyy', 'date_mm').distinct()
    else:
        if pk2:
            results = MetricsSetup.objects.filter(Q(work_code=pk1),Q(date_yyyy=pk2)).values_list('work_code', 'date_yyyy','date_mm').order_by('work_code', 'date_yyyy', 'date_mm').distinct()
        else:
            if pk1:
                # results = MetricsSetup.objects.filter(work_code=pk1).values_list('work_code', 'date_yyyy','date_mm').order_by('work_code', 'date_yyyy', 'date_mm').distinct()
                #20210222 4:58修改
                results = MetricsSetup.objects.filter(Q(work_code=pk1), Q(date_yyyy=pk2)).values_list('work_code','date_yyyy','date_mm').order_by('work_code', 'date_yyyy', 'date_mm').distinct()
            else:
                results=None


    dataList = []
    if results!=None:
        for dataTuple in results:
            dataList.append({
                'work_code': dataTuple[0],
                'date_yyyy': dataTuple[1],
                'date_mm': dataTuple[2],
        })
    return JsonResponse(dataList,safe=False)



def metrics_setup_score_sheet_pm402(request):
    get_params = request.GET.dict()
    pk1 = get_params.get('work_code_id')
    pk2 = get_params.get('date_yyyy')
    pk3 = get_params.get('date_mm')

    fields = ['metrics_id',
              'metrics_type',
              'work_code',
              'date_yyyy',
              'date_mm',
              'order_number',
              'order_item',
              'metrics_content',
              'metrics_txt1',
              'metrics_number',
              'metrics_txt2',
              'unit_Mcalc',
              'allocation',
              'score_sheet__calc_content',
              'score_sheet__metrics_calc',
              'score_sheet__actual_score',
              'score_sheet__last_status',
              'score_type',
              ]
    #PM402 : 使用者自評
    #score_type : 共同指標的評核方式
    # commQ = ~Q( score_type__in=['B','C'])    # B:主管評核  C:外部匯入
    commQ = ~Q( score_type__in=['B'])    # B:主管評核 2022/04/08 修改外部匯入可看見,不可自評
    results = MetricsSetup.objects.filter( Q(work_code=pk1), Q(date_yyyy=pk2), Q(date_mm=pk3), commQ).\
        values_list(*fields).order_by('work_code', 'date_yyyy', 'date_mm', 'order_number', 'order_item')

    score_type_choices = (
        ('A','自評實績') ,
        ('B','主管評核') ,
        ('C','匯入實績'),
    )

    dataList = []
    if results:
        for dataTuple in results:
            unit_desc = UserDefCode.objects.get(id=dataTuple[11]).desc1 if dataTuple[11] else None
            metrics_desc = UserDefCode.objects.get(id=dataTuple[1]).desc1 if dataTuple[1] else None
            score_type_desc = ''
            for T in score_type_choices:
                if dataTuple[17]==T[0]:
                    score_type_desc=T[1]
                    break
            dataList.append({
                'metrics_id':dataTuple[0],
                'date_yyyy':dataTuple[3],
                'date_mm':dataTuple[4],
                'order_number': dataTuple[5],
                'order_item': dataTuple[6],
                'metrics_content': dataTuple[7],
                'metrics_txt1':dataTuple[8],
                'metrics_number':dataTuple[9],
                'metrics_txt2':dataTuple[10],
                'unit_Mcalc': unit_desc,
                'allocation': dataTuple[12],
                'calc_content':dataTuple[13] if dataTuple[13] else None,
                'metrics_calc':dataTuple[14] if dataTuple[14] else ( 0 if dataTuple[14]==0 else None),    # 0 和 null 都會為false, 若不加第二段if, 傳到前端會變成null  2021/07/27
                'actual_score':dataTuple[15] if dataTuple[15] else ( 0 if dataTuple[15]==0 else None),    # 0 和 null 都會為false, 若不加第二段if, 傳到前端會變成null
                'last_status':RegActRules.objects.get(id=dataTuple[16]).last_status if dataTuple[16] else None,
                'status_desc':RegActRules.objects.get(id=dataTuple[16]).status_desc if dataTuple[16] else None,
                'score_type': score_type_desc,
                'metrics_type':metrics_desc,
            })


    return JsonResponse(dataList,safe=False)


def metrics_setup_score_sheet_pm406(request):
    get_params = request.GET.dict()
    pk1 = get_params.get('work_code_id')
    pk2 = get_params.get('date_yyyy')
    pk3 = get_params.get('date_mm')

    fields = ['metrics_id',
              'metrics_type',
              'work_code',
              'date_yyyy',
              'date_mm',
              'order_number',
              'order_item',
              'metrics_content',
              'metrics_txt1',
              'metrics_number',
              'metrics_txt2',
              'unit_Mcalc',
              'allocation',
              'score_sheet__calc_content',
              'score_sheet__metrics_calc',
              'score_sheet__actual_score',
              'score_sheet__last_status',
              'score_type'
              ]
    #PM402 : 使用者自評
    #score_type : 共同指標的評核方式
    # commQ = ~Q(score_type='C')  # B:主管評核  C:外部匯入
    # results = MetricsSetup.objects.filter( Q(work_code=pk1), Q(date_yyyy=pk2), Q(date_mm=pk3), commQ).\
    #     values_list(*fields).order_by('work_code', 'date_yyyy', 'date_mm', 'order_number', 'order_item')
    results = MetricsSetup.objects.filter( Q(work_code=pk1), Q(date_yyyy=pk2), Q(date_mm=pk3)).\
        values_list(*fields).order_by('work_code', 'date_yyyy', 'date_mm', 'order_number', 'order_item')

    score_type_choices = (
        ('A','自評實績') ,
        ('B','主管評核') ,
        ('C','匯入實績'),
    )
    dataList = []
    if results:
        for dataTuple in results:
            score_type_desc = ''
            for T in score_type_choices:
                if dataTuple[17]==T[0]:
                    score_type_desc=T[1]
                    break
            unit_desc = UserDefCode.objects.get(id=dataTuple[11]).desc1 if dataTuple[11] else None
            metrics_desc = UserDefCode.objects.get(id=dataTuple[1]).desc1 if dataTuple[1] else None
            dataList.append({
                'metrics_id':dataTuple[0],
                'date_yyyy':dataTuple[3],
                'date_mm':dataTuple[4],
                'order_number': dataTuple[5],
                'order_item': dataTuple[6],
                'metrics_content': dataTuple[7],
                'metrics_txt1':dataTuple[8],
                'metrics_number':dataTuple[9],
                'metrics_txt2':dataTuple[10],
                'unit_Mcalc': unit_desc,
                'allocation': dataTuple[12],
                'calc_content':dataTuple[13] if dataTuple[13] else None,
                'metrics_calc':dataTuple[14] if dataTuple[14] else  ( 0 if dataTuple[14]==0 else None),   # 0 和 null 都會為false, 若不加第二段if, 傳到前端會變成null  2021/08/02
                'actual_score':dataTuple[15] if dataTuple[15] else  ( 0 if dataTuple[15]==0 else None),
                'last_status':RegActRules.objects.get(id=dataTuple[16]).last_status if dataTuple[16] else None,
                'status_desc':RegActRules.objects.get(id=dataTuple[16]).status_desc if dataTuple[16] else None,
                'score_type':score_type_desc,
                'metrics_type': metrics_desc,
            })


    return JsonResponse(dataList,safe=False)







#指標的展開至每位員工
class MetricsSetupExpandCommonView(View):
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        post_param = request.POST.dict()
        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')

        expandYear =  post_param.get('commonExpandYear', None)
        expandMonth = post_param.get('commonExpandMonth', None)
        commonWK = request.POST.get('expandTo_Common1', None)  # copyToId : Form item's name    工號

        expandDept = request.POST.get('expandDept', None)
        expandWorkCode = request.POST.get('expandWorkCode', None)

        toYear = request.POST.getlist('commonExpandToYear', [])
        sMonth = request.POST.get('commonExpandStart', None)
        eMonth = request.POST.get('commonExpandEnding', None)

        fieldListX = ['work_code_id',
                      'metrics_type',
                      'order_number',
                      'metrics_content',
                      'metrics_txt1',
                      'metrics_number',
                      'metrics_txt2',
                      'allocation',
                      'low_limit',
                      'order_item',
                      'date_yyyy',
                      'date_mm',
                      'metrics_id',
                      'unit_Mcalc_id',
                      'score_type',
                      'asc_desc',
                      'auto_alloc',
                      'alloc_range',
                      'ef_date',
                      'exp_date'
                      ]

        if expandYear and expandMonth:
            from_results = MetricsSetup.objects.values_list(*fieldListX). \
                order_by('work_code', 'date_yyyy', 'date_mm','order_number', 'order_item'). \
                filter(Q(work_code=commonWK), Q(date_yyyy=expandYear), Q(date_mm=expandMonth) ,  Q(order_number__lt=900))


        # 共同指標有設定, 才繼續做
        if from_results:
            factory = commonWK.split('-')[0]
            commQ =  Q(eval_class__in=UserDefCode.objects.filter(Q(topic_code_id='eval_class_id'),~Q(desc1='BSC')))
            commQ1 = ~Q(work_code__endswith='-000')
            commQ2 = Q(factory_id=factory)
            commQ3 = Q(resign_date__isnull=True)
            #沒有離職日期,
            if ( (expandDept=='') & (expandWorkCode=='') ):
                to_results= EmployeeInfoEasy.objects.values_list('work_code', flat=True).filter(commQ1, commQ2,commQ3,commQ)
            elif ( (not expandDept=='') & (expandWorkCode=='')):
                commQQ = Q(dept_id=(UserDefCode.objects.get(topic_code_id='dept_id', udc=expandDept).id))
                to_results= EmployeeInfoEasy.objects.values_list('work_code', flat=True).filter(commQ1, commQ2,commQ3,commQ,commQQ)
            elif ( (not expandDept=='') & (not expandWorkCode=='') ):
                commQ1 = Q(work_code=expandWorkCode)
                to_results = EmployeeInfoEasy.objects.values_list('work_code', flat=True).filter(commQ1, commQ2, commQ3,commQ)



            expandFromDict = from_results.__dict__
            datasFrom = expandFromDict.get('_result_cache')

            # f = open("A_results_from_to.txt", 'w+')
            # f.write('=' * 600 + '\n')
            lastOrderNumber = 0
            expandToList = []
            if to_results:
                for Tfrom in datasFrom:
                    # f.write(str(datasFrom.index(Tfrom)) + '=' + str(Tfrom) + '\n')
                    # f.write('-' * 10 + '\n')
                    for Tto in to_results:
                        for month in range(int(sMonth),int(eMonth)+1):
                            expandToList.append({
                                'work_code': Tto,
                                # 'date_yyyy': expandYear,
                                'date_yyyy': toYear[0],
                                'date_mm': month,
                                'metrics_type': Tfrom[1],
                                'order_number': Tfrom[2] + 900,  # 9開頭，為共同指標
                                'metrics_content': Tfrom[3],
                                'metrics_txt1': Tfrom[4],
                                'metrics_number': Tfrom[5],
                                'metrics_txt2': Tfrom[6],
                                'allocation': Tfrom[7],
                                'low_limit': Tfrom[8],
                                'order_item': Tfrom[9],
                                'unit_Mcalc_id': Tfrom[13],
                                'score_type': Tfrom[14],
                                'asc_desc': Tfrom[15],
                                'auto_alloc': Tfrom[16],
                                'alloc_range': Tfrom[17],
                                'ef_date': Tfrom[18],
                                'exp_date': Tfrom[19],
                                'urrf':'Common metrics expand',
                                'urrf1':expandYear+' '+expandMonth+' '+commonWK,
                                'urrf2':sMonth+'~'+eMonth
                            })
                            # f.write(str(Tto)+'-'+str(toYear[0])+'-'+str(month)+'\n')
                # f.write('=' * 600)
                # f.close()

                # f = open("B_expandToList.txt", 'w+')
                # f.write('=' * 600 + '\n')
                # for ss in expandToList:
                    # f.write(str(expandToList.index(ss)) + '=' + str(ss) + '\n')
                # f.write('=' * 600)
                # f.close()

                bulk_data = [MetricsSetup(
                    change_time=now,
                    changer=request.user.username,
                    create_time=now,
                    creator=request.user.username,
                    work_code_id=expandTo.get('work_code'),
                    date_yyyy=expandTo.get('date_yyyy'),
                    date_mm=expandTo.get('date_mm'),
                    metrics_type_id=expandTo.get('metrics_type'),
                    order_number=expandTo.get('order_number'),
                    metrics_content=expandTo.get('metrics_content'),
                    metrics_txt1=expandTo.get('metrics_txt1'),
                    metrics_number=expandTo.get('metrics_number'),
                    metrics_txt2=expandTo.get('metrics_txt2'),
                    allocation=expandTo.get('allocation'),
                    low_limit=expandTo.get('low_limit'),
                    order_item=expandTo.get('order_item'),
                    unit_Mcalc_id=expandTo.get('unit_Mcalc'),
                    score_type=expandTo.get('score_type'),
                    asc_desc=expandTo.get('asc_desc'),
                    auto_alloc=expandTo.get('auto_alloc'),
                    alloc_range=expandTo.get('alloc_range'),
                    ef_date=expandTo.get('ef_date'),
                    exp_date=expandTo.get('exp_date'),
                    urrf=expandTo.get('urrf'),
                    urrf1=expandTo.get('urrf1'),
                    urrf2=expandTo.get('urrf2')
                ) for expandTo in expandToList]  # copyTo是一個dict

                # f = open("C_bulk_data.txt", 'w+')
                # f.write('=' * 600 + '\n')
                # for ss in bulk_data:
                #     f.write(str(bulk_data.index(ss)) + '=' + str(ss) + '\n')
                # f.write('=' * 600)
                # f.close()

            if len(toYear) >=2:
                err_message = "展開資料不成功，"
                try:
                    MetricsSetup.objects.bulk_create(bulk_data)
                except ValidationError as e:
                    logger.debug(e)
                    if e.code == 'invalid_time':
                        return render(request, 'kpi/copy_error.html', context={"success": False,
                                                                               "message": err_message + "\"%s\" 格式正確，卻非有效時間。" %
                                                                                          e.params['value']})
                    else:
                        return render(request, 'kpi/copy_error.html',
                                      context={"success": False, "message": err_message + "請檢查欄位數值是否正確。"})
                except ValueError as e:
                    logger.debug(e)
                    return render(request, 'kpi/copy_error.html',
                                  context={"success": False, "message": err_message + "欄位不能為空，且只能是數字。"})
                except pymssql.DatabaseError as e:
                    logger.debug(e)
                    return render(request, 'kpi/copy_error.html',
                                  context={"success": False, "message": err_message + "已存在資料庫中..."})
                except pymssql.IntegrityError as e:
                    logger.debug(e)
                    return render(request, 'kpi/copy_error.html',
                                  context={"success": False, "message": err_message + "已存在資料庫中..."})
                except Exception as e:
                    logger.debug(e)
                    return render(request, 'kpi/copy_error.html',
                                  context={"success": False, "message": err_message + "已存在資料庫中..."})
                else:   # 展開成功,繼續展開計算方式
                    fieldList = ['metrics', 'order_number', 'calc_content',
                                 'lower_limit', 'upper_limit', 'score',
                                 'metrics__work_code', 'metrics__date_yyyy', 'metrics__date_mm',
                                 'metrics__order_number', 'metrics__order_item']
                    expandToCalc = []
                    # 找出MetricsSetup新的metrics_id
                    every_metrics = len(to_results) * ( int(eMonth)-int(sMonth)+ 1)     # 工號*月份=每一份共同指標要展開的筆數
                    xx = 0
                    for copied in expandToList:
                        new_metrics = MetricsSetup.objects.\
                            order_by('work_code', 'date_yyyy', 'date_mm', 'order_number', 'order_item').\
                            get(work_code=copied.get('work_code'),
                                date_yyyy=copied.get('date_yyyy'),
                                date_mm=copied.get('date_mm'),
                                order_number=copied.get('order_number'),
                                order_item=copied.get('order_item')
                                )
                        idx = expandToList.index(copied) // every_metrics   #  // 取得商(無條件捨去小數)
                        key_id = datasFrom[ idx ][12]

                        old_calc = MetricsCalc.objects.order_by('metrics_id', 'order_number').filter(
                            metrics_id=key_id).values_list(*fieldList)  # 找出source的MetricsCalc

                        for T in old_calc:
                            expandToCalc.append({
                                'metrics': new_metrics,
                                'order_number': T[1],
                                'calc_content': T[2],
                                'lower_limit': T[3],
                                'upper_limit': T[4],
                                'score': T[5]
                            })

                    # bulk_data : 要做bulk_create的data，將之整理成List({})<--Dict內含於List，再做Iteration(for)
                    bulk_data_ToCalc = [MetricsCalc(
                        change_time=now,
                        changer=request.user.username,
                        create_time=now,
                        creator=request.user.username,
                        metrics=ToCalc.get('metrics'),
                        order_number=ToCalc.get('order_number'),
                        calc_content=ToCalc.get('calc_content'),
                        lower_limit=ToCalc.get('lower_limit'),
                        upper_limit=ToCalc.get('upper_limit'),
                        score=ToCalc.get('score'),
                    ) for ToCalc in expandToCalc]  # copyTo是一個dict

                    # f = open("D_bulk_data_ToCalc.txt", 'w+')
                    # f.write('=' * 600 + '\n')
                    # for ss in bulk_data_ToCalc:
                    #     f.write(str(bulk_data_ToCalc.index(ss)) + '=' + str(ss) + '\n')
                    # f.write('=' * 600)
                    # f.close()

                    try:
                        MetricsCalc.objects.bulk_create(bulk_data_ToCalc)
                    except ValidationError as e:
                        logger.debug(e)
                        if e.code == 'invalid_time':
                            return render(request, 'kpi/copy_error.html', context={"success": False,
                                                                                   "message": err_message + "\"%s\" 格式正確，卻非有效時間。" %
                                                                                              e.params['value']})
                        else:
                            return render(request, 'kpi/copy_error.html',
                                          context={"success": False, "message": err_message + "請檢查欄位數值是否正確。"})
                    except ValueError as e:
                        logger.debug(e)
                        return render(request, 'kpi/copy_error.html',
                                      context={"success": False, "message": err_message + "欄位不能為空，且只能是數字。"})
                    except pymssql.DatabaseError as e:
                        logger.debug(e)
                        return render(request, 'kpi/copy_error.html',
                                      context={"success": False, "message": err_message + "已存在資料庫中..."})
                    except pymssql.IntegrityError as e:
                        logger.debug(e)
                        return render(request, 'kpi/copy_error.html',
                                      context={"success": False, "message": err_message + "已存在資料庫中..."})
                    except Exception as e:
                        logger.debug(e)
                        return render(request, 'kpi/copy_error.html',
                                      context={"success": False, "message": err_message + "已存在資料庫中..."})
                    else:
                        cache.clear()
                        # 複製成功,返回上一頁
                         # return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'), {"success": True})
                        err_message = "展開成功"
                        return render(request, 'kpi/import_success.html', context={"success": False, "message": err_message})
            else:
                return render(request, 'kpi/copy_error.html',context={"success": False, "message": "展開不成功"})


#指標的從某些員工收回
class MetricsSetupRecallCommonView(View):
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        get_param = request.GET.dict()

    def post(self, request, *args, **kwargs):
        post_param = request.POST.dict()
        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')

        commonWK = request.POST.get('recallCommon', None)
        recallDept = request.POST.get('recallDept', None)
        recallWorkCode = request.POST.get('recallWorkCode', None)

        recallYear = request.POST.get('recallYear', None)
        sMonth = request.POST.get('recallStart', None)
        eMonth = request.POST.get('recallEnding', None)


        #只要找到同一公司的所有員工即可
        factory=commonWK.split('-')[0]
        if commonWK:
            work_code_toDe = ''
            if ( (recallDept=='') & (recallWorkCode=='') ):
                work_code_toDel = Q(work_code__in=EmployeeInfoEasy.objects.values_list('work_code', flat=True).
                                    filter(Q(factory_id=factory),~Q(work_code__endswith='-000')))
                results = MetricsSetup.objects.filter(work_code_toDel, Q(date_yyyy=recallYear), Q(date_mm__gte=sMonth),
                                                      Q(date_mm__lte=eMonth), Q(order_number__gt=900))
            elif ((not recallDept=='') & (recallWorkCode=='')):
                commQQ = Q(dept_id=(UserDefCode.objects.get(topic_code_id='dept_id',udc=recallDept).id))
                work_code_toDel = Q(work_code__in=EmployeeInfoEasy.objects.values_list('work_code', flat=True).
                                    filter(Q(factory_id=factory),~Q(work_code__endswith='-000'),commQQ))
                results = MetricsSetup.objects.filter(work_code_toDel, Q(date_yyyy=recallYear), Q(date_mm__gte=sMonth),
                                                      Q(date_mm__lte=eMonth), Q(order_number__gt=900))
            elif ( (not recallDept=='') & (not recallWorkCode=='')):
                results = MetricsSetup.objects.filter(work_code_id=recallWorkCode, date_yyyy=recallYear,date_mm__gte=sMonth,date_mm__lte=eMonth,order_number__gt=900)
            results_calc = MetricsCalc.objects.filter(metrics_id__in=results)
            try:
                results_calc.delete()
                results.delete()
            except:
                err_message = "收回錯誤"
                return render(request, 'kpi/import_error.html', context={"success": False, "message": err_message})
            else:
                err_message = "收回完成"
                return render(request, 'kpi/import_success.html', context={"success": False, "message": err_message})



#員工基本資料的複製
class EmployeeCopyView(View):
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        post_param = request.POST.dict()
        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')

        copyFrom_id = post_param.get('copyFrom_id', None)
        copyFrom = {}
        if copyFrom_id:
            results = EmployeeInfoEasy.objects.get(work_code=copyFrom_id)     #自db取得來源row
            if (results):
                copyFrom = results.__dict__
            else:
                copyFrom = {}

        work_codes = request.POST.getlist('copyToId', [])            # copyToId : Form item's name
        chi_names = request.POST.getlist('copyToVal1', [])            # copyToVal1 : Form item's name
        # bulk_data : 要做bulk_create的data，將之整理成List({})<--Dict內含於List，再做Iteration(for)

        bulk_data = [EmployeeInfoEasy(
            change_time=now,
            changer=request.user.username,
            create_time=now,
            creator=request.user.username,
            work_code=copyToId,
            chi_name=copyToVal,
            corp_id=copyFrom.get('corp_id'),
            factory_id=copyFrom.get('factory_id'),
            dept_id=copyFrom.get('dept_id'),
            pos_id=copyFrom.get('pos_id'),
            director_id=copyFrom.get('director_id'),
            arrival_date=copyFrom.get('arrival_date'),
            resign_date=copyFrom.get('resign_date'),
            rank_id=copyFrom.get('rank_id'),
            nat_id=copyFrom.get('nat_id'),
            bonus_factor_id=copyFrom.get('bonus_factor_id'),
            eval_class_id=copyFrom.get('eval_class_id'),
            labor_type_id=copyFrom.get('labor_type_id'),
            bonus_type_id=copyFrom.get('bonus_type_id'),
            email=copyFrom.get('email'),

            factory_area_id=copyFrom.get('factory_area_id'),
            dept_flevel_id=copyFrom.get('dept_flevel_id'),
            dept_desc_id=copyFrom.get('dept_desc_id'),
            direct_supv_id=copyFrom.get('direct_supv_id'),
            # trans_date=copyFrom.get('trans_date'),
            # trans_type=copyFrom.get('trans_type'),
            service_status_id=copyFrom.get('service_status_id'),
            kpi_diy=copyFrom.get('kpi_diy'),
            # skill_diy=copyFrom.get('skill_diy'),
        ) for (copyToId,copyToVal) in zip(work_codes,chi_names)]


        if len(work_codes)>=3:    # 第二次post的值(不知為何會有二次post, 第一次post的data只有一組)
            for val in work_codes:
                if val=="":
                    blank_index = work_codes.index(val)
                    bulk_data.pop(blank_index)      # 移除空值(沒有輸入料號的list)

            err_message = "複製資料不成功，"
            try:
                EmployeeInfoEasy.objects.bulk_create(bulk_data)
            except ValidationError as e:
                logger.debug(e)
                if e.code == 'invalid_time':
                    return render(request, 'kpi/copy_error.html', context={"success": False, "message": err_message + "\"%s\" 格式正確，卻非有效時間。" % e.params['value']})
                else:
                    return render(request, 'kpi/copy_error.html', context={"success": False,"message": err_message + "請檢查欄位數值是否正確。" })
            except ValueError as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "欄位不能為空，且只能是數字。"})
            except pymssql.DatabaseError as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "已存在資料庫中..."})
            except pymssql.IntegrityError as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "已存在資料庫中..."})
            except Exception as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "已存在資料庫中..."})
            else:
                cache.clear()
                # 匯入成功,返回上一頁
                cache.clear()
                return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'), {"success": True})


#指標的展開1~12月
class MetricsSetupExpandView(View):
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        post_param = request.POST.dict()
        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')

        expandYear = post_param.get('expandYear', None)

        work_codes = request.POST.getlist('expandTo_WorkCode', [])  # copyToId : Form item's name    工號

        copyFromDict = {}
        copyToList = []

        fieldListX = ['work_code_id',
                      'metrics_type',
                      'order_number',
                      'metrics_content',
                      'metrics_txt1',
                      'metrics_number',
                      'metrics_txt2',
                      'allocation',
                      'low_limit',
                      'order_item',
                      'date_yyyy',
                      'date_mm',
                      'metrics_id',
                      'unit_Mcalc_id',
                      'score_type',
                      'asc_desc',
                      'auto_alloc',
                      'alloc_range',
                      'ef_date',
                      'exp_date',
                      'asc_desc_score',
                      ]

        if expandYear:
            # 年度一定要輸入，前端控制，工號未輸入，不予複製
            # 年度，月份=0，一定要存在
            # for work_code in work_codes:    #改為只有一個工號,所有移除2021/03/30
            #     #每一個工號，都有一組新的1~12個月，(List)不累加
            if work_codes[0] != '':
                results = MetricsSetup.objects.values_list(*fieldListX).order_by('work_code', 'date_yyyy', 'date_mm','order_number','order_item').\
                    filter(Q(work_code=work_codes[0]), Q(date_yyyy=expandYear), Q(date_mm=0), Q(order_number__lt=900))
                if (results):
                    copyFromDict = results.__dict__
                    datas = copyFromDict.get('_result_cache')
                    # 產生各工號(work_code)的metrics
                    YM = WorkingYM.objects.values_list('date_yyyy', 'date_mm').get(id=1)
                    if ( int(expandYear)==YM[0] ):
                        startMonth = YM[1]+1
                    else:
                        startMonth = 1

                    for month in range(startMonth,13):  #月份設定為1~12月
                        for dataTuple in datas:
                            # 每個工號(work_code)的metrics
                            # metrics都相同，但date_mm不同
                            copyToList.append({
                                'work_code': work_codes[0],
                                'metrics_type': dataTuple[1],
                                'order_number': dataTuple[2],
                                'metrics_content': dataTuple[3],
                                'metrics_txt1': dataTuple[4],
                                'metrics_number': dataTuple[5],
                                'metrics_txt2':dataTuple[6],
                                'allocation': dataTuple[7],
                                'low_limit': dataTuple[8],
                                'order_item': dataTuple[9],
                                'date_yyyy': dataTuple[10],
                                'date_mm': month,
                                'unit_Mcalc':dataTuple[13],
                                'score_type': dataTuple[14],  # 2021/06/01增加欄位
                                'asc_desc': dataTuple[15],
                                'auto_alloc': dataTuple[16],
                                'alloc_range': dataTuple[17],
                                'ef_date': dataTuple[18],
                                'exp_date': dataTuple[19],
                                'asc_desc_score': dataTuple[20],
                            })
        # bulk_data : 要做bulk_create的data，將之整理成List({})<--Dict內含於List，再做Iteration(for)

        bulk_data = [MetricsSetup(
            change_time=now,
            changer=request.user.username,
            create_time=now,
            creator=request.user.username,
            work_code_id=copyTo.get('work_code'),
            metrics_type_id=copyTo.get('metrics_type'),
            order_number=copyTo.get('order_number'),
            metrics_content=copyTo.get('metrics_content'),
            metrics_txt1=copyTo.get('metrics_txt1'),
            metrics_number=copyTo.get('metrics_number'),
            metrics_txt2=copyTo.get('metrics_txt2'),
            allocation=copyTo.get('allocation'),

            low_limit=copyTo.get('low_limit'),
            order_item=copyTo.get('order_item'),
            # date_yyyy=  0 if date_yyyy_all[copyToList.index(copyTo)]==0 else date_yyyy_all[copyToList.index(copyTo)],
            date_yyyy=copyTo.get('date_yyyy'),
            # date_mm=0 if len(date_mm_all)==0 else date_mm_all[copyToList.index(copyTo)],
            date_mm=copyTo.get('date_mm'),
            unit_Mcalc_id=copyTo.get('unit_Mcalc'),
            score_type=copyTo.get('score_type'),  # 2021/06/01增加欄位
            asc_desc=copyTo.get('asc_desc'),
            auto_alloc=copyTo.get('auto_alloc'),
            alloc_range=copyTo.get('alloc_range'),
            ef_date=copyTo.get('ef_date'),
            exp_date=copyTo.get('exp_date'),
            asc_desc_score=copyTo.get('asc_desc_score'),
        ) for copyTo in copyToList]  # copyTo是一個dict

        if len(work_codes) >= 2:  # 第二次post的值(不知為何會有二次post, 第一次post的data只有一組)
            err_message = "複製資料不成功，"
        try:
            MetricsSetup.objects.bulk_create(bulk_data)
        except ValidationError as e:
            logger.debug(e)
            if e.code == 'invalid_time':
                return render(request, 'kpi/copy_error.html', context={"success": False, "message": err_message + "\"%s\" 格式正確，卻非有效時間。" % e.params['value']})
            else:
                return render(request, 'kpi/copy_error.html', context={"success": False,"message": err_message + "請檢查欄位數值是否正確。" })
        except ValueError as e:
            logger.debug(e)
            return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "欄位不能為空，且只能是數字。"})
        except pymssql.DatabaseError as e:
            logger.debug(e)
            return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "已存在資料庫中..."})
        except pymssql.IntegrityError as e:
            logger.debug(e)
            return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "已存在資料庫中..."})
        except Exception as e:
            logger.debug(e)
            return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "已存在資料庫中..."})
        else:
            # MetricsCalc計算方式expand..............................................................................................................................Begin
            fieldList = ['metrics', 'order_number', 'calc_content', \
                     'lower_limit', 'upper_limit', 'score', \
                     'metrics__work_code', 'metrics__date_yyyy', 'metrics__date_mm', \
                     'metrics__order_number', 'metrics__order_item']
            copyToCalc = []
            # 找出MetricsSetup新的metrics_id
            count = 0
            for copied in copyToList:
                new_metrics = MetricsSetup.objects. \
                    order_by('work_code', 'date_yyyy', 'date_mm','order_number','order_item'). \
                    get(work_code=copied.get('work_code'), \
                        date_yyyy=copied.get('date_yyyy'), \
                        date_mm=copied.get('date_mm'), \
                        order_number=copied.get('order_number'), \
                        order_item=copied.get('order_item')
                        )

                # copyToList.index(copied)新的

                key_id = datas[copyToList.index(copied) % len(datas)][12]  # 取得list裏的Tuple( datas 是MetricsSetup's CopyFromData)
                old_calc = MetricsCalc.objects.order_by('metrics_id', 'order_number').filter(
                    metrics_id=key_id).values_list(*fieldList)  # 找出source的MetricsCalc
                for t in old_calc:
                    copyToCalc.append({
                        'metrics': new_metrics,
                        'order_number': t[1],
                        'calc_content': t[2],
                        'lower_limit': t[3],
                        'upper_limit': t[4],
                        'score': t[5]
                    })


            # bulk_data : 要做bulk_create的data，將之整理成List({})<--Dict內含於List，再做Iteration(for)
            calc_bulk_data = [MetricsCalc(
                change_time=now,
                changer=request.user.username,
                create_time=now,
                creator=request.user.username,
                metrics=copyTo.get('metrics'),
                order_number=copyTo.get('order_number'),
                calc_content=copyTo.get('calc_content'),
                lower_limit=copyTo.get('lower_limit'),
                upper_limit=copyTo.get('upper_limit'),
                score=copyTo.get('score'),
            ) for copyTo in copyToCalc]  # copyTo是一個dict


            try:
              MetricsCalc.objects.bulk_create(calc_bulk_data)
            except ValidationError as e:
                logger.debug(e)
                if e.code == 'invalid_time':
                    return render(request, 'kpi/copy_error.html', context={"success": False,
                                                                           "message": err_message + "\"%s\" 格式正確，卻非有效時間。" %
                                                                                      e.params['value']})
                else:
                    return render(request, 'kpi/copy_error.html',
                                  context={"success": False, "message": err_message + "請檢查欄位數值是否正確。"})
            except ValueError as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',
                              context={"success": False, "message": err_message + "欄位不能為空，且只能是數字。"})
            except pymssql.DatabaseError as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',
                              context={"success": False, "message": err_message + "已存在資料庫中..."})
            except pymssql.IntegrityError as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',
                              context={"success": False, "message": err_message + "已存在資料庫中..."})
            except Exception as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',
                              context={"success": False, "message": err_message + "已存在資料庫中..."})
            else:
                cache.clear()
                # MetricsCalc計算方式copy..............................................................................................................................Ending
                # 複製成功,返回上一頁
                return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'), {"success": True})
            # MetricsCalc計算方式Expand..............................................................................................................................Ending
        # else:
        #     return render(request, 'kpi/copy_error.html',context={"success": False, "message": "展開年度未輸入"})


#指標的複製
class MetricsSetupCopyView(View):
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        post_param = request.POST.dict()
        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')

        copyFrom_id = post_param.get('copyFrom_id', None)
        copyFrom_val1 = post_param.get('copyFrom_val1', None)     #年度
        copyFrom_val2 = post_param.get('copyFrom_val2', None)     #月份
        work_codes = request.POST.getlist('copyToId', [])   # copyToId : Form item's name    工號
        date_yyyy_all = request.POST.getlist('copyToVal1', [])  # copyToVal1 : Form item's name  年份
        date_mm_all = request.POST.getlist('copyToVal2', [])    # copyToVal2 : Form item's name  月份
        results = None
        copyFromDict = {}
        copyToList = []

        # MerticsSetup複製----------------------------------------------------------------------------------------------------------------------------------------Begin
        fieldListX = ['work_code_id',
                     'metrics_type',
                     'order_number',
                     'metrics_content',
                     'metrics_txt1',
                     'metrics_number',
                     'metrics_txt2',
                     'allocation',
                     'low_limit',
                     'order_item',
                     'date_yyyy',
                     'date_mm',
                     'metrics_id',
                     'unit_Mcalc_id',
                     'score_type',
                     'asc_desc',
                     'auto_alloc',
                     'alloc_range',
                     'ef_date',
                     'exp_date',
                     'asc_desc_score',
                      ]
        if copyFrom_id:
            # 工號一定要輸入，前端控制，工號未輸入，不予複製
            if copyFrom_val1:
                if copyFrom_val2:
                    # 有工號,有年度,有月份
                    results = MetricsSetup.objects.values_list(*fieldListX) \
                        .filter(Q(work_code=copyFrom_id),Q(date_yyyy=copyFrom_val1),Q(date_mm=copyFrom_val2),Q(order_number__lt=900))
                else:
                    # 有工號,有年度
                    results = MetricsSetup.objects.values_list(*fieldListX)\
                        .filter(Q(work_code=copyFrom_id),Q(date_yyyy=copyFrom_val1),Q(order_number__lt=900))

        if (results):
            copyFromDict = results.__dict__
            datas = copyFromDict.get('_result_cache')

            #產生各工號(work_code)的metrics
            count = 0
            for work_code in work_codes:     # 若在coptTo的工號相同於copyFrom時 work_codes.index(work_code), 永遠是0, 所以要額外計數( 不論工號 coptT 是否和 copyFrom 相同 )
                # 有輸入複製工號的，才產生資料
                # Tuple轉成List({})
                # date_yyyy_to = ''
                # if (len(date_yyyy_all)) > 0:  # 避免LIST index out of range
                #     date_yyyy_to = date_yyyy_all[count]


                date_yyyy_to =  date_yyyy_all[count] if (len(date_yyyy_all)) > 0 else ''

                # date_mm_to = ''
                # if (len(date_mm_all)) > 0:  # 避免LIST index out of range
                #     date_mm_to = date_mm_all[count]
                date_mm_to = date_mm_all[count] if (len(date_mm_all)) > 0  else ''

                if work_code!='':
                    for dataTuple in datas:
                        # 每個工號(work_code)的metrics
                        # metrics都相同，但work_code不同
                        copyToList.append({
                                'work_code': work_code,
                                'metrics_type':dataTuple[1],
                                'order_number':dataTuple[2],
                                'metrics_content': dataTuple[3],
                                'metrics_txt1':dataTuple[4],
                                'metrics_number':dataTuple[5],
                                'metrics_txt2':dataTuple[6],
                                'allocation': dataTuple[7],
                                'low_limit':dataTuple[8],
                                'order_item':dataTuple[9],
                                # 2021/03/15 修改
                                # work_codes.index 為何不用此取值 : 當work_codes這個list裏的每個"工號"都一樣的時候, 永遠都只會取到index=0, 所以同一個工號複製到不同月份就會有問題了
                                # 'date_yyyy': date_yyyy_all[work_codes.index(work_code)],
                                # 'date_mm': date_mm_all[work_codes.index(work_code)],
                                'date_yyyy': dataTuple[10] if date_yyyy_to == '' else date_yyyy_to,
                                'date_mm': dataTuple[11] if date_mm_to == '' else date_mm_to,
                                'unit_Mcalc': dataTuple[13],
                                'score_type': dataTuple[14],  #2021/06/01增加欄位
                                'asc_desc': dataTuple[15],
                                'auto_alloc': dataTuple[16],
                                'alloc_range': dataTuple[17],
                                'ef_date': dataTuple[18],
                                'exp_date': dataTuple[19],
                                'asc_desc_score':dataTuple[20],
                        })
                count += 1  # 所以要額外計數( 不論工號 coptTo 是否和 copyFrom 相同 )
            else:
                err_message = '工號無此年度或月份的衡量指標，'

        # bulk_data : 要做bulk_create的data，將之整理成List({})<--Dict內含於List，再做Iteration(for)
        bulk_data = [MetricsSetup(
            change_time=now,
            changer=request.user.username,
            create_time=now,
            creator=request.user.username,
            work_code_id=copyTo.get('work_code'),
            metrics_type_id=copyTo.get('metrics_type'),
            order_number=copyTo.get('order_number'),
            metrics_content=copyTo.get('metrics_content'),
            metrics_txt1=copyTo.get('metrics_txt1'),
            metrics_number=copyTo.get('metrics_number'),
            metrics_txt2=copyTo.get('metrics_txt2'),
            allocation=copyTo.get('allocation'),
            low_limit=copyTo.get('low_limit'),
            order_item=copyTo.get('order_item'),
            # date_yyyy=  0 if date_yyyy_all[copyToList.index(copyTo)]==0 else date_yyyy_all[copyToList.index(copyTo)],
            date_yyyy= copyTo.get('date_yyyy'),
            # date_mm=0 if len(date_mm_all)==0 else date_mm_all[copyToList.index(copyTo)],
            date_mm=copyTo.get('date_mm'),
            unit_Mcalc_id=copyTo.get('unit_Mcalc'),
            score_type=copyTo.get('score_type'),    #2021/06/01增加欄位
            asc_desc=copyTo.get('asc_desc'),
            auto_alloc=copyTo.get('auto_alloc'),
            alloc_range=copyTo.get('alloc_range'),
            ef_date=copyTo.get('ef_date'),
            exp_date=copyTo.get('exp_date'),
            asc_desc_score= copyTo.get('asc_desc_score'),
        ) for copyTo in copyToList]    # copyTo是一個dict


        if len(work_codes) >= 6:  # 第二次post的值(不知為何會有二次post, 第一次post的data只有一組)
            err_message = "複製資料不成功，"
            try:
                MetricsSetup.objects.bulk_create(bulk_data)
            except ValidationError as e:
                logger.debug(e)
                if e.code == 'invalid_time':
                    return render(request, 'kpi/copy_error.html', context={"success": False, "message": err_message + "\"%s\" 格式正確，卻非有效時間。" % e.params['value']})
                else:
                    return render(request, 'kpi/copy_error.html', context={"success": False,"message": err_message + "請檢查欄位數值是否正確。" })
            except ValueError as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "欄位不能為空，且只能是數字。"})
            except pymssql.DatabaseError as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "已存在資料庫中..."})
            except pymssql.IntegrityError as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "已存在資料庫中..."})
            except Exception as e:
                logger.debug(e)
                return render(request, 'kpi/copy_error.html',context={"success": False, "message": err_message + "已存在資料庫中..."})
            else:
                # MetricsCalc計算方式copy..............................................................................................................................Begin
                fieldList = ['metrics', 'order_number', 'calc_content',\
                             'lower_limit', 'upper_limit', 'score',\
                             'metrics__work_code', 'metrics__date_yyyy', 'metrics__date_mm',\
                             'metrics__order_number', 'metrics__order_item']

                copyToCalc = []
                # 找出MetricsSetup新的metrics_id
                count = 0
                for copied in copyToList:
                    new_metrics = MetricsSetup.objects. \
                        get(work_code=copied.get('work_code'), \
                            date_yyyy=copied.get('date_yyyy'), \
                            date_mm=copied.get('date_mm'), \
                            order_number=copied.get('order_number'), \
                            order_item=copied.get('order_item')
                            )
                    # copyToList.index(copied)新的

                    # key_id  = datas[copyToList.index(copied)][12]   #取得list裏的Tuple( datas 是MetricsSetup's CopyFromData)
                    key_id  = datas[copyToList.index(copied) % len(datas)][12]   #取得list裏的Tuple( datas 是MetricsSetup's CopyFromData)
                    # for dataTuple in datas:  # datas : MetricsSetup 的source資料
                    old_calc = MetricsCalc.objects.order_by('metrics_id', 'order_number').filter(
                        metrics_id=key_id).values_list(*fieldList)  # 找出source的MetricsCalc

                    for t in old_calc:
                        copyToCalc.append({
                            'metrics': new_metrics,
                            'order_number':t[1],
                            'calc_content':t[2],
                            'lower_limit':t[3],
                            'upper_limit':t[4],
                            'score':t[5]
                        })

                # bulk_data : 要做bulk_create的data，將之整理成List({})<--Dict內含於List，再做Iteration(for)
                calc_bulk_data = [MetricsCalc(
                    change_time=now,
                    changer=request.user.username,
                    create_time=now,
                    creator=request.user.username,
                    metrics=copyTo.get('metrics'),
                    order_number=copyTo.get('order_number'),
                    calc_content=copyTo.get('calc_content'),
                    lower_limit= copyTo.get('lower_limit'),
                    upper_limit=copyTo.get('upper_limit'),
                    score=copyTo.get('score'),
                ) for copyTo in copyToCalc]  # copyTo是一個dict


                try:
                    MetricsCalc.objects.bulk_create(calc_bulk_data)
                except ValidationError as e:
                    logger.debug(e)
                    if e.code == 'invalid_time':
                        return render(request, 'kpi/copy_error.html', context={"success": False,
                                                                               "message": err_message + "\"%s\" 格式正確，卻非有效時間。" %
                                                                                          e.params['value']})
                    else:
                        return render(request, 'kpi/copy_error.html',
                                      context={"success": False, "message": err_message + "請檢查欄位數值是否正確。"})
                except ValueError as e:
                    logger.debug(e)
                    return render(request, 'kpi/copy_error.html',
                                  context={"success": False, "message": err_message + "欄位不能為空，且只能是數字。"})
                except pymssql.DatabaseError as e:
                    logger.debug(e)
                    return render(request, 'kpi/copy_error.html',
                                  context={"success": False, "message": err_message + "已存在資料庫中..."})
                except pymssql.IntegrityError as e:
                    logger.debug(e)
                    return render(request, 'kpi/copy_error.html',
                                  context={"success": False, "message": err_message + "已存在資料庫中..."})
                except Exception as e:
                    logger.debug(e)
                    return render(request, 'kpi/copy_error.html',
                                  context={"success": False, "message": err_message + "已存在資料庫中..."})
                else:
                    cache.clear()
                # MetricsCalc計算方式copy..............................................................................................................................Ending
                    # 複製成功,返回上一頁
                    return HttpResponseRedirect(request.META.get('HTTP_REFERER', '/'), {"success": True})


#衡量指標批次刪除
class MetricsBatchDeleView(View):
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        post_param = request.POST.dict()
        qry1 = Q( work_code__in = post_param.get('deleWorkCode', None).split(',') )
        qry2 = Q( date_yyyy = post_param.get('deleYear', None) )
        qry3 = Q( date_mm__gte = post_param.get('deleMonthBegin', None) )
        qry4 = Q( date_mm__lte = post_param.get('deleMonthEnd', None))
        qry5 = Q( order_number__lt = 900)    #2021/08/04 add，不刪除共同指標
        # results = MetricsSetup.objects.values_list('metrics_id').filter(qry1,qry2,qry3,qry4)
        results = MetricsSetup.objects.values_list('metrics_id').filter(qry1,qry2,qry3,qry4,qry5)    #2021/08/04 add，不刪除共同指標
        username = request.user.username
        data1 = FileActionLedger(creator=username,changer=username,
                                    app_name='KPI',model_name='MetricsSetup',
                                    pg_name='PM208',act_btn='batch_dele_btn',btn_name='批次刪除',
                                    url='/api/metrics_setup_batch_dele/',
                                    formdata=str(post_param))
        # metrics_id_dele = [x for x in results]'
        other_data = ''
        # other_data += str( [x[0] for x in results] )

        data2 = FileActionLedger(creator=username,changer=username,
                                    app_name='KPI',model_name='MetricsCalc',
                                    pg_name='PM208',act_btn='batch_dele_btn',btn_name='批次刪除',
                                    url='/api/metrics_setup_batch_dele/',
                                    formdata=str(post_param),
                                    otherdata="{'metrics_id':"+ str( [x[0] for x in results] ) )
        data1.save()
        data2.save()
        MetricsCalc.objects.filter(metrics_id__in=results).delete()
        MetricsSetup.objects.filter(qry1,qry2,qry3,qry4,qry5).delete()

        return JsonResponse({"success": True})



#實績評核
class ScoreSheetProcessView(View):
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        post_param = request.POST.dict()
        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')
        last_status = post_param.get('last_status',None)
        last_status_instance = RegActRules.objects.get(last_status=last_status) if last_status else None
        order_type=ProcessOptionsTxtDef.objects.get \
            (app_model='KPI', view_code='PM402', model_class='RegActRules', topic_code='order_type',
             action='blank').po_code

        #取得主管審核的狀態,只要小於該狀態,每次儲存更新都會備份
        pm406_last_status_new=ProcessOptionsTxtDef.objects.get \
            (app_model='KPI', view_code='PM406', model_class='RegActRules', topic_code='last_status',
             action='new').po_code

        Num1 = post_param.get('metrics_calc')
        Num2 = post_param.get('actual_score')
        metrics_calc = None if Num1=='NaN' or Num1=='' else Num1
        actual_score = None if Num2=='NaN' or Num2=='' else Num2
        if ( last_status < pm406_last_status_new):   # 自評,皆做備份 bu_calc_content,bu_metrics_calc,bu_actual_score
            post_param.update({
                'change_time': now,
                'changer': request.user.username,
                'create_time': now,
                'creator': request.user.username,
                'order_type': order_type,
                'calc_content': '' if post_param.get('calc_content',None)=='' else post_param.get('calc_content',None),       #將''空字串, 轉為'', 才能存檔
                'metrics_calc': metrics_calc,
                'actual_score': actual_score,
                'last_status': last_status_instance,
                # 2021/08/03 add : 將自評儲存起來(主管審核覆蓋後可備查),可以和『主管審核』做比對( 備而不用 )
                'bu_calc_content': '' if post_param.get('calc_content', None) == '' else post_param.get('calc_content', None),
                'bu_metrics_calc': metrics_calc,
                'bu_actual_score': actual_score,
            })
        else:     #主管審核結果,直接存於原自評的位置
            post_param.update({
                'change_time': now,
                'changer': request.user.username,
                'create_time': now,
                'creator': request.user.username,
                'order_type': order_type,
                'calc_content': '' if post_param.get('calc_content',None)=='' else post_param.get('calc_content',None),           #將''空字串, 轉為'', 才能存檔
                'metrics_calc': metrics_calc,
                'actual_score': actual_score,
                'last_status': last_status_instance,
            })

        try:
            ScoreSheet.objects.create(**post_param)
        except ValidationError as e:
            logger.debug(e)
            if e.code == 'invalid_time':
                return JsonResponse({"success": False, "message": "\"%s\" xxx格式正確，卻非有效時間。" % e.params['value']})
            else:
                return JsonResponse({"success": False, "message": "xxx請檢查欄位數值是否接正確。"})
        except ValueError as e:
            logger.debug(e)
            return JsonResponse({"success": False, "message": "xxx欄位不能為空，且只能是數字。"})
        except pymssql.DatabaseError as e:
            logger.debug(e)
            return JsonResponse({"success": False, "message": "xxx已存在資料庫中"})
        except Exception as e:
            if post_param.get('metrics_id', None):
                pk = post_param.get('metrics_id')
                post_param.pop('metrics_id')
            if pk:
                # instance = ScoreSheet.objects.get(metrics_id=pk)    objects.get沒有update attribute
                instance = ScoreSheet.objects.filter(metrics_id=pk)
                if instance:
                    try:
                        instance.update(**post_param)
                    except ValidationError as e:
                        logger.debug(e)
                        if e.code == 'invalid_time':
                            return JsonResponse(
                                {"success": False, "message": "\"%s\" 格式正確，卻非有效時間。" % e.params['value']})
                        else:
                            return JsonResponse({"success": False, "message": "請檢查欄位數值是否接正確。"})
                    except ValueError as e:
                        logger.debug(e)
                        return JsonResponse({"success": False, "message": "欄位不能為空，且只能是數字。"})
                    else:
                        cache.clear()
                        return JsonResponse({"success": True})
        else:
            cache.clear()
            return JsonResponse({"success": True,"message":"更新成功"})

# 關帳前的檢核
def valid_all_metrics(request,pk1=None,pk2=None,pk3=None,pk4=None):
    current_tz = timezone.get_current_timezone()
    now = timezone.now().astimezone(current_tz).strftime('%Y-%m-%d %H:%M:%S')

    next_year = pk1+1 if pk2==12 else pk1
    next_month= 1 if pk2==12 else pk2+1

    # 目前評核年月
    #   (1)-主管是否已審核:找出主管未審核的MetricsSetup
    #       找出MetricsSetup的id
    fields = ['metrics_id',                 # 0
              'work_code',                  # 1
              'date_yyyy',                  # 2
              'date_mm',                    # 3
              'order_number',               # 4
              'order_item',                 # 5
              'allocation',                 # 6
              'metrics_content',            # 7
              'work_code__chi_name',        # 8
              'work_code__factory',         # 9
              'work_code__dept',            #10
              'work_code__pos',             #11
              'work_code__director',        #12
              'work_code__arrival_date',    #13
              'work_code__rank',            #14
              'work_code__nat',             #15
              'work_code__labor_type',      #16
              'work_code__bonus_type',      #17
              'work_code__bonus_factor',    #18
              'work_code__urrf',            #19
              'work_code__urrf1',           #20
              'work_code__urrf2',           #21
              'score_type',                 #22
              'work_code__email',            #23
              ]

    fields_allocation = [ 'work_code',                  # 0
                          'date_yyyy',                  # 1
                          'date_mm',                    # 2
                          'work_code__chi_name',        # 3
                          'work_code__factory',         # 4
                          'work_code__dept',            # 5
                          'work_code__pos',             # 6
                          'work_code__director',        # 7
                          'work_code__arrival_date',    # 8
                          'work_code__rank',            # 9
                          'work_code__nat',             #10
                          'work_code__labor_type',      #11
                          'work_code__bonus_type',      #12
                          'work_code__bonus_factor',    #13
                          'work_code__urrf',            #14
                          'work_code__urrf1',           #15
                          'work_code__urrf2',           #16
                          'work_code__email',           #17
                        ]

    fields_ee = [ 'work_code',                  # 0
                  'chi_name',                   # 1
                  'factory',                    # 2
                  'dept',                       # 3
                  'pos',                        # 4
                  'director',                   # 5
                  'arrival_date',               # 6
                  'rank',                       # 7
                  'nat',                        # 8
                  'labor_type',                 # 9
                  'bonus_type',                 #10
                  'bonus_factor',               #11
                  'urrf',                       #12
                  'urrf1',                      #13
                  'urrf2',                      #14
                  'email',                      #15
              ]

    commQ = Q(eval_class__in=UserDefCode.objects.filter(Q(topic_code_id='eval_class_id'), ~Q(desc1='BSC')))
    commQ1 = Q(work_code__in=EmployeeInfoEasy.objects.values_list('work_code', flat=True).filter(~Q(work_code__endswith='-000'), Q(resign_date__isnull=True),commQ) )

    #   ('A','自評實績') ,('B','主管評核') , ('C','匯入實績')
    # result_metrics = MetricsSetup.objects.values_list(*fields).filter(commQ1 , date_yyyy=pk1 , date_mm=pk2).exclude(score_type='C')
    result_metrics = MetricsSetup.objects.values_list(*fields).filter(commQ1 , date_yyyy=pk1 , date_mm=pk2 , order_number__lt = 900)   #共同衡量指標, 不算
    result_metrics_next = MetricsSetup.objects.values_list(*fields).filter(commQ1 , date_yyyy=next_year , date_mm=next_month , order_number__lt = 900)   #共同衡量指標, 不算

    result_allocation = MetricsSetup.objects.order_by('work_code', 'date_yyyy', 'date_mm').values_list(*fields_allocation).filter(commQ1 , date_yyyy=pk1 , date_mm=pk2 , order_number__lt = 900).annotate(Sum('allocation'))   #共同衡量指標, 不算
    result_allocation_next = MetricsSetup.objects.order_by('work_code', 'date_yyyy', 'date_mm').values_list(*fields_allocation).filter(commQ1 , date_yyyy=next_year , date_mm=next_month , order_number__lt = 900).annotate(Sum('allocation'))   #共同衡量指標, 不算

    # 本期的評核年月
    commQ2 = Q(work_code__in=MetricsSetup.objects.values_list('work_code',flat=True).filter(date_yyyy=pk1 , date_mm=pk2 , order_number__lt = 900).order_by('work_code','date_yyyy','date_mm').distinct() )
    result_ee = EmployeeInfoEasy.objects.values_list(*fields_ee).filter(~Q(work_code__endswith='-000'), Q(resign_date__isnull=True),commQ).exclude(commQ2)

    # 下個評核年月
    commQ3 = Q(work_code__in=MetricsSetup.objects.values_list('work_code',flat=True).filter(date_yyyy=next_year , date_mm=next_month , order_number__lt = 900).order_by('work_code','date_yyyy','date_mm').distinct() )
    result_ee_next = EmployeeInfoEasy.objects.values_list(*fields_ee).filter(~Q(work_code__endswith='-000'), Q(resign_date__isnull=True),commQ).exclude(commQ3)

    # 寫Excel
    wb = openpyxl.Workbook()
    # 建立sheet
    # sheetA = wb.create_sheet('A下期_得分無最高',0)                     #add in 2021/09/29
    # sheet9 = wb.create_sheet('9下期_得分無0',0)                       #add in 2021/09/29
    # sheet8 = wb.create_sheet('8下期_指標未滿100分',0)                  #add in 2021/09/29

    sheet8 = wb.create_sheet('8下期_計算方式錯誤',0)                      #add in 2021/09/29
    sheet7 = wb.create_sheet('7本期_計算方式錯誤',0)                      #add in 2021/09/29
    sheet6 = wb.create_sheet('6下期_指標未滿100分',0)                       #add in 2021/09/29
    sheet5 = wb.create_sheet('5本期_指標未滿100分',0)                  #add in 2021/09/29

    sheet4 = wb.create_sheet('4下期無指標',0)                         #add in 2021/09/28
    sheet3 = wb.create_sheet('3本期無指標',0)                          #add in 2021/09/28

    sheet2 = wb.create_sheet('2自評未開始',0)
    sheet1 = wb.create_sheet('1自評未結案',0)

    title3 = ['工號','姓名','部門','評核主管','職稱(位)','職等','公司','國籍','直接/間接','獎金類型','獎金點數','備註1','備註2','備註3','E-mail']    #add in 2021/09/28  本期
    title4 = title3                                                                                                                    #add in 2021/09/28  下期

    title5 = ['工號','姓名','部門','評核主管','職稱(位)','職等','公司','國籍','直接/間接','獎金類型','獎金點數','備註1','備註2','備註3','E-mail','年','月','總分','分數差距']
    title2 = ['工號','姓名','部門','評核主管','職稱(位)','職等','公司','國籍','直接/間接','獎金類型','獎金點數','備註1','備註2','備註3','年','月','順序','細項','配分','衡量指標','評核方式','E-mail','備註']
    title1 = ['工號','姓名','部門','評核主管','職稱(位)','職等','公司','國籍','直接/間接','獎金類型','獎金點數','備註1','備註2','備註3','年','月','順序','細項','配分','衡量指標','計算方式','得分','實績', '狀態碼','狀態內容','評核方式','E-mail']

    sheet8.append(title2)
    sheet7.append(title2)
    sheet6.append(title5)
    sheet5.append(title5)

    sheet1.append(title1)

    Color = ['000000','ffffff','ff0000','0000ff','00ff00','ffff00','000055']  # 黑,白,紅,藍,綠,黃
    fille = PatternFill('solid', fgColor=Color[6])
    font = Font(u'微軟正黑體', size=12, bold=True, italic=False, strike=False, color=Color[5])  # 设置字体样式

    for i in range(0,len(title1)):
        sheet1.cell(row=1, column=i+1).fill = fille
        sheet1.cell(row=1, column=i+1).font = font

    sheet2.append(title2)
    for i in range(0,len(title2)):
        sheet2.cell(row=1, column=i+1).fill = fille
        sheet2.cell(row=1, column=i+1).font = font

        sheet8.cell(row=1, column=i+1).fill = fille
        sheet8.cell(row=1, column=i+1).font = font

        sheet7.cell(row=1, column=i+1).fill = fille
        sheet7.cell(row=1, column=i+1).font = font


    for i in range(0, len(title5)):
        sheet6.cell(row=1, column=i+1).fill = fille
        sheet6.cell(row=1, column=i+1).font = font

        sheet5.cell(row=1, column=i+1).fill = fille
        sheet5.cell(row=1, column=i+1).font = font

    sheet3.append(title3)
    sheet4.append(title4)
    for i in range(0,len(title3)):
        sheet3.cell(row=1, column=i+1).fill = fille
        sheet3.cell(row=1, column=i+1).font = font

        sheet4.cell(row=1, column=i+1).fill = fille
        sheet4.cell(row=1, column=i+1).font = font

    # 檢核錯誤筆數  計算開始
    error_count = 0
    for TT in result_ee:
        # sheet3 : 本期無指標
        error_count += 1
        factory_desc = Factory.objects.get(id=TT[2]).name
        dept_desc = UserDefCode.objects.get(id=TT[3]).desc1
        pos_desc = UserDefCode.objects.get(id=TT[4]).desc1
        director_desc = "" if TT[5] == None else EmployeeInfoEasy.objects.get(work_code=TT[5]).chi_name
        rank_desc = UserDefCode.objects.get(id=TT[7]).desc1
        nat_desc = UserDefCode.objects.get(id=TT[8]).desc1
        labor_type_desc = UserDefCode.objects.get(id=TT[9]).desc1
        bonus_type_desc = "" if TT[10] == None else UserDefCode.objects.get(id=TT[10]).desc1
        bonus_factor_desc = UserDefCode.objects.get(id=TT[11]).desc1
        rowZ = [TT[0],TT[1],dept_desc,director_desc,pos_desc,rank_desc,factory_desc,nat_desc,labor_type_desc,
                        bonus_type_desc,bonus_factor_desc,TT[12],TT[13],TT[14],TT[15],]
        sheet3.append(rowZ)

    for TT in result_ee_next:
        # sheet4 : 下期無指標
        error_count += 1
        factory_desc = Factory.objects.get(id=TT[2]).name
        dept_desc = UserDefCode.objects.get(id=TT[3]).desc1
        pos_desc = UserDefCode.objects.get(id=TT[4]).desc1
        director_desc = "" if TT[5] == None else EmployeeInfoEasy.objects.get(work_code=TT[5]).chi_name
        rank_desc = UserDefCode.objects.get(id=TT[7]).desc1
        nat_desc = UserDefCode.objects.get(id=TT[8]).desc1
        labor_type_desc = UserDefCode.objects.get(id=TT[9]).desc1
        bonus_type_desc = "" if TT[10] == None else UserDefCode.objects.get(id=TT[10]).desc1
        bonus_factor_desc = UserDefCode.objects.get(id=TT[11]).desc1
        rowZ = [TT[0],TT[1],dept_desc,director_desc,pos_desc,rank_desc,factory_desc,nat_desc,labor_type_desc,
                        bonus_type_desc,bonus_factor_desc,TT[12],TT[13],TT[14],TT[15],]
        sheet4.append(rowZ)

    for TX in result_allocation:
        if (TX[18] < 100):
            # sheet5 : 本期指標未滿100
            error_count += 1
            factory_desc = Factory.objects.get(id=TX[4]).name
            dept_desc = UserDefCode.objects.get(id=TX[5]).desc1
            pos_desc = UserDefCode.objects.get(id=TX[6]).desc1
            director_desc = "" if TX[7] == None else EmployeeInfoEasy.objects.get(work_code=TX[7]).chi_name
            rank_desc = UserDefCode.objects.get(id=TX[9]).desc1
            nat_desc = UserDefCode.objects.get(id=TX[10]).desc1
            labor_type_desc = UserDefCode.objects.get(id=TX[11]).desc1
            bonus_type_desc = "" if TX[12] == None else UserDefCode.objects.get(id=TX[12]).desc1
            bonus_factor_desc = UserDefCode.objects.get(id=TX[13]).desc1

            rowZ = [TX[0], TX[3], dept_desc, director_desc, pos_desc, rank_desc, factory_desc, nat_desc, labor_type_desc,
                    bonus_type_desc, bonus_factor_desc, TX[14], TX[15], TX[16],TX[17], TX[1], TX[2],TX[18], 100-TX[18] ]
            sheet5.append(rowZ)

    for TX in result_allocation_next:
        if (TX[18] < 100):
            # sheet6 : 下期指標未滿100
            error_count += 1
            factory_desc = Factory.objects.get(id=TX[4]).name
            dept_desc = UserDefCode.objects.get(id=TX[5]).desc1
            pos_desc = UserDefCode.objects.get(id=TX[6]).desc1
            director_desc = "" if TX[7] == None else EmployeeInfoEasy.objects.get(work_code=TX[7]).chi_name
            rank_desc = UserDefCode.objects.get(id=TX[9]).desc1
            nat_desc = UserDefCode.objects.get(id=TX[10]).desc1
            labor_type_desc = UserDefCode.objects.get(id=TX[11]).desc1
            bonus_type_desc = "" if TX[12] == None else UserDefCode.objects.get(id=TX[12]).desc1
            bonus_factor_desc = UserDefCode.objects.get(id=TX[13]).desc1

            rowZ = [TX[0], TX[3], dept_desc, director_desc, pos_desc, rank_desc, factory_desc, nat_desc, labor_type_desc,
                    bonus_type_desc, bonus_factor_desc, TX[14], TX[15], TX[16],TX[17], TX[1], TX[2],TX[18], 100-TX[18] ]
            sheet6.append(rowZ)

    for T in result_metrics_next:
        count_score0 = MetricsCalc.objects.filter(metrics_id=T[0], score=0).count()
        count_scoreX = MetricsCalc.objects.filter(metrics_id=T[0], score=T[6]).count()
        scoreMax = MetricsCalc.objects.values_list('score',flat=True).filter(metrics_id=T[0]).annotate(Max('score'))

        factory_desc = Factory.objects.get(id=T[9]).name
        dept_desc = UserDefCode.objects.get(id=T[10]).desc1
        pos_desc = UserDefCode.objects.get(id=T[11]).desc1
        director_desc = "" if T[12] == None else EmployeeInfoEasy.objects.get(work_code=T[12]).chi_name
        rank_desc = UserDefCode.objects.get(id=T[14]).desc1
        nat_desc = UserDefCode.objects.get(id=T[15]).desc1
        labor_type_desc = UserDefCode.objects.get(id=T[16]).desc1
        bonus_type_desc = "" if T[17] == None else UserDefCode.objects.get(id=T[17]).desc1
        bonus_factor_desc = UserDefCode.objects.get(id=T[18]).desc1
        score_type = '自評實績' if T[22] == 'A' else ('主管評核' if T[22] == 'B' else ('匯入實績' if T[22] == 'C' else ''))

        if count_score0==0:
            # sheet8 : 下期 得分無0
            error_count += 1
            rowA = [T[1], T[8], dept_desc, director_desc, pos_desc, rank_desc, factory_desc, nat_desc, labor_type_desc,
                    bonus_type_desc, bonus_factor_desc, T[19], T[20], T[21], T[2], T[3], T[4], T[5], T[6],
                    T[7].replace("\r", " "), T[22] + ' ' + score_type, T[23] , '得分無0']
            sheet8.append(rowA)

        if count_scoreX==0:
            # sheet8 : 下期 得分無最高
            error_count += 1
            rowB = [T[1], T[8], dept_desc, director_desc, pos_desc, rank_desc, factory_desc, nat_desc, labor_type_desc,
                    bonus_type_desc, bonus_factor_desc, T[19], T[20], T[21], T[2], T[3], T[4], T[5], T[6],
                    T[7].replace("\r", " "), T[22] + ' ' + score_type, T[23] , '得分無"最高"配分']
            sheet8.append(rowB)

        if ( scoreMax ):
            if scoreMax[0] > T[6]:
                # sheet8 : 下期 最高得分比衡量指標的配分高
                error_count += 1
                rowC = [T[1], T[8], dept_desc, director_desc, pos_desc, rank_desc, factory_desc, nat_desc, labor_type_desc,
                        bonus_type_desc, bonus_factor_desc, T[19], T[20], T[21], T[2], T[3], T[4], T[5], T[6],
                        T[7].replace("\r", " "), T[22] + ' ' + score_type, T[23] , '得分( '+str(scoreMax[0])+' ) 大於"最高"配分']
                sheet8.append(rowC)
        else:
            # sheet8 : 下期 最高得分比衡量指標的配分高
            error_count += 1
            rowC = [T[1], T[8], dept_desc, director_desc, pos_desc, rank_desc, factory_desc, nat_desc, labor_type_desc,
                    bonus_type_desc, bonus_factor_desc, T[19], T[20], T[21], T[2], T[3], T[4], T[5], T[6],
                    T[7].replace("\r", " "), T[22] + ' ' + score_type, T[23], '無計算方式']
            sheet8.append(rowC)

    for T in result_metrics:
        count_score0 = MetricsCalc.objects.filter(metrics_id=T[0], score=0).count()
        count_scoreX = MetricsCalc.objects.filter(metrics_id=T[0], score=T[6]).count()
        scoreMax = MetricsCalc.objects.values_list('score',flat=True).filter(metrics_id=T[0]).annotate(Max('score'))

        factory_desc = Factory.objects.get(id=T[9]).name
        dept_desc = UserDefCode.objects.get(id=T[10]).desc1
        pos_desc = UserDefCode.objects.get(id=T[11]).desc1
        director_desc = "" if T[12] == None else EmployeeInfoEasy.objects.get(work_code=T[12]).chi_name
        rank_desc = UserDefCode.objects.get(id=T[14]).desc1
        nat_desc = UserDefCode.objects.get(id=T[15]).desc1
        labor_type_desc = UserDefCode.objects.get(id=T[16]).desc1
        bonus_type_desc = "" if T[17] == None else UserDefCode.objects.get(id=T[17]).desc1
        bonus_factor_desc = UserDefCode.objects.get(id=T[18]).desc1
        score_type = '自評實績' if T[22] == 'A' else ('主管評核' if T[22] == 'B' else ('匯入實績' if T[22] == 'C' else ''))

        if count_score0==0:
            # sheet７ : 得分無0
            error_count += 1
            rowA = [T[1], T[8], dept_desc, director_desc, pos_desc, rank_desc, factory_desc, nat_desc, labor_type_desc,
                    bonus_type_desc, bonus_factor_desc, T[19], T[20], T[21], T[2], T[3], T[4], T[5], T[6],
                    T[7].replace("\r", " "), T[22] + ' ' + score_type, T[23] , '得分無0']
            sheet7.append(rowA)

        if count_scoreX==0:
            # sheet７ : 得分無最高
            error_count += 1
            rowB = [T[1], T[8], dept_desc, director_desc, pos_desc, rank_desc, factory_desc, nat_desc, labor_type_desc,
                    bonus_type_desc, bonus_factor_desc, T[19], T[20], T[21], T[2], T[3], T[4], T[5], T[6],
                    T[7].replace("\r", " "), T[22] + ' ' + score_type, T[23] , '得分無"最高"配分']
            sheet7.append(rowB)

        if ( scoreMax):
            if ( scoreMax[0] > T[6] ):
            # sheet７ : 最高得分比衡量指標的配分高
                error_count += 1
                rowC = [T[1], T[8], dept_desc, director_desc, pos_desc, rank_desc, factory_desc, nat_desc, labor_type_desc,
                        bonus_type_desc, bonus_factor_desc, T[19], T[20], T[21], T[2], T[3], T[4], T[5], T[6],
                        T[7].replace("\r", " "), T[22] + ' ' + score_type, T[23] , '得分( '+str(scoreMax[0])+' ) 大於"最高"配分']
                sheet7.append(rowC)
        else:
            rowC = [T[1], T[8], dept_desc, director_desc, pos_desc, rank_desc, factory_desc, nat_desc, labor_type_desc,
                    bonus_type_desc, bonus_factor_desc, T[19], T[20], T[21], T[2], T[3], T[4], T[5], T[6],
                    T[7].replace("\r", " "), T[22] + ' ' + score_type, T[23], '無計算方式']
            sheet7.append(rowC)

        count = 0
        po_code = ProcessOptionsTxtDef.objects.get(app_model='KPI',view_code='PM610',action='return').po_code    #找出處理選項的狀態碼
        try:
            count = ScoreSheet.objects.filter(metrics_id=T[0]).count()
        except:
            pass
        else:
            if count==0:
                # sheet2 : 自評未開始
                error_count += 1
                rowX = [T[1],T[8],dept_desc,director_desc,pos_desc,rank_desc,factory_desc,nat_desc,labor_type_desc,
                        bonus_type_desc,bonus_factor_desc,T[19],T[20],T[21],T[2],T[3],T[4],T[5],T[6],
                        T[7].replace("\r"," "),T[22]+' '+score_type,T[23]]
                sheet2.append(rowX)
            else:
                # sheet1 : 自評未結案
                fields = ['metrics_id',  #0
                          'work_code',  #1
                          'date_yyyy',  #2
                          'date_mm',  #3
                          'order_number',  #4
                          'order_item',  #5
                          'allocation',  #6
                          'metrics_content',  #7
                          'score_sheet__calc_content',  #8
                          'score_sheet__metrics_calc',  #9
                          'score_sheet__actual_score',  #10
                          'score_sheet__last_status',  #11
                          'work_code__chi_name',  #12
                          'work_code__factory',  # 13
                          'work_code__dept',  # 14
                          'work_code__pos',  # 15
                          'work_code__director',  # 16
                          'work_code__arrival_date',  # 17
                          'work_code__rank',  # 18
                          'work_code__nat',  # 19
                          'work_code__labor_type',  # 20
                          'work_code__bonus_type',  # 21
                          'work_code__bonus_factor',  # 22
                          'work_code__urrf',  # 23
                          'work_code__urrf1',  # 24
                          'work_code__urrf2',  # 25
                          'score_type', #26
                          'work_code__email',          #27
                          ]
                results = MetricsSetup.objects.filter(metrics_id=T[0]).values_list(*fields).order_by('work_code', 'date_yyyy', 'date_mm', 'order_number', 'order_item')
                for TT in results:
                    factory_desc = Factory.objects.get(id=TT[13]).name
                    dept_desc = UserDefCode.objects.get(id=TT[14]).desc1
                    pos_desc = UserDefCode.objects.get(id=TT[15]).desc1
                    director_desc =  "" if TT[16]==None else EmployeeInfoEasy.objects.get(work_code=TT[16]).chi_name
                    rank_desc = UserDefCode.objects.get(id=TT[18]).desc1
                    nat_desc = UserDefCode.objects.get(id=TT[19]).desc1
                    labor_type_desc = UserDefCode.objects.get(id=TT[20]).desc1
                    bonus_type_desc = UserDefCode.objects.get(id=TT[21]).desc1
                    bonus_factor_desc = UserDefCode.objects.get(id=TT[22]).desc1
                    last_status = RegActRules.objects.get(id=TT[11]).last_status  # 最後狀態碼
                    score_type = '自評實績' if TT[26] == 'A' else ('主管評核' if TT[26] == 'B' else ('匯入實績' if TT[26] == 'C' else ''))
                    rowY = [TT[1],TT[12],dept_desc,director_desc,pos_desc,rank_desc,factory_desc,nat_desc,labor_type_desc,bonus_type_desc,bonus_factor_desc,TT[23],TT[24],TT[25],TT[2], TT[3], TT[4],
                            TT[5],TT[6],None if TT[7]==None else TT[7].replace("\r"," ") ,None if TT[8]==None else TT[8].replace("\r"," "),TT[9],
                            TT[10],last_status,RegActRules.objects.get(id=TT[11]).status_desc,TT[26]+' '+score_type,TT[27]]
                    if (last_status<po_code):     #少於BPM已簽核完的狀態碼, 才列出
                        # sheet1 : 自評未結案
                        sheet1.append(rowY)
                        error_count += 1

    err_release = WorkingYM.objects.values_list('err_release').get(id=1)
    if err_release[0] is True:        # 若為true,則放行
        error_count=0
    if (error_count==0):
        # 都無錯誤，檢核完成，進行關帳，狀態改至９９９
        update_params = {}
        po_code_close = ProcessOptionsTxtDef.objects.get(app_model='KPI', view_code='PM203',action='close').po_code  # 找出處理選項的狀態碼
        last_status_close_id = RegActRules.objects.get(last_status=po_code_close).id
        commQx = Q(metrics_id__in=MetricsSetup.objects.filter(commQ1, date_yyyy=pk1, date_mm=pk2).order_by('work_code', 'date_yyyy', 'date_mm'))

        instance = ScoreSheet.objects.filter(commQx)
        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz)
        update_params.update({
            'change_time': now,
            'changer': request.user.username,
            'last_status': last_status_close_id,
        })
        instance.update(**update_params)
        gen_score_sheet(next_year, next_month, request.user.username, now)  # 產生『下個評核核年月』的, 無法『自評』的工作底稿
        return JsonResponse({"success": True})
    else:
        current_tz = timezone.get_current_timezone()
        now = timezone.now().astimezone(current_tz).strftime('%Y%m%d_%H%M%S')
        filename = 'pm203_validAllMetrics_' + now + '.xlsx'
        output_file = '%s/reports/xlsx/%s' % (settings.MEDIA_ROOT, filename)
        wb.save(output_file)
        openFile = '/media-files/reports/xlsx/'+filename
        return JsonResponse({"success": False,"filename":filename,"openFile":openFile})



def gen_score_sheet(pk1,pk2,username,now):
    # 產生『下個評核核年月』的, 無法『自評』的工作底稿
    commQ = Q(eval_class__in=UserDefCode.objects.filter(Q(topic_code_id='eval_class_id'), ~Q(desc1='BSC')))
    commQ1 = Q(work_code__in=EmployeeInfoEasy.objects.values_list('work_code', flat=True).filter(~Q(work_code__endswith='-000'),Q(kpi_diy=False), Q(resign_date__isnull=True),commQ) )

    #   ('A','自評實績') ,('B','主管評核') , ('C','匯入實績')
    # result_metrics = MetricsSetup.objects.values_list(*fields).filter(commQ1 , date_yyyy=pk1 , date_mm=pk2).exclude(score_type='C')
    # 沒有設定衡量指標的人, 不會有工作底稿
    # result_metrics = MetricsSetup.objects.values_list('metrics_id').filter(commQ1 , date_yyyy=pk1 , date_mm=pk2 , order_number__lt = 900)   #共同衡量指標, 不算
    result_metrics = MetricsSetup.objects.filter(commQ1 , date_yyyy=pk1 , date_mm=pk2 , order_number__lt = 900)   #共同衡量指標, 不算

    po_code = ProcessOptionsTxtDef.objects.get(app_model='KPI', view_code='PM402',action='self_X').po_code  # 不可自評的狀態碼
    last_status_instance = RegActRules.objects.get(last_status=po_code)  # 不可自評的狀態碼

    bulk_data = [ ScoreSheet(
        change_time=now,
        changer=username,
        create_time=now,
        creator=username,
        order_type='KPI',
        metrics= instance,
        last_status=last_status_instance,
    ) for instance in result_metrics ]

    try:
        ScoreSheet.objects.bulk_create(bulk_data)
    except ValidationError as e:
        logger.debug(e)
        if e.code == 'invalid_time':
            return False
        else:
            return True
    except ValueError as e:
        logger.debug(e)
        return False
    except pymssql.DatabaseError as e:
        logger.debug(e)
        return False
    except pymssql.IntegrityError as e:
        logger.debug(e)
        return False
    except Exception as e:
        logger.debug(e)
        return False
    else:
        return True



def set_score_sheet_status(request,pk1=None,pk2=None,pk3=None):
    get_param = request.POST.dict()
    order_type = ProcessOptionsTxtDef.objects.get \
        (app_model='KPI', view_code=pk2, model_class='RegActRules', topic_code='order_type',
         action='blank').po_code
    this_status = ProcessOptionsTxtDef.objects.get \
        (app_model='KPI', view_code=pk2, model_class='RegActRules', topic_code='last_status',
         action=pk3).po_code
    get_param.update({
        'change_time': now,
        'changer': request.user.username,
        'create_time': now,
        'creator': request.user.username,
        'last_status': RegActRules.objects.get(last_status=this_status) if this_status else None,
    })
    if pk1:
        instance = ScoreSheet.objects.filter(metrics_id=pk1)
        if instance:
            try:
                instance.update(**get_param)
            except ValidationError as e:
                logger.debug(e)
                if e.code == 'invalid_time':
                    return JsonResponse(
                        {"success": False, "message": "\"%s\" 格式正確，卻非有效時間。" % e.params['value']})
                else:
                    return JsonResponse({"success": False, "message": "請檢查欄位數值是否接正確。"})
            except ValueError as e:
                logger.debug(e)
                return JsonResponse({"success": False, "message": "欄位不能為空，且只能是數字。"})
            else:
                cache.clear()
                return JsonResponse({"success": True})
        else:
            return JsonResponse({"success": False})


#(人事基本資料  20210531修改)取得下屬的mail, pk1:下屬的工號
def get_score_sheet_sendmail(request,pk1=None,pk2=None,pk3=None):
    if pk1:
        try:
            mail_address=EmployeeInfoEasy.objects.get(work_code=pk1).email
            # send_mail--------------------------------------------------------------------------------------Begin
            try:
                mail_subject = "*** "+str(pk2)+"年"+str(pk3)+"月KPI ***"
                report_month = [3,6,9,12]
                if pk3 in report_month:
                    mail_messages = "主管已審核完成\n\n＊＊＊ 請注意:季報表已送至BPM ＊＊＊"
                else:
                    mail_messages = "主管已審核完成"

                mail_results = send_mail(
                                    mail_subject,
                                    mail_messages,
                                    settings.DEFAULT_FROM_EMAIL,
                                    [mail_address],
                                    fail_silently=False,
                                )
                cache.clear()
                return JsonResponse({"success": True})
            except BadHeaderError:
                return JsonResponse({"success": False})
            #send_mail--------------------------------------------------------------------------------------Ending
        except:
            cache.clear()
            return JsonResponse({"success": False})


#(人事基本資料  20210531修改)取得下屬的mail, pk1:下屬的工號
def get_score_sheet_sendmail_to_director(request,pk1=None,pk2=None,pk3=None):
    cache.clear()
    if pk1:
        try:
            chi_name = EmployeeInfoEasy.objects.get(work_code=pk1).chi_name
            director = EmployeeInfoEasy.objects.get(work_code=pk1).director_id
            mail_address=EmployeeInfoEasy.objects.get(work_code=director).email
            # send_mail--------------------------------------------------------------------------------------Begin
            try:
                mail_subject = "*** "+pk1+chi_name+"，"+str(pk2)+"年"+str(pk3)+"月，KPI自評已完成 ***"
                report_month = [3, 6, 9, 12]
                if pk3 in report_month:
                    mail_messages = "請登入績效管理系統\nPM406 KPI主管審核\n評核實績/儲存更新/送出\n以完成評核作業\n\n＊＊＊ 請注意:此月在送出後，有季報表送至BPM簽核 ＊＊＊"
                else:
                    mail_messages = "請登入績效管理系統\nPM406 KPI主管審核\n評核實績/儲存更新/送出\n以完成評核作業"
                mail_results = send_mail(
                                    mail_subject,
                                    mail_messages,
                                    settings.DEFAULT_FROM_EMAIL,
                                    [mail_address],
                                    fail_silently=False,
                                )
                return JsonResponse({"success": True})
            except BadHeaderError:
                return JsonResponse({"success": False})
            #send_mail--------------------------------------------------------------------------------------Ending
        except:
            cache.clear()
            return JsonResponse({"success": False})


#order_number自動增加
def get_metrics_order_number(request,pk1=None,pk2=None,pk3=None):
    commQ = (Q(work_code=pk1) & Q(date_yyyy=pk2) & Q(date_mm=pk3))
    # result = MetricsSetup.objects.filter(commQ).values_list('order_number',).order_by('work_code','date_yyyy','date_mm' ,'-order_number')[0]
    result = MetricsSetup.objects.filter(commQ,order_number__lt=900).values_list('order_number',).order_by('work_code','date_yyyy','date_mm' ,'order_number').last()
    if result:
        next_number = result[0] + 1
    else:
        next_number = 1
    return JsonResponse({"next_number": next_number})


def get_metrics_order_item(request,pk1=None,pk2=None,pk3=None,pk4=None):
    commQ = (Q(work_code=pk1) & Q(date_yyyy=pk2) & Q(date_mm=pk3) & Q(order_number=pk4))
    result = MetricsSetup.objects.filter(commQ).values_list('order_item',).order_by('work_code','date_yyyy','date_mm' ,'order_number','order_item').last()
    if result:
        next_number = result[0] + 1
    else:
        next_number = 0
    return JsonResponse({"next_number": next_number})



#order_number自動增加
def get_metricsCalc_order_number(request,pk1=None):
    fieldList = [ 'order_number',
                  'lower_limit',
                  'upper_limit',
                  'score'
                 ]
    result = MetricsCalc.objects.filter(metrics_id=pk1).values_list(*fieldList).order_by('metrics_id','order_number').last()
    res_param = {}
    if result:
        next_number = result[0] + 1
        res_param.update({
                'next_number': next_number,
                'lower_limit': result[1],
                'upper_limit': result[2],
                'score': result[3],
            })
    else:
        next_number= 1
        res_param.update({
                'next_number': next_number,
                'lower_limit': -999,
                'upper_limit': -999,
                'score': -999,
            })
    return JsonResponse(res_param)



def get_common_udc(request,pk=None):
    fieldList = [ 'id',
                  'udc',
                  'desc1',
                  'desc2',
                 ]
    results = UserDefCode.objects.filter(topic_code_id=pk).values_list(*fieldList).order_by('topic_code_id','udc')
    dataList = []
    for dataTuple in results:
        dataList.append({
            'value': dataTuple[0],
            # 'text': str(dataTuple[0])+" "+dataTuple[1],
            'text': str(dataTuple[1])+" "+dataTuple[2],
        })
    return JsonResponse(dataList,safe=False)



def list_filter_dict(list_object,key):
    for itm in list_object:
        if itm.get('work_code')==key:
            director = itm.get('director',None)
            director_supv = itm.get('direct_supv',None)
            rtn_data = (
                itm.get('work_code'),
                itm.get('chi_name'),
                Factory.objects.get(id=itm.get('factory',None)).name,
                UserDefCode.objects.get(id=itm.get('nat')).udc,
                UserDefCode.objects.get(id=itm.get('service_status')).desc1,
                UserDefCode.objects.get(id=itm.get('factory_area')).desc1,
                itm.get('email'),
                # UserDefCode.objects.get(id=itm.get('kpi_diy')).desc1,
                itm.get('kpi_diy'),
                UserDefCode.objects.get(id=itm.get('dept_flevel')).udc,
                UserDefCode.objects.get(id=itm.get('dept_flevel')).desc1,
                UserDefCode.objects.get(id=itm.get('dept')).udc,
                UserDefCode.objects.get(id=itm.get('dept')).desc1,
                # UserDefCode.objects.get(id=itm.get('dept_desc')).desc1,
                itm.get('dept_description'),     #2022/05/11起,部門全稱,直接依『HR人事基本資料』
                director,
                EmployeeInfoEasy.objects.get(work_code=itm.get('director')).chi_name if director else "",
                director_supv,
                EmployeeInfoEasy.objects.get(work_code=itm.get('direct_supv')).chi_name if director_supv else "",
                UserDefCode.objects.get(id=itm.get('pos')).desc1,
                UserDefCode.objects.get(id=itm.get('labor_type')).desc1,
                itm.get('arrival_date',None),
                itm.get('resign_date',None),
                itm.get('trans_date',None),
                itm.get('trans_type',None),
                UserDefCode.objects.get(id=itm.get('bonus_type')).desc1,
                UserDefCode.objects.get(id=itm.get('bonus_factor')).desc1,
                UserDefCode.objects.get(id=itm.get('rank')).desc1,
                UserDefCode.objects.get(id=itm.get('eval_class')).desc1,
            )
            return rtn_data
    return False


def search_supervisor(work_code_x,dept_code,manager_type):
    try:
        dept_id = UserDefCode.objects.get(udc=dept_code).id
    except:
        dept_id = 0

    if dept_id==0:
        return ""
    else:
        try :
            r = model_to_dict(DeptSupervisor.objects.get(dept_id=dept_id))
            supervisor_x = r.get('dept_supervisor')
            if manager_type=="評核主管":
                if r.get('dept') == r.get('dept_flevel'):
                    return supervisor_x
                else:
                    search_supervisor(work_code_x,r.get('dept_flevel'), manager_type)
            elif manager_type=="直屬主管":
                if  supervisor_x==work_code_x or supervisor_x==None or supervisor_x=="":     #主管工號==員工工號
                    search_supervisor(work_code_x, r.get('dept_upper'), manager_type)
                else:
                    return supervisor_x
        except:
            return ""


def synchronize_hr_employee(request):
    # 定義excel檔的sheet,寫入sheet-----------------------------------------------------------------------------------------Begin
    Color = ['000000', 'ffffff', 'ff0000', '0000ff', '00ff00', 'ffff00', '000055','ffffaa']
    #          0黑,       1白,       2紅,      3藍,      4綠,       5黃      6BlackBlue
    fille = PatternFill('solid', fgColor=Color[0])
    fille_diff = PatternFill('solid', fgColor=Color[5])
    font_efields = Font(u'微軟正黑體', size=10.5, color=Color[3])  # 设置字体样式
    font_etitles = Font(u'微軟正黑體', size=11.5, bold=True, color=Color[2])  # 设置字体样式  # 设置字体样式
    font_with_fill = Font(u'微軟正黑體', size=11.5, bold=False, color=Color[1])  # 设置字体样式  # 设置字体样式
    font_with_fill_diff = Font(u'微軟正黑體', size=11.5, bold=False, color=Color[0])  # 设置字体样式  # 设置字体样式
    font_data = Font(u'微軟正黑體', size=12, bold=True, italic=False, strike=False, color=Color[0])  # 设置字体样式
    align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    efields = ['work_code', 'chi_name', 'factory_id', 'nat_id', 'service_status_id', 'factory_area_id', 'email',
               'kpi_diy','dept_flevel_id','--dept_flevel', 'dept_id','--dept', 'dept_description', 'director_id', '--direct', 'direct_supv_id',
               '--direct_supv', 'pos_id', 'labor_type_id', 'arrival_date', 'resign_date', 'trans_date', 'trans_type',
               'bonus_type_id', 'bonus_factor_id', 'rank_id', 'eval_class_id']
    etitles = ['工號', '姓名', '公司', '國籍', '服務狀態', '廠區', 'E-mail', '自評', '部門編號','--一級部門', '部門編號','--部門別', '部門全稱', '評核主管工號', '--評核主管',
               '直接主管工號', '--直接主管', '職稱(位)', '直接/間接', '到職日', '離職日', '異動日', '異動類別', '獎金型態', '獎金點數', '職等', 'BSC/KPI']
    wb = openpyxl.Workbook()

    # 三個sheet,  一個"Add新增", 一個"Change異動", 一個"HR_ALL全部"
    sheetHR = wb.create_sheet("HR_All", 0)
    sheetHR.append(efields)
    sheetHR.append(etitles)

    sheetC = wb.create_sheet("Change", 0)
    sheetC.append(efields+['＊']+efields)
    sheetC.append(etitles+['＊']+etitles)

    sheetA = wb.create_sheet("Add", 0)
    sheetA.append(efields)
    sheetA.append(etitles)

    title_len = len(efields)
    for i in range(0, title_len ):
        if efields[i].find("--") > -1:  # --匯入時要刪除
            sheetA.cell(row=1, column=i + 1).fill = fille
            sheetC.cell(row=1, column=i + 1).fill = fille
            sheetHR.cell(row=1, column=i + 1).fill = fille

            sheetA.cell(row=1, column=i + 1).font = font_with_fill
            sheetC.cell(row=1, column=i + 1).font = font_with_fill
            sheetHR.cell(row=1, column=i + 1).font = font_with_fill

            sheetA.cell(row=2, column=i + 1).fill = fille
            sheetC.cell(row=2, column=i + 1).fill = fille
            sheetHR.cell(row=2, column=i + 1).fill = fille

            sheetA.cell(row=2, column=i + 1).font = font_with_fill
            sheetC.cell(row=2, column=i + 1).font = font_with_fill
            sheetHR.cell(row=2, column=i + 1).font = font_with_fill
        else:
            sheetA.cell(row=1, column=i + 1).font = font_efields
            sheetA.cell(row=1, column=i + 1).alignment = align
            sheetA.cell(row=2, column=i + 1).font = font_etitles
            sheetA.cell(row=2, column=i + 1).alignment = align

            sheetC.cell(row=1, column=i + 1).font = font_efields
            sheetC.cell(row=1, column=i + 1).alignment = align
            sheetC.cell(row=2, column=i + 1).font = font_etitles
            sheetC.cell(row=2, column=i + 1).alignment = align

            sheetHR.cell(row=1, column=i + 1).font = font_efields
            sheetHR.cell(row=1, column=i + 1).alignment = align
            sheetHR.cell(row=2, column=i + 1).font = font_etitles
            sheetHR.cell(row=2, column=i + 1).alignment = align

        title2_len = title_len*2+1
        for i in range(title_len+2 , title2_len):
            if efields[i-title_len-2].find("--") > -1:  # --匯入時要刪除
                sheetC.cell(row=1, column=i ).fill = fille
                sheetC.cell(row=1, column=i ).font = font_with_fill
                sheetC.cell(row=2, column=i).fill = fille
                sheetC.cell(row=2, column=i).font = font_with_fill
            else:
                sheetC.cell(row=1, column=i).font = font_efields
                sheetC.cell(row=1, column=i).alignment = align
                sheetC.cell(row=2, column=i).font = font_etitles
                sheetC.cell(row=2, column=i).alignment = align
    # 定義excel檔的sheet,寫入sheet-----------------------------------------------------------------------------------------Ending

    #抓pms employeeInfoEasy----------------------------------------------------------------------------------------------Begin
    pms_employee = EmployeeInfoEasy.objects.all()
    pms_data = []
    for idx, itm in enumerate(pms_employee):
        pms_dict = model_to_dict(itm)
        pms_data.append(pms_dict)
    # 抓pms employeeInfoEasy---------------------------------------------------------------------------------------------Ending

    # hr : View_Employee_Short-----------------export to python----------------------------------------------------------Begin
    sql = """select a.* 
            from View_Employee_Short a
            order by a.Code
          """
    # batch_db_setting = settings.DATABASES['batch']
    # conn = pyodbc.connect(batch_db_setting['CONNECT_STR'],timeout=300)
    conn = connections['external_db_hr']
    cursor_hr = conn.cursor()
    cursor_hr.execute(sql)
    records = cursor_hr.fetchall()
    insertObject = []
    columnNames = [column[0] for column in cursor_hr.description]
    for record in records:
        insertObject.append(dict(zip(columnNames, record)))

    hr_all_data = []
    # for row in insertObject:
    change_in_row = 2
    print("\n" * 2)
    print("=" * 150)
    for row in insertObject:
        #HR全倒出
        work_code_hr = row.get("Code")
        department_id = row.get("DepartmentID", "")
        department = row.get("Department", "")
        direct_department_id = row.get("DirectDepartmentID", "")    #一級部門編號
        direct_department = row.get("DirectDepartment", "")    #一級部門編號

        pos_desc = row.get("Job", "")
        find_pos = 0 if pos_desc == None else pos_desc.find("等")
        if find_pos == 0:
            # None
            pos_desc = ""
        elif find_pos == -1:
            # 沒有"等"
            pass
        else:
            # 有找到"等"
            pos_desc = pos_desc[find_pos+1 :len(pos_desc)]

        director_id = row.get("DirectorID")
        director_name = row.get("Director")
        if director_id:
            # 尋找台幹對應的台灣工號
            r = WorkcodeMapping.objects.filter(work_code=director_id).values_list('work_code_x_id')
            if len(r)>0 :
                work_code_x = r[0][0]
                director_id = work_code_x
                director_name = EmployeeInfoEasy.objects.get(work_code=director_id).chi_name


        if director_id=="" or director_id==None:     #HR沒有主管時, 自動尋找主管
            try:
                work_code_data = model_to_dict(EmployeeInfoEasy.objects.get(work_code=work_code_hr),fields=['chi_name', 'urrf'])
                if work_code_data.get('urrf') == "***":
                    is_boss = True
                else:
                    is_boss = False
            except:
                is_boss = False

            if is_boss:
                director_id_A = work_code_hr
                director_name_A = work_code_data.get("chi_name")
                director_id_B = director_id_A
                director_name_B = director_name_A
            else:
                # 找評核主管
                director_id_A = search_supervisor(work_code_hr, department_id, '評核主管')
                if director_id_A == "" or director_id_A == None:
                    director_name_A = ""
                else:
                    try:
                        director_name_A = EmployeeInfoEasy.objects.get(work_code=director_id_A).chi_name
                        # print("評核主管:",director_id_A,director_name_A)
                    except:
                        print("評核主管,無此工號...", director_id_A)

                # 找直接主管
                director_id_B = search_supervisor(work_code_hr, department_id, '直屬主管')
                if director_id_B == "" or director_id_B == None:
                    director_name_B = ""
                else:
                    try:
                        director_name_B = EmployeeInfoEasy.objects.get(work_code=director_id_B).chi_name
                        # print("直屬主管:",director_id_B, director_name_B)
                    except:
                        print("直屬主管,無此工號...", director_id_B)
        else:
            director_id_A = director_id
            director_name_A = director_name
            director_id_B = director_id_A
            director_name_B = director_name_A


        arrival_date = row.get("Date", None)
        trans_date = row.get("TransformDate", None)
        resign_date = row.get("LastWorkDate",None)
        hr_tuple = (                    # field(hr-->pms)
            work_code_hr,  # -->work_code
            row.get("CnName"),  # -->chi_name
            row.get("Corporation"),  # -->factory_id
            row.get("CountryId"),  # -->nat_id
            row.get("EmployeeState"),  # -->service_status_id
            row.get("Factory"),  # -->factory_area_id
            "HR無此資料",
            "HR無此資料",
            # HR的一級部門,部門別...對調
            direct_department_id,  # -->dept_flevel_id
            direct_department,
            department_id,  #  -->dept_id
            department,
            row.get("DepartmentFullName",""),  # -->dept_desc_id
            #
            director_id_A,  # recursive程式找出來的-->Director
            director_name_A,
            director_id_B,  # recursive程式找出來的-->Director
            director_name_B,
            pos_desc,  # (要把等次拿掉)-->pos_id
            row.get("ZhiJian"),  # -->labor_type_id
            datetime.datetime.date(arrival_date) if arrival_date else None,  # -->arrival_date
            None if row.get("LastWorkDate","") == datetime.datetime(9999,12,31) else datetime.datetime.date(resign_date),  # -->resign_date
            datetime.datetime.date(trans_date) if trans_date else None,  # (轉正日期)-->trans_date(異動日期)
            "轉正",
            row.get("PoliticalIdentity"),  # -->bonus_type_id
            row.get("DecisionLevel"),  # -->bonus_factor_id
            row.get("Grade"),  # -->ank_id
            "HR無此資料",
        )
        hr_all_data.append(hr_tuple)
        # hr : View_Employee_Short-----------------export to python----------------------------------------------------------Ending

        # 比對 hr -->pms-----------------------------------------------------------------------------------------------------Begin
        sheetHR.append(hr_tuple)
        find_result=list_filter_dict(pms_data,row.get("Code"))       #row[2] : work_code_id 搜尋PMS的工號
        hr_none = [6,7,8,9,21,22,26]      #email,自評,部門編號,一級部門,異動日,異動類別,BSC/KPI

        if find_result:
            #異動
            diff_count = 0
            for k in range(0,title_len):
                if k not in hr_none:
                    if hr_tuple[k]!=find_result[k]:
                        if diff_count==0:
                            change_in_row += 1
                            sheetC.append(find_result+tuple(("＊"))+hr_tuple)
                        diff_count += 1
                        sheetC.cell(row=change_in_row, column=k+1).fill = fille_diff
                        sheetC.cell(row=change_in_row, column=k+1).font = font_with_fill_diff
                        sheetC.cell(row=change_in_row, column=title_len+k+2).fill = fille_diff
                        sheetC.cell(row=change_in_row, column=title_len+k+2).font = font_with_fill_diff
        else:
            #新增
            if row.get("LastWorkDate","") == datetime.datetime(9999,12,31):
                # add_data.append(hr_tuple)
                sheetA.append(hr_tuple)
        # 比對 hr -->pms-----------------------------------------------------------------------------------------------------Ending
    print("=" * 150)
    print("\n" * 2)
    # conn.close()
    close_old_connections()


    sheetC_rows = len ( list(sheetC.rows) )
    for i in range(1, sheetC_rows):
        title2_len = title_len*2+2
        for j in range(1 , title2_len):
            if sheetC.cell(row=i, column=j).value=="＊":
                sheetC.cell(row=i, column=j).fill = fille
                sheetC.cell(row=i, column=j).font = font_with_fill


    #將定義好的的sheet, 寫入檔案
    current_tz = timezone.get_current_timezone()
    now = timezone.now().astimezone(current_tz).strftime('%Y%m%d_%H%M%S')
    filename = 'HR2PMS_COMPARE_' + now + '.xlsx'
    output_file = '%s/reports/xlsx/tt002/export/%s' % (settings.MEDIA_ROOT, filename)
    wb.save(output_file)
    openFile = '/media-files/reports/xlsx/tt002/export/' + filename

    message = ""
    return JsonResponse({"message":message,"success":True,"filename":filename,"openFile":openFile})


def get_all_factory(request):
    fieldList = [ 'id',
                  'name',
                 ]
    results = Factory.objects.all().values_list(*fieldList)
    dataList = []
    for dataTuple in results:
        dataList.append({
            'value': dataTuple[0],
            'text': str(dataTuple[0])+" "+dataTuple[1],
        })
    return JsonResponse(dataList,safe=False)